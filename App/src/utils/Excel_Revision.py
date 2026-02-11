# Principales/Excel_Revision.py
# -*- coding: utf-8 -*-
"""
==============================================================================
                      EXCEL_REVISION.PY - NOZHGESS v1.2
==============================================================================
Módulo de generación de Excel con estilos profesionales V2 (Mega Plan).

Funcionalidades:
- Exportación de resultados por misión en hojas separadas
- Estilos automáticos de headers por tipo de columna
- Colores Estrictos V2 (Azul, Verde, Rosa, Mostaza, Naranjo)
- Detección de Alertas Etarias

Autor: Sistema Nozhgess
==============================================================================
"""
from __future__ import annotations
import os
import re
from datetime import datetime
from typing import Dict, List, Any, Optional

import pandas as pd

from src.utils.Terminal import log_info, log_ok, log_error

# =============================================================================
#                      IMPORTAR ESTILOS DE OPENPYXL
# =============================================================================
try:
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    STYLES_AVAILABLE = True
except ImportError:
    PatternFill = Font = Alignment = Border = Side = None
    get_column_letter = None
    STYLES_AVAILABLE = False

# =============================================================================
#                      PALETA DE COLORES V2 (MEGA PLAN)
# =============================================================================

# GLOBAL FLAG
es_carga_masiva = False

COLORS = {
    # Grupo 1 (Headers Base): Azul Oscuro
    "grupo_azul_oscuro": {"fill": "4472C4", "font": "FFFFFF", "bold": True},
    
    # Grupo 2 (Familia/Esp/Fecha/Rut/Edad/Estado/Tipo): Azul Intenso
    "grupo_azul_familia": {"fill": "002060", "font": "FFFFFF", "bold": True},
    
    # Grupo 3 (Caso/Estado/SIC): Verde con Texto Blanco Negrita
    "grupo_verde_bold": {"fill": "00B050", "font": "FFFFFF", "bold": True},
    
    # Grupo 4 (Lógica Apto): Rosado Oscuro (para texto blanco)
    "grupo_apto": {"fill": "C71585", "font": "FFFFFF", "bold": True},
    
    # Grupo 5 (Tiempo - Mensual/CodAño): Café
    "grupo_cafe": {"fill": "806000", "font": "FFFFFF", "bold": True},
    
    # Grupo 6 (Críticos/Habilitantes): Rojo
    "grupo_rojo": {"fill": "FF0000", "font": "FFFFFF", "bold": True},
    
    # Grupo 7 (OA/Objetivos): Azul Eléctrico
    "grupo_azul_oa": {"fill": "0101FF", "font": "FFFFFF", "bold": True},
    
    # Grupo 8 (APS/Contra): Morado (Solicitud Usuario para CONTRA)
    "grupo_morado": {"fill": "7030A0", "font": "FFFFFF", "bold": True},
    
    # Grupo Naranjo (APS solo o Excluyentes si aplica)
    "grupo_naranjo": {"fill": "F26D00", "font": "FFFFFF", "bold": True},
    
    # Carga Masiva
    "grupo_cyan": {"fill": "00FFFF", "font": "000000", "bold": True}
}

# =============================================================================
#                           LOGICA DE ESTILOS
# =============================================================================

def _get_header_style(column_name: str) -> dict:
    """
    Determina el estilo de header según el nombre de la columna.
    """
    name = (column_name or "").lower().strip()
    
    # 1. CASO EN CONTRA (Prioridad Alta) -> Morado
    if "contra" in name:
        return COLORS["grupo_morado"]

    # 2. FRECUENCIAS (MENSUAL / CODIGO AÑO / PERIODICIDAD) -> Café
    if name in ["mensual", "periodicidad"] or "año" in name or "anio" in name or name.startswith("freq") or name.startswith("period"):
        return COLORS["grupo_cafe"]

    # 3. FECHA / RUT / EDAD / FAMILIA / ESPECIALIDAD / ESTADO / TIPO -> Azul Intenso (grupo_azul_familia)
    if name in ["fecha", "rut", "edad", "familia", "especialidad", "estado", "tipo"]:
        return COLORS["grupo_azul_familia"]

    # 4. GRUPO VERDE (Texto blanco negrita)
    # Fallecido, Caso, Apertura, SIC
    if name in ["fallecido", "caso", "apertura"] or "sic" in name:
        return COLORS["grupo_verde_bold"]

    # 5. APTOS (Texto blanco negrita) -> Rosado Oscuro
    if "apto" in name:
        return COLORS["grupo_apto"]

    # 6. OBSERVACIÓN -> Texto Blanco Negrita
    if "observ" in name:
        return COLORS["grupo_azul_oscuro"]

    # 7. HABILITANTES / EXCLUYENTES
    if "hab" in name and "excl" not in name:
         return COLORS["grupo_rojo"]
    if "excl" in name or "oa" in name or name.startswith("obj") or "objetivo" in name:
         return COLORS["grupo_azul_oa"]

    # 8. APS (Si no es contra)
    if "aps" in name:
        return COLORS["grupo_naranjo"]
    
    # Default: usar grupo azul oscuro
    return COLORS["grupo_azul_oscuro"]

def _apply_header_filter(ws) -> None:
    """Aplica filtro solo al rango utilizado (encabezados visibles)."""
    try:
        if not get_column_letter:
            return
        max_row = ws.max_row or 1
        max_col = ws.max_column or 1
        if max_row < 1 or max_col < 1:
            return
        last_col = get_column_letter(max_col)
        # Rango utilizado: encabezados + data (sin abarcar fila completa vacía)
        ws.auto_filter.ref = f"A1:{last_col}{max_row}"
    except Exception:
        pass


def _aplicar_estilos(ws, rows_metadata: List[Dict] = None) -> None:
    """
    Aplica estilos profesionales a una hoja de Excel.
    """
    if not STYLES_AVAILABLE:
        return
    
    # Fix: Define es_carga_masiva default
    es_carga_masiva = False
    
    # Borde sutil
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    # Pre-instanciar estilos para rendimiento (Caching)
    # Evita crear millones de objetos PatternFill/Font en bucles largos
    style_cache = {}
    
    # Crear cache de objetos para cada grupo de color definido
    for group_name, props in COLORS.items():
        fill_obj = PatternFill(start_color=props["fill"], end_color=props["fill"], fill_type="solid")
        font_obj = Font(color=props["font"], bold=props.get("bold", True), size=9) # Default size 9 for body
        # Cachear tupla (fill, font)
        style_cache[group_name] = (fill_obj, font_obj)
        
    # Cache especial para headers (size 10)
    header_style_cache = {}
    for group_name, props in COLORS.items():
         fill_obj = PatternFill(start_color=props["fill"], end_color=props["fill"], fill_type="solid")
         font_obj = Font(color=props["font"], bold=props.get("bold", True), size=10)
         header_style_cache[group_name] = (fill_obj, font_obj)
         
    # Cache para Alertas de Edad
    # Definir Fills para Semáforo (Pastel standard para legibilidad)
    age_green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    age_yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    age_red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    age_font_green = Font(color="006100", bold=True, size=9)
    age_font_yellow = Font(color="9C5700", bold=True, size=9)
    age_font_red = Font(color="9C0006", bold=True, size=9)
    
    # Cache para Futuros (Explicit Black)
    future_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    future_font = Font(color="000000", size=9)
    
    # Default Font (Explicit Black)
    default_font = Font(color="000000", size=9)
    # Alineaciones reutilizables (mejor rendimiento)
    align_header = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_center = Alignment(horizontal='center', vertical='center')

    # Map column names to indices
    header_map = {}
    for cell in ws[1]:
        header_map[cell.column] = str(cell.value or "").lower().strip()

    # Estilizar headers (fila 1)
    for cell in ws[1]:
        # Determinar grupo
        if es_carga_masiva:
            group_key = "grupo_cyan"
        else:
             # _get_header_style devuelve el dict de propiedades directamente, no la key.
             # Necesitamos refactorizar levemente o buscar la key inversa.
             # Mejor optimización: _get_header_style ahora devuelve solo la KEY del grupo.
             # Pero para no romper compatibilidad, buscaremos el match o calcularemos al vuelo si no hay cache.
             
             # Fallback manual por ahora: Recalcular style dict
             style_props = _get_header_style(str(cell.value or ""))
             
             # Crear objetos al vuelo solo para headers (son pocos, ~50 max)
             # No vale la pena complicar el cache inverso.
             cell.fill = PatternFill(start_color=style_props["fill"], end_color=style_props["fill"], fill_type="solid")
             is_bold = style_props.get("bold", True)
             cell.font = Font(color=style_props["font"], bold=is_bold, size=10)
        
        # Si fuera carga masiva, usamos el cache directo
        if es_carga_masiva:
             # Carga masiva usa grupo_cyan directo
             hf, ht = header_style_cache["grupo_cyan"]
             cell.fill = hf
             cell.font = ht

        cell.alignment = align_header
        cell.border = thin_border
    
    # Estilizar celdas de datos - Centrado horizontal y vertical
    # Aquí es donde importa el rendimiento (10k+ celdas)
    
    # Pre-calcular el grupo de estilo para cada COLUMNA (índice)
    # Lista de tuplas (FillObj, FontObj) por índice de columna (1-based)
    col_styles_map = {} # {col_idx: (fill, font)}
    
    for col_idx, col_name in header_map.items(): # header_map usa cell.column (int)
         # Determinar el estilo base de esta columna
         style_props = _get_header_style(col_name) # Devuelve dict
         
         # Buscar si este dict coincide con algún grupo cacheado para reutilizar objetos
         # Heurística: chequear por 'fill' color
         found_objs = None
         color_fill = style_props["fill"]
         
         # Buscar en style_cache
         for g_key, (c_fill, c_font) in style_cache.items():
              if COLORS[g_key]["fill"] == color_fill:
                  found_objs = (c_fill, c_font)
                  break
         
         if found_objs:
             col_styles_map[col_idx] = found_objs
         else:
             # Si no está en cache standard, crear uno nuevo
             f = PatternFill(start_color=color_fill, end_color=color_fill, fill_type="solid")
             t = Font(color=style_props["font"], bold=style_props.get("bold", True), size=9)
             col_styles_map[col_idx] = (f, t)


    for i, row in enumerate(ws.iter_rows(min_row=2), start=0):
        age_status = None
        has_age_alert = False
        status_red = False
        codxanio_source = None
        
        if rows_metadata and i < len(rows_metadata):
            # Preferir el status nuevo, fallback al booleano antiguo
            age_status = rows_metadata[i].get("_age_validation_status", None)
            has_age_alert = rows_metadata[i].get("_age_alert", False)
            status_red = rows_metadata[i].get("_status_red", False)
            codxanio_source = rows_metadata[i].get("_codxanio_source", None)
            
        for cell in row:
            val = cell.value
            col_name = header_map.get(cell.column, "")
            col_idx = cell.column
            
            # Recuperar estilo base (cacheado) de la columna
            base_fill, base_font = col_styles_map.get(col_idx, (None, default_font))
            
            # Aplicar base
            # FIX: Usuario pidió celdas de datos SIN color (transparente), salvo excepciones.
            # if base_fill: cell.fill = base_fill  <-- REMOVED
            # Correcto: Usar font negro por defecto para cuerpo, ignorar el del header
            # cell.font = base_font  <-- ERROR: base_font tiene color blanco (header)
            cell.font = default_font # Default is generic black
            
            # --- SOBRE-ESCRIBIR CON CONDICIONES DE CELDA ---
            
            # 1. Alerta de Edad (Prioridad Alta)
            if col_name == "edad":
                if age_status == "green":
                    cell.fill = age_green_fill
                    cell.font = age_font_green
                elif age_status == "yellow":
                    cell.fill = age_yellow_fill
                    cell.font = age_font_yellow
                elif age_status == "red" or has_age_alert: # Fallback a alerta antigua si es red
                    cell.fill = age_red_fill
                    cell.font = age_font_red
                else: 
                     # Si no es alerta y es la columna edad -> VERDE (Cumple)
                     # Solo si tiene valor
                     if val is not None and val != "":
                         cell.fill = age_green_fill
                         cell.font = age_font_green
            
            # 2. Detección de Prestaciones Futuras (! )
            elif isinstance(val, str) and val.startswith("! "):
                # Aplicar color naranja "Advertencia/Futuro"
                cell.fill = future_fill
                # Quitar marcador para que quede limpio
                cell.value = val[2:]
                cell.font = future_font
            
            # 5. Estilizado de CodxAño (Prioridad IPD/APS)
            elif "codxaño" in col_name.lower() and codxanio_source:
                if codxanio_source == "ipd":
                    # Verde (IPD)
                    cell.fill = age_green_fill
                    cell.font = age_font_green
                elif codxanio_source == "aps":
                    # Naranja (APS)
                    cell.fill = future_fill
                    cell.font = future_font

            # 3. Optimización: Si no hubo override, ya tiene el base_font/fill puestos.
                
            # 3. Folios VIH Usados (Sombreado Verde)
            elif "folio oa" in col_name and rows_metadata and i < len(rows_metadata):
                folios_usados = rows_metadata[i].get("_folios_usados", [])
                # Limpiar valor para comparaciÃ³n
                val_clean = str(val).strip().lower().replace("oa", "").strip()
                if val_clean and val_clean in [str(f).strip().lower().replace("oa", "").strip() for f in folios_usados]:
                    cell.fill = age_green_fill
                    cell.font = age_font_green
            
            # 4. Celda ROJA crítica (Sin Mini-Tabla)
            elif status_red and "observ" in col_name:
                rf, rt = style_cache.get("grupo_rojo", (None, None))
                if rf:
                    cell.fill = rf
                    cell.font = rt
            
            cell.alignment = align_center
            cell.border = thin_border
    
    # Filtro en encabezados (rango utilizado)
    _apply_header_filter(ws)

def _escribir_y_estilizar(writer, resultados_por_mision, mission_list: List[Dict] = None):
    """Escribe los dataframes y llama a estilizar. Retorna columnas encontradas (ordenadas)."""
    columnas_encontradas = []
    columnas_set = set()
    for m_name, items in resultados_por_mision.items():
        if not items: continue # Skip empty
        
        # Determine Sheet Name
        sheet_label = str(m_name)
        # If m_name is an index (legacy), try to get name from mission_list
        if str(m_name).isdigit() and mission_list and int(m_name) < len(mission_list):
            sheet_label = mission_list[int(m_name)].get("nombre", f"Mision_{m_name}")
        
        # Extract metadata (age alert + columns order)
        metadata_list = []
        clean_items = []
        cols_order = None
        for it in items:
            meta = {
                "_age_alert": it.get("_age_alert", False),
                "_age_validation_status": it.get("_age_validation_status", None),
                "_folios_usados": it.get("_folios_usados", [])
            }
            if cols_order is None and "_cols_order" in it:
                cols_order = it["_cols_order"]
            # Create clean copy removing internal keys
            clean_it = {k: v for k, v in it.items() if not k.startswith("_")}
            clean_items.append(clean_it)
            metadata_list.append(meta)
            
        df_clean = pd.DataFrame(clean_items)
        if cols_order:
            # Reindex to expected order, keeping any extra columns at the end (except internal/unwanted)
            extra_cols = [c for c in df_clean.columns if c not in cols_order and c != "Nombre"]
            df_clean = df_clean.reindex(columns=cols_order + extra_cols)
        # Eliminar columnas duplicadas (mantener la primera aparición)
        if df_clean.columns.duplicated().any():
            df_clean = df_clean.loc[:, ~df_clean.columns.duplicated()]
        
        # Sheet name cleaning
        # Limitar a 31 chars y quitar caracteres inválidos
        safe_name = str(sheet_label).replace(":", "").replace("/", "").replace("\\", "")[:30]
        
        df_clean.to_excel(writer, sheet_name=safe_name, index=False)
        
        # Metadata validation
        if hasattr(writer, 'book') and hasattr(writer.book, 'properties'):
            writer.book.properties.author = "NZLP-7733-CL"
            writer.book.properties.comments = "Build (PROD): 2026-NZT-M01"
        
        ws = writer.sheets[safe_name]
        _aplicar_estilos(ws, metadata_list)
        
    # Return columns found from the last df processed or accumulate them?
    # Actually df_clean might not be defined if no items. Adding protection.
    try: 
        for col in df_clean.columns:
            if col not in columnas_set:
                columnas_set.add(col)
                columnas_encontradas.append(col)
    except Exception: pass
    return columnas_encontradas


# =============================================================================
#                    HOJA DICCIONARIO DE COLUMNAS
# =============================================================================
def _describe_column(col: str) -> dict:
    """
    Retorna metadatos detallados para una columna.
    Devuelve dict: {
        "Categoria": str, 
        "Descripcion": str, 
        "Fuente": str, 
        "Nota": str
    }
    """
    name = (col or "").lower().strip()
    
    # -------------------------------------------------------------------------
    # 1. HEURÍSTICAS DINÁMICAS
    # -------------------------------------------------------------------------
    if name.startswith("código año") or name.startswith("cod año"):
        return {
            "Categoria": "Lógica de Negocio",
            "Descripcion": "Código seleccionado automáticamente según la antigüedad del tratamiento.",
            "Fuente": "Cálculo Interno",
            "Nota": "Se calcula: [Año Objetivo] - [Año IPD]. Ejemplo: IPD 2021 vs Obj 2024 = 3 años de antigüedad."
        }
    if name.startswith("freq res") or name.startswith("freq_res"):
        return {
            "Categoria": "Lógica de Negocio",
            "Descripcion": "Resultado de la validación de frecuencia específica.",
            "Fuente": "Cálculo Interno",
            "Nota": "Muestra: Cantidad Encontrada / Cantidad Requerida (Tipo). Ej: '1/2 Mes'."
        }
    if name.startswith("freq per") or name.startswith("freq_per"):
        return {
            "Categoria": "Configuración",
            "Descripcion": "Periodicidad de la regla de frecuencia.",
            "Fuente": "Misión",
            "Nota": "Etiqueta informativa (ej: 'Mensual', 'Por Ciclo')."
        }
    if name == "freq codxaño":
        return {
            "Categoria": "Lógica de Negocio",
            "Descripcion": "Frecuencia requerida según la antigüedad actual del paciente.",
            "Fuente": "Configuración (Código por Año)",
            "Nota": "Configuración específica para el año en curso del paciente."
        }
    if name.startswith("obj") or "objetivo" in name:
        return {
            "Categoria": "Gestión y Seguimiento",
            "Descripcion": "Fechas encontradas para el cumplimiento de Objetivos Sanitarios.",
            "Fuente": "Prestaciones (SIGGES)",
            "Nota": "Se buscan códigos específicos configurados en la misión. Muestra la fecha más reciente."
        }
    if name.startswith("c hab") or name.startswith("f hab") or name.startswith("hab vi"):
        return {
            "Categoria": "Criterios de Inclusión",
            "Descripcion": "Evaluación de Habilitantes (Diagnósticos o prestaciones previas requeridas).",
            "Fuente": "Prestaciones (SIGGES)",
            "Nota": "Si no se encuentran habilitantes vigentes, el paciente podría no calificar."
        }
    if name.startswith("c excl") or name.startswith("f excl"):
        return {
            "Categoria": "Criterios de Exclusión",
            "Descripcion": "Detección de Excluyentes (Códigos que invalidan la misión para este paciente).",
            "Fuente": "Prestaciones (SIGGES)",
            "Nota": "La presencia de estos códigos generalmente descarta el caso."
        }
    if "folio vih" in name:
         return {
            "Categoria": "Datos Clínicos Específicos",
            "Descripcion": "Folio asociado a prestaciones o diagnósticos de VIH (si aplica).",
            "Fuente": "SIGGES OA/Prestaciones",
            "Nota": "Rastreo específico para misiones que requieren trazabilidad de VIH."
         }

    # -------------------------------------------------------------------------
    # 2. DICCIONARIO ESTÁTICO
    # -------------------------------------------------------------------------
    mapping = {
        # --- IDENTIFICACIÓN ---
        "rut": {
            "Categoria": "Identificación Paciente",
            "Descripcion": "RUT del paciente normalizado.",
            "Fuente": "Excel de Entrada",
            "Nota": "Clave principal de búsqueda."
        },
        "nombre": {
            "Categoria": "Identificación Paciente",
            "Descripcion": "Nombre completo del paciente.",
            "Fuente": "Excel de Entrada",
            "Nota": ""
        },
        "edad": {
            "Categoria": "Identificación Paciente",
            "Descripcion": "Edad actual del paciente registrada en sistema.",
            "Fuente": "SIGGES (Datos Demográficos)",
            "Nota": "Puede activar alertas de colores si está fuera del rango configurado."
        },
        "fallecido": {
            "Categoria": "Identificación Paciente",
            "Descripcion": "Fecha de fallecimiento (si aplica) o 'No'.",
            "Fuente": "SIGGES (Historia)",
            "Nota": "Dato crítico para la gestión."
        },

        # --- DATOS DEL CASO ---
        "caso": {
            "Categoria": "Datos del Caso GES",
            "Descripcion": "Nombre del problema de salud GES activo.",
            "Fuente": "SIGGES (Mini-tabla)",
            "Nota": "Debe coincidir con las 'keywords' de la misión."
        },
        "estado": {
            "Categoria": "Datos del Caso GES",
            "Descripcion": "Estado administrativo del caso (ej. Vigente, Cerrado).",
            "Fuente": "SIGGES (Mini-tabla)",
            "Nota": "Si no hay caso, se marca como 'Sin Caso'."
        },
        "tipo": {
            "Categoria": "Configuración",
            "Descripcion": "Tipo de cobertura/flujo (valor fijo por defecto).",
            "Fuente": "Sistema",
            "Nota": "Por defecto: 'Auge'."
        },
        "apertura": {
            "Categoria": "Datos del Caso GES",
            "Descripcion": "Fecha de inicio del caso GES.",
            "Fuente": "SIGGES (Mini-tabla)",
            "Nota": "Determina la antigüedad administrativa."
        },


        # --- DATOS CLÍNICOS (IPD) ---
        "fecha ipd": {
            "Categoria": "Datos Clínicos (IPD)",
            "Descripcion": "Fecha del Informe de Proceso Diagnóstico.",
            "Fuente": "SIGGES (Pestaña IPD)",
            "Nota": "Base para calcular antigüedad clínica."
        },
        "estado ipd": {
            "Categoria": "Datos Clínicos (IPD)",
            "Descripcion": "Resultado del IPD (ej. 'Sí', 'No', 'En estudio').",
            "Fuente": "SIGGES (Pestaña IPD)",
            "Nota": "Debe ser 'Sí' para considerar al paciente Apto RE (generalmente)."
        },
        "diagnóstico ipd": {
            "Categoria": "Datos Clínicos (IPD)",
            "Descripcion": "Detalle del diagnóstico confirmado en IPD.",
            "Fuente": "SIGGES (Pestaña IPD)",
            "Nota": ""
        },

        # --- DATOS CLÍNICOS (OA) ---
        "fecha oa": {
            "Categoria": "Datos Clínicos (OA)",
            "Descripcion": "Fecha de emisión de la Orden de Atención.",
            "Fuente": "SIGGES (Pestaña OA)",
            "Nota": ""
        },
        "código oa": {
            "Categoria": "Datos Clínicos (OA)",
            "Descripcion": "Código de prestación asociado a la OA.",
            "Fuente": "SIGGES (Pestaña OA)",
            "Nota": ""
        },
        "folio oa": {
            "Categoria": "Datos Clínicos (OA)",
            "Descripcion": "Folio único de la Orden de Atención.",
            "Fuente": "SIGGES (Pestaña OA)",
            "Nota": "Usado para trazabilidad y cruce con prestaciones realizadas."
        },
        "derivado oa": {
            "Categoria": "Datos Clínicos (OA)",
            "Descripcion": "Estado de derivación/tratamiento OA.",
            "Fuente": "SIGGES (Pestaña OA)",
            "Nota": "Si indica 'Caso en Tratamiento', suma puntos para Apto RE."
        },
        "diagnóstico oa": {
            "Categoria": "Datos Clínicos (OA)",
            "Descripcion": "Diagnóstico asociado a la OA.",
            "Fuente": "SIGGES (Pestaña OA)",
            "Nota": ""
        },

        # --- DATOS CLÍNICOS (APS) ---
        "fecha aps": {
            "Categoria": "Datos Clínicos (APS)",
            "Descripcion": "Fecha de registro en Atención Primaria.",
            "Fuente": "SIGGES (Pestaña APS)",
            "Nota": ""
        },
        "estado aps": {
            "Categoria": "Datos Clínicos (APS)",
            "Descripcion": "Estado de confirmación APS (ej. 'Caso Confirmado').",
            "Fuente": "SIGGES (Pestaña APS)",
            "Nota": "Clave para determinar elegibilidad en ciertos cánceres."
        },

        # --- DATOS CLÍNICOS (SIC) ---
        "fecha sic": {
            "Categoria": "Datos Clínicos (SIC)",
            "Descripcion": "Fecha de Solicitud de Interconsulta.",
            "Fuente": "SIGGES (Pestaña SIC)",
            "Nota": ""
        },
        "derivado sic": {
            "Categoria": "Datos Clínicos (SIC)",
            "Descripcion": "Destino o motivo de la interconsulta.",
            "Fuente": "SIGGES (Pestaña SIC)",
            "Nota": ""
        },

        # --- LÓGICA DE NEGOCIO (CALCULADA) ---
        "apto elección": {
            "Categoria": "Lógica de Negocio",
            "Descripcion": "Evaluación compuesta de requisitos IPD/APS configurables.",
            "Fuente": "Algoritmo Interno",
            "Nota": "Muestra si cumple requisitos activados (ej: 'SI IPD | NO APS'). Útil para decisiones flexibles."
        },
        "apto se": {
            "Categoria": "Lógica de Negocio",
            "Descripcion": "Apto para Seguimiento (Histórico).",
            "Fuente": "Algoritmo Interno",
            "Nota": "Indica si el paciente ha tenido historial de seguimiento (OA/SIC) previamente."
        },
        "apto re": {
            "Categoria": "Lógica de Negocio",
            "Descripcion": "Apto para Resolución (Criterios Estrictos).",
            "Fuente": "Algoritmo Interno",
            "Nota": "Evalúa cumplimiento simultáneo: IPD + (Sí), OA + (En Tratamiento), APS + (Confirmado)."
        },
        "apto caso": {
            "Categoria": "Lógica de Negocio",
            "Descripcion": "Comparativa con casos 'En Contra'.",
            "Fuente": "Algoritmo Interno",
            "Nota": "Indica si existe un caso contradictorio más reciente que el actual (ej. un cáncer nuevo)."
        },
        "mensual": {
            "Categoria": "Lógica de Negocio",
            "Descripcion": "Verificación de cobertura periódica (Mensual/Anual).",
            "Fuente": "Cruce Prestaciones",
            "Nota": "Muestra si se encontró la prestación requerida en la frecuencia configurada (ej. 1/1 Mes)."
        },
        "periodicidad": {
            "Categoria": "Lógica de Negocio",
            "Descripcion": "Frecuencia teórica configurada.",
            "Fuente": "Configuración Misión",
            "Nota": "Informativo."
        },

        # --- AUDITORÍA Y TRAZABILIDAD ---
        "observación": {
            "Categoria": "Notas / Auditoría",
            "Descripcion": "Campo reservado para notas manuales o errores críticos de sistema.",
            "Fuente": "Sistema / Manual",
            "Nota": "Por defecto vacío, salvo error de conexión o 'Sin Caso'."
        },
        "observación folio": {
            "Categoria": "Auditoría Interna",
            "Descripcion": "Verificación cruzada de Folio OA.",
            "Fuente": "Sistema",
            "Nota": "Confirma si el folio de la OA fue utilizado en alguna prestación facturada."
        },
        "familia": {
             "Categoria": "Configuración",
             "Descripcion": "Familia GES asignada.",
             "Fuente": "Misión",
             "Nota": ""
        },
        "especialidad": {
             "Categoria": "Configuración",
             "Descripcion": "Especialidad asignada.",
             "Fuente": "Misión",
             "Nota": ""
        }
    }

    # Búsqueda en mapa estático
    if name in mapping:
        return mapping[name]
    
    # Búsqueda parcial para "En Contra"
    if "contra" in name:
        return {
            "Categoria": "Caso en Contra (Incompatibilidad)",
            "Descripcion": "Datos provenientes de un caso secundario detectado como conflicto.",
            "Fuente": "SIGGES (Caso Secundario)",
            "Nota": "Usado para descartar pacientes que ya tienen otro diagnóstico prevalente."
        }

    # Fallback Genérico
    return {
        "Categoria": "Otros Datos",
        "Descripcion": "Columna generada dinámicamente o dato adicional.",
        "Fuente": "Procesamiento General",
        "Nota": ""
    }

DICT_FIELDS = [
    "Categoría",
    "Columna Excel",
    "Descripción (Qué es)",
    "Función / Para qué sirve",
    "Cómo se obtiene",
    "Cuándo se calcula",
    "Dónde se usa",
    "Por qué es importante",
    "Tipo/Formato",
    "Fuente de datos",
    "Validaciones/Reglas",
    "Ejemplo",
    "Notas",
]


def _expand_column_info(col: str) -> dict:
    base = _describe_column(col)
    name = (col or "").lower().strip()
    categoria = base.get("Categoria", "Otros Datos")
    descripcion = base.get("Descripcion", "")
    fuente = base.get("Fuente", "")
    nota = base.get("Nota", "")

    cuando_map = {
        "Identificación Paciente": "Al cargar la nómina y consultar datos del paciente.",
        "Datos del Caso GES": "Al leer la mini-tabla del caso en SIGGES.",
        "Datos Clínicos (IPD)": "Durante la extracción de IPD.",
        "Datos Clínicos (OA)": "Durante la extracción de OA.",
        "Datos Clínicos (APS)": "Durante la extracción de APS.",
        "Datos Clínicos (SIC)": "Durante la extracción de SIC.",
        "Gestión y Seguimiento": "Durante la revisión de objetivos y seguimiento.",
        "Criterios de Inclusión": "Durante la validación de habilitantes.",
        "Criterios de Exclusión": "Durante la validación de excluyentes.",
        "Caso en Contra (Incompatibilidad)": "Durante el análisis de casos en contra.",
        "Auditoría Interna": "Cuando ocurre un evento relevante en el proceso.",
        "Lógica de Negocio": "Durante la evaluación de reglas de la misión.",
        "Configuración": "Al cargar la misión.",
    }
    donde_map = {
        "Identificación Paciente": "En búsquedas, cruces y reportes.",
        "Datos del Caso GES": "En validaciones y decisiones de gestión.",
        "Datos Clínicos (IPD)": "En cálculos clínicos y reglas de apto.",
        "Datos Clínicos (OA)": "En trazabilidad y reglas clínicas.",
        "Datos Clínicos (APS)": "En reglas de confirmación APS.",
        "Datos Clínicos (SIC)": "En trazabilidad de interconsultas.",
        "Gestión y Seguimiento": "En reportes y auditorías de cumplimiento.",
        "Criterios de Inclusión": "En reglas de elegibilidad.",
        "Criterios de Exclusión": "En reglas de descarte.",
        "Caso en Contra (Incompatibilidad)": "En validación final y descartes.",
        "Auditoría Interna": "En revisión manual y depuración.",
        "Lógica de Negocio": "En resultados de aptitud y resumen.",
        "Configuración": "En segmentación y contexto del reporte.",
    }
    por_que_map = {
        "Identificación Paciente": "Es clave para identificar al paciente y consultar su historial.",
        "Datos del Caso GES": "Define el contexto administrativo y clínico principal.",
        "Datos Clínicos (IPD)": "Permite evaluar diagnóstico y antigüedad.",
        "Datos Clínicos (OA)": "Aporta trazabilidad de atención.",
        "Datos Clínicos (APS)": "Define confirmaciones en atención primaria.",
        "Datos Clínicos (SIC)": "Aporta evidencia de derivación.",
        "Gestión y Seguimiento": "Mide cumplimiento de objetivos sanitarios.",
        "Criterios de Inclusión": "Determina si el caso puede ser considerado.",
        "Criterios de Exclusión": "Evita seleccionar casos inválidos.",
        "Caso en Contra (Incompatibilidad)": "Detecta diagnósticos conflictivos.",
        "Auditoría Interna": "Explica errores o hallazgos críticos.",
        "Lógica de Negocio": "Resume decisiones calculadas por el sistema.",
        "Configuración": "Aporta contexto de la misión.",
    }

    def _infer_como() -> str:
        if "SIGGES" in fuente:
            return "Se consulta en SIGGES según la pestaña/fuente indicada."
        if "Misión" in fuente:
            return "Se lee desde la configuración de la misión."
        if "Sistema" in fuente or "Cálculo" in fuente or "Algoritmo" in fuente:
            return "Se calcula internamente durante la revisión."
        return "Se obtiene durante el procesamiento."

    def _infer_tipo() -> str:
        if "fecha" in name:
            return "Fecha"
        if name == "edad":
            return "Numérico (años)"
        if "rut" in name:
            return "Texto (RUT)"
        if "apto" in name or "fallecido" in name or "¿cerrado?" in name:
            return "Sí/No"
        if "folio" in name or "código" in name or "codigo" in name:
            return "Texto/Numérico"
        if "mensual" in name or "freq" in name or "periodicidad" in name:
            return "Texto (regla)"
        return "Texto"

    funcion = descripcion or "Dato utilizado en el proceso de revisión."
    if nota:
        funcion = f"{funcion} {nota}"
    como = _infer_como()
    cuando = cuando_map.get(categoria, "Durante la revisión de la misión.")
    donde = donde_map.get(categoria, "En la hoja de resultados.")
    por_que = por_que_map.get(categoria, "Aporta contexto a la decisión.")
    tipo = _infer_tipo()
    validaciones = nota or "Depende de la misión."

    ejemplo = ""
    if "fecha" in name:
        ejemplo = "15-01-2024"
    elif name == "rut":
        ejemplo = "12345678-9"
    elif name == "edad":
        ejemplo = "55"
    elif "apto" in name:
        ejemplo = "Sí"
    elif "mensual" in name or "freq" in name:
        ejemplo = "1/2 Mes"
    elif "folio" in name:
        ejemplo = "OA-2024-001234"
    elif "código" in name or "codigo" in name:
        ejemplo = "A12345"
    elif name == "tipo":
        ejemplo = "Auge"
    elif name == "estado":
        ejemplo = "Vigente / Sin Caso"

    # Overrides más específicos
    if name == "tipo":
        tipo = "Texto fijo"
        validaciones = "Valor por defecto: Auge."
    if name == "estado":
        validaciones = "Si no hay caso, se marca 'Sin Caso'."
    if name == "observación":
        por_que = "Explica errores, hallazgos o datos faltantes."
    if "apto" in name:
        por_que = "Resume la decisión de aptitud y permite priorizar casos."
    if "freq" in name or "mensual" in name or "periodicidad" in name:
        por_que = "Permite verificar cumplimiento de frecuencia/periodicidad."
    if name.startswith("c hab") or name.startswith("f hab"):
        por_que = "Define si el paciente cumple habilitantes requeridos."
    if name.startswith("c excl") or name.startswith("f excl"):
        por_que = "Si hay excluyentes, el caso se descarta."

    return {
        "Categoría": categoria,
        "Descripción (Qué es)": descripcion,
        "Función / Para qué sirve": funcion,
        "Cómo se obtiene": como,
        "Cuándo se calcula": cuando,
        "Dónde se usa": donde,
        "Por qué es importante": por_que,
        "Tipo/Formato": tipo,
        "Fuente de datos": fuente,
        "Validaciones/Reglas": validaciones,
        "Ejemplo": ejemplo,
        "Notas": nota,
    }


def _escribir_diccionario(writer, columnas: List[str]) -> None:
    """
    Genera la hoja 'Diccionario' con formato Premium.
    """
    if not columnas:
        return

    # 1. Preparar datos
    data = []
    seen = set()
    
    # Ordenar columnas: Primero por categoría lógica (hardcoded sort), luego alfabético
    # Definir prioridad de categorías
    cat_priority = {
        "Identificación Paciente": 1,
        "Lógica de Negocio": 2,
        "Datos del Caso GES": 3,
        "Datos Clínicos (IPD)": 4,
        "Datos Clínicos (OA)": 5,
        "Datos Clínicos (APS)": 6,
        "Datos Clínicos (SIC)": 7,
        "Gestión y Seguimiento": 8,
        "Criterios de Inclusión": 9,
        "Criterios de Exclusión": 10,
        "Caso en Contra (Incompatibilidad)": 11,
        "Auditoría Interna": 99,
        "Otros Datos": 100
    }

    for col in columnas:
        if col in seen: continue
        seen.add(col)
        
        info = _expand_column_info(col)
        # Asignar prioridad de orden
        prio = cat_priority.get(info["Categoría"], 50)
        
        row = {"_sort": prio, "Categoría": info["Categoría"], "Columna Excel": col}
        for field in DICT_FIELDS:
            if field in ("Categoría", "Columna Excel"):
                continue
            row[field] = info.get(field, "")
        data.append(row)

    # Ordenar data
    data.sort(key=lambda x: (x["_sort"], x["Categoría"], x["Columna Excel"]))
    
    # Eliminar llave temporal corrección
    clean_data = [{k: v for k, v in d.items() if k != "_sort"} for d in data]

    # 2. Crear DataFrame
    df = pd.DataFrame(clean_data, columns=DICT_FIELDS)
    
    # 3. Escribir a Excel
    sheet_name = "Diccionario"
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # 4. Estilizar (Premium)
    ws = writer.sheets[sheet_name]
    
    if STYLES_AVAILABLE:
        # Definir estilos
        header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid") # Azul oscuro
        header_font = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
        
        cat_font = Font(bold=True, color="000000", size=10)
        text_font = Font(size=10, name="Calibri")
        
        border = Border(
            left=Side(style='thin', color='BFBFBF'),
            right=Side(style='thin', color='BFBFBF'),
            top=Side(style='thin', color='BFBFBF'),
            bottom=Side(style='thin', color='BFBFBF')
        )
        
        # Estilizar Headers
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
            
        # Estilizar Cuerpo
        for row in ws.iter_rows(min_row=2):
            for i, cell in enumerate(row):
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                
                # Columna A (Categoría) en Negrita
                if i == 0: 
                    cell.font = cat_font
                    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                else:
                    cell.font = text_font
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Dimensiones (diccionario extendido)
        widths = {
            "A": 22,  # Categoría
            "B": 22,  # Columna Excel
            "C": 40,  # Descripción
            "D": 40,  # Función
            "E": 36,  # Cómo se obtiene
            "F": 28,  # Cuándo
            "G": 28,  # Dónde
            "H": 32,  # Por qué
            "I": 18,  # Tipo
            "J": 22,  # Fuente
            "K": 32,  # Validaciones
            "L": 20,  # Ejemplo
            "M": 32,  # Notas
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        
        # Altura de filas automática (aproximada, openpyxl no hace autofit height real bien, pero wrap_text ayuda)

    # Filtro en encabezados
    _apply_header_filter(ws)


# =============================================================================
#                     FUNCIN PRINCIPAL DE EXPORTACIN
# =============================================================================


def _crear_hoja_carga_masiva(writer) -> None:
    """
    Crea la hoja 'Carga Masiva' solo con los encabezados solicitados.
    """
    headers = ["Fecha", "Rut", "DV", "Prestaciones", "Tipo", "PS-Fam", "Especialidad"]
    df = pd.DataFrame(columns=headers)
    
    sheet_name = "Carga Masiva"
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # Estilizar headers (Cyan)
    if STYLES_AVAILABLE:
        ws = writer.sheets[sheet_name]
        header_fill = PatternFill(start_color="00FFFF", end_color="00FFFF", fill_type="solid") # Cyan
        header_font = Font(color="000000", bold=True, size=11, name="Calibri")
        
        border = Border(
            left=Side(style='thin', color='BFBFBF'),
            right=Side(style='thin', color='BFBFBF'),
            top=Side(style='thin', color='BFBFBF'),
            bottom=Side(style='thin', color='BFBFBF')
        )
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
            
        # Ajustar anchos
        for col_idx, _ in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 20

    # Filtro en encabezados
    ws = writer.sheets[sheet_name]
    _apply_header_filter(ws)


def generar_excel_revision(
    resultados_por_mision: Dict[int, List[Dict[str, Any]]],
    MISSIONS: List[Dict[str, Any]],
    NOMBRE_DE_LA_MISION: str,
    RUTA_CARPETA_SALIDA: str
) -> Optional[str]:
    """
    Genera el Excel final de revisión con estilos profesionales.
    """
    # Crear carpeta si no existe
    os.makedirs(RUTA_CARPETA_SALIDA, exist_ok=True)

    # Generar nombre de archivo con timestamp
    stamp = datetime.now().strftime("%d.%m.%Y_%H.%M")
    base_nombre = re.sub(r"[^\wáéíóúÁÉÍÓÚñÑ]+", "_", NOMBRE_DE_LA_MISION.strip())
    filename = f"Rev_{base_nombre}_{stamp}.xlsx"
    ruta_salida = os.path.join(RUTA_CARPETA_SALIDA, filename)

    try:
        # INTENTO 1: Ruta Original
        with pd.ExcelWriter(ruta_salida, engine="openpyxl") as writer:
            cols = _escribir_y_estilizar(writer, resultados_por_mision, MISSIONS)
            _crear_hoja_carga_masiva(writer) 
            _escribir_diccionario(writer, cols)
            
        log_ok(f" Excel guardado: {filename}")
        return ruta_salida
        
    except Exception as e:
        log_error(f" Error guardando en ruta original: {e}")
        
        # INTENTO 2: Fallback (Carpeta Backups en Proyecto)
        try:
            root_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))) # App/src/utils -> App/src -> App -> Root
            fecha_hoy = datetime.now().strftime("%Y-%m-%d")
            backup_dir = os.path.join(root_dir, "Backups", fecha_hoy)
            os.makedirs(backup_dir, exist_ok=True)
            
            ruta_backup = os.path.join(backup_dir, filename)
            
            log_info(f" Intentando guardar en respaldo: {ruta_backup}")
            
            with pd.ExcelWriter(ruta_backup, engine="openpyxl") as writer:
                cols = _escribir_y_estilizar(writer, resultados_por_mision, MISSIONS)
                _crear_hoja_carga_masiva(writer) 
                _escribir_diccionario(writer, cols)
                
            log_ok(f" RESCATADO: Excel guardado en respaldo: {ruta_backup}")
            return ruta_backup
            
        except Exception as e2:
            log_error(f" CRÍTICO: Falló también el respaldo: {e2}")
            return None
