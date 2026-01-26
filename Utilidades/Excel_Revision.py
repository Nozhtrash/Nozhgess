# Principales/Excel_Revision.py
# -*- coding: utf-8 -*-
"""
==============================================================================
                      EXCEL_REVISION.PY - NOZHGESS v1.0
==============================================================================
Módulo de generación de Excel con estilos profesionales.

Funcionalidades:
- Exportación de resultados por misión en hojas separadas
- Estilos automáticos de headers por tipo de columna
- Formato de anchos de columna
- Hoja de carga masiva opcional

Autor: Sistema Nozhgess
==============================================================================
"""
from __future__ import annotations
import os
import re
from datetime import datetime
from typing import Dict, List, Any, Optional

import pandas as pd

from Z_Utilidades.Principales.Terminal import log_info, log_ok, log_error

# =============================================================================
#                      IMPORTAR ESTILOS DE OPENPYXL
# =============================================================================
try:
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    STYLES_AVAILABLE = True
except ImportError:
    PatternFill = Font = Alignment = Border = Side = None
    get_column_letter = None
    STYLES_AVAILABLE = False


# =============================================================================
#                         PALETA DE COLORES POR GRUPO
# =============================================================================

# Colores profesionales para headers - AGRUPADOS
COLORS = {
    # Grupo Básico (Fecha, Rut, Edad) - Verde Oscuro con texto blanco en negrita
    "grupo_basico":      {"fill": "228B22", "font": "FFFFFF", "bold": True},
    
    # Grupo Objetivos (F Obj 1/2/3/4) - Celeste con texto blanco
    "grupo_objetivos":   {"fill": "00B0F0", "font": "FFFFFF", "bold": True},
    
    # Grupo Clasificación (Familia, Especialidad) - Amarillo con texto negro
    "grupo_clasificacion": {"fill": "FFFF00", "font": "000000", "bold": True},
    
    # Grupo Estado (Caso Encontrado, Mensual) - Morado con texto blanco
    "grupo_estado":      {"fill": "7030A0", "font": "FFFFFF", "bold": True},
    
    # Grupo Habilitantes (C Hab, F Hab, Hab Vi) - Naranjo con texto negro en negrita
    "grupo_habilitantes": {"fill": "FF8C00", "font": "000000", "bold": True},
    
    # Grupo Excluyentes (C Excluyente, F Excluyente) - Azul con texto blanco
    "grupo_excluyentes": {"fill": "4472C4", "font": "FFFFFF", "bold": True},
    
    # Grupo IPD (Fecha IPD, Estado IPD, Diagnóstico IPD) - Rojo fuerte con texto negro en negrita
    "grupo_ipd":         {"fill": "FF0000", "font": "000000", "bold": True},
    
    # Grupo OA (Código OA, Fecha OA, Folio OA, etc.) - Azul oscuro con texto blanco en negrita
    "grupo_oa":          {"fill": "00008B", "font": "FFFFFF", "bold": True},
    
    # Grupo APS (Fecha APS, Estado APS) - Amarillo con texto negro
    "grupo_aps":         {"fill": "FFC000", "font": "000000", "bold": True},
    
    # Grupo Observaciones (Observación, Observación Folio) - Rosado con texto negro en negrita
    "grupo_observacion": {"fill": "FFB6C1", "font": "000000", "bold": True},
    
    # Colores de celdas (para datos)
    "exito":             "C6EFCE",   # Verde claro
    "advertencia":       "FFEB9C",   # Amarillo claro
    "error":             "FFC7CE",   # Rojo claro
}


# =============================================================================
#                    MAPEO DE COLUMNAS A ESTILOS
# =============================================================================

def _get_header_style(column_name: str) -> dict:
    """
    Determina el estilo de header según el nombre de la columna.
    Agrupado por categorías para mejor visualización.
    
    Args:
        column_name: Nombre de la columna
        
    Returns:
        Dict con fill (color de fondo) y font (color de texto)
    """
    name = (column_name or "").lower().strip()
    
    # Grupo Básico: Fecha, Rut, Edad
    if name in ["fecha", "rut", "edad"]:
        return COLORS["grupo_basico"]
    
    # Grupo Objetivos: F Obj 1/2/3/4
    if name.startswith("f obj") or "objetivo" in name:
        return COLORS["grupo_objetivos"]
    
    # Grupo Clasificación: Familia, Especialidad
    if name in ["familia", "especialidad"]:
        return COLORS["grupo_clasificacion"]
    
    # Grupo Estado: Caso Encontrado, Estado Caso, Fecha Apertura, Fecha Cierre, Mensual
    if name in ["caso encontrado", "estado caso", "fecha apertura", "fecha cierre", "mensual"]:
        return COLORS["grupo_estado"]
    
    # Grupo Habilitantes: C Hab, F Hab, Hab Vi
    if "hab" in name and "exclu" not in name:
        return COLORS["grupo_habilitantes"]
    
    # Grupo Excluyentes: C Excluyente, F Excluyente
    if "excluy" in name:
        return COLORS["grupo_excluyentes"]
    
    # Grupo IPD
    if "ipd" in name:
        return COLORS["grupo_ipd"]
    
    # Grupo OA
    if " oa" in name or name.endswith("oa") or name.startswith("código oa") or name.startswith("fecha oa") or name.startswith("folio oa") or name.startswith("derivado oa") or name.startswith("diagnóstico oa"):
        return COLORS["grupo_oa"]
    
    # Grupo APS
    if "aps" in name:
        return COLORS["grupo_aps"]
    
    # Grupo Observaciones
    if "observ" in name:
        return COLORS["grupo_observacion"]
    
    # Default: usar grupo básico
    return COLORS["grupo_basico"]


# =============================================================================
#                      APLICAR ESTILOS A HOJA
# =============================================================================

def _aplicar_estilos(ws) -> None:
    """
    Aplica estilos profesionales a una hoja de Excel.
    
    - Headers con colores por tipo de columna
    - Bordes sutiles
    - Alineación centrada en headers
    - Ajuste automático de anchos
    """
    if not STYLES_AVAILABLE:
        return
    
    # Borde sutil
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    # Estilizar headers (fila 1)
    for cell in ws[1]:
        style = _get_header_style(str(cell.value or ""))
        
        cell.fill = PatternFill(start_color=style["fill"], end_color=style["fill"], fill_type="solid")
        # Usar negrita según configuración del estilo
        is_bold = style.get("bold", True)
        cell.font = Font(color=style["font"], bold=is_bold, size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    # Estilizar celdas de datos - Centrado horizontal y vertical
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            cell.font = Font(size=9)
    
    # Ajustar anchos de columna
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        
        for cell in column_cells:
            try:
                cell_length = len(str(cell.value or ""))
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
        
        # Ancho mínimo 10, máximo 50
        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[column_letter].width = adjusted_width


# =============================================================================
#                     FUNCIÓN PRINCIPAL DE EXPORTACIÓN
# =============================================================================

def generar_excel_revision(
    resultados_por_mision: Dict[int, List[Dict[str, Any]]],
    MISSIONS: List[Dict[str, Any]],
    NOMBRE_DE_LA_MISION: str,
    RUTA_CARPETA_SALIDA: str
) -> Optional[str]:
    """
    Genera el Excel final de revisión con estilos profesionales.
    
    Args:
        resultados_por_mision: Dict {indice_mision: [filas_dict]}
        MISSIONS: Lista de configuraciones de misión
        NOMBRE_DE_LA_MISION: Nombre general de la misión
        RUTA_CARPETA_SALIDA: Carpeta destino
        
    Returns:
        Ruta del archivo generado, o None si hay error
    """
    # Crear carpeta si no existe
    os.makedirs(RUTA_CARPETA_SALIDA, exist_ok=True)

    # Generar nombre de archivo con timestamp
    stamp = datetime.now().strftime("%d.%m.%Y_%H.%M")
    base_nombre = re.sub(r"[^\wáéíóúÁÉÍÓÚñÑ]+", "_", NOMBRE_DE_LA_MISION.strip())
    filename = f"Rev_{base_nombre}_{stamp}.xlsx"
    ruta_salida = os.path.join(RUTA_CARPETA_SALIDA, filename)

    try:
        with pd.ExcelWriter(ruta_salida, engine="openpyxl") as writer:
            
            # Una hoja por misión
            for i_mision, filas in resultados_por_mision.items():
                nombre_hoja = f"Mision {i_mision + 1}"
                
                if not filas:
                    # Hoja vacía con mensaje
                    df = pd.DataFrame([{
                        "Fecha": "",
                        "Rut": "",
                        "Nombre": "",
                        "Observación": "Sin datos procesados"
                    }])
                else:
                    df = pd.DataFrame(filas)
                
                df.to_excel(writer, index=False, sheet_name=nombre_hoja)
                
                # Aplicar estilos
                try:
                    ws = writer.sheets[nombre_hoja]
                    _aplicar_estilos(ws)
                except Exception:
                    pass  # Continuar sin estilos si hay error
            
            # Hoja de Carga Masiva
            try:
                ws_masiva = writer.book.create_sheet("Carga Masiva")
                ws_masiva.append(["Fecha", "Rut", "Dv", "Prestaciones", "Tipo", "Ps-Fam", "Especialidad"])
                _aplicar_estilos(ws_masiva)
            except Exception:
                pass

        log_ok(f"📁 Excel guardado: {filename}")
        return ruta_salida
        
    except Exception as e:
        log_error(f"Error guardando Excel: {str(e)[:50]}")
        return None
