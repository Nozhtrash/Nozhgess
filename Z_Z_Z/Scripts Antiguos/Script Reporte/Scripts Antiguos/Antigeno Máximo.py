from __future__ import annotations # Para type hints más limpios
import os, re, time, unicodedata, gc
from datetime import datetime
from typing import List, Optional, Dict, Tuple, Any
from collections import defaultdict

import pandas as pd
import warnings
# Evitar warning de openpyxl sobre 'Data Validation extension is not supported'
warnings.filterwarnings(
    "ignore",
    message="Data Validation extension is not supported and will be removed",
    category=UserWarning,
)
# ==============================================================================
# --- IMPORTACIÓN NUEVA PARA ESTILOS DE EXCEL ---
# ==============================================================================
try:
    from openpyxl.styles import PatternFill, Font, Alignment
except ImportError:
    print("ADVERTENCIA: 'openpyxl' no está instalado. El script funcionará, pero el Excel no tendrá colores.")
    # Clases Falsas (dummies) para que el script no falle si no está openpyxl
    class DummyFill:
        def __init__(self, *args, **kwargs): pass
    class DummyFont:
        def __init__(self, *args, **kwargs): pass
    class DummyAlignment:
        def __init__(self, *args, **kwargs): pass
    PatternFill, Font, Alignment = DummyFill, DummyFont, DummyAlignment

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException, ElementClickInterceptedException, ElementNotInteractableException,
    StaleElementReferenceException, WebDriverException, NoSuchWindowException
)

# =============================== PANEL DE CONTROL (R9) ====================================
# (Sin cambios)
NOMBRE_DE_LA_MISION = "Antigeno Prostatico"
RUTA_ARCHIVO_ENTRADA = r"C:\Users\usuariohgf\OneDrive\Documentos\Trabajo Oficina\Archivos\Reportes\Antígeno Prostático 1° Quincena (300) - Diciembre.xlsx"
RUTA_CARPETA_SALIDA  = r"C:\Users\usuariohgf\OneDrive\Documentos\Trabajo Oficina\Revisiones, Falta Enviar, CM, Listos\Revisiones\Revisiones Reportes"
DIRECCION_DEBUG_EDGE = "localhost:9222"
EDGE_DRIVER_PATH     = r"C:\\Windows\\System32\\msedgedriver.exe"

BASE_URL      = "https://www.sigges.cl" 
BUSQUEDA_URL  = f"{BASE_URL}/#/busqueda-de-paciente"
CERRAR_NAVEGADOR_AL_FINAL = False 

# Excel (0-based)
INDICE_COLUMNA_FECHA  = 0
INDICE_COLUMNA_RUT    = 3
INDICE_COLUMNA_NOMBRE = 2

# Reglas / ventanas
VENTANA_VIGENCIA_DIAS = 90

# Reintentos
MAX_REINTENTOS_GRAVES_POR_PACIENTE = 4

# Spinner
SPINNER_APPEAR_TIMEOUT = 2 
SPINNER_CSS = "dialog.loading"

# === Nuevos toggles / configuraciones ===
REVISAR_HISTORIA_COMPLETA: bool = False
MAX_OBJETIVOS: int = 5
XP_EDAD_PACIENTE = "/html/body/div/main/div[3]/div[3]/div/div[1]/div[7]/div[2]/div/p"
MINI_TABLA_STABLE_READS = 2 
TOP_K_HABILITANTES = 3
REVISAR_IPD = True
REVISAR_OA  = True
IPD_FILAS_REVISION = 1
OA_FILAS_REVISION  = 1
# VALIDAR_RUT = False # <-- ELIMINADO R9 (Innecesario)

MISSIONS: List[Dict[str, Any]] = [
    {
        "nombre": "Antigeno",
        "keywords": ["cancer de prostata", "cáncer de próstata"],
        "objetivo": "0305070",
        "usar_habilitantes": True,
        "habilitantes": ["0503101", "0504014", "0507506", "1703038", "1902055", "1902057", "1902065","2901001", "2901003", "2902001", 
            "2902002", "2902003", "2902004", "2902005","2902006", "2902007", "2902008", "3003412", "3106102",],
        "familia": "28",
        "especialidad": "07-800-2",
        "frecuencia": "Mensual",
        "edad_min": 15,
        "edad_max": None,
        "excluyentes": [],
        "min_habilitantes": 1
    },
]

# =============================== PANEL DE ESPERAS (R9) ==================================
# (Sin cambios)
ESPERAS: Dict[str, Dict[str, float]] = {
    # --- Tiempos Globales ---
    "page_load":           {"wait": 2, "sleep": 0.5},
    "default":             {"wait": 2, "sleep": 3}, # Base de 3s
    "spinner":             {"wait": 5.0, "sleep": 1}, 
    "spinner_largo":       {"wait": 120,"sleep": 3}, 
    "spinner_stuck":       {"wait": 300,"sleep": 3},

    # --- Navegación y Conexión ---
    "inicio_driver":       {"wait": 2,  "sleep": 0.5},
    "nav_a_busqueda":      {"wait": 2, "sleep": 0.5},
    "nav_a_busqueda_fast": {"wait": 2,  "sleep": 0.5}, 
    
    # --- Login (Se mantienen sleeps por seguridad/estabilidad de login) ---
    "login_check":         {"wait": 2,  "sleep": 0.5},
    "login_ingresar":      {"wait": 2,  "sleep": 0.5},
    "login_sel_unidad":    {"wait": 2,  "sleep": 0.5},
    "login_op_hospital":   {"wait": 2,  "sleep": 0.5},
    "login_ingreso_sigges":{"wait": 2,  "sleep": 0.5},
    "login_conectar":      {"wait": 2,  "sleep": 1.0},

    # --- Búsqueda de Paciente ---
    "find_rut":            {"wait": 1, "sleep": 0.3},
    "clear_rut":           {"wait": 1,  "sleep": 0.3},
    "send_rut":            {"wait": 1,  "sleep": 0.3},
    "click_buscar":        {"wait": 1, "sleep": 0.5},
    "mini_tabla_read":     {"wait": 3,  "sleep": 3}, # 2s -> 6.0s (Wait 3s / Sleep 3s)
    "leer_edad":           {"wait": 2,  "sleep": 0.5},
    
    # --- Cartola y Casos ---
    "click_ir_a_cartola":  {"wait": 2, "sleep": 0.5},
    "click_activar_hitos": {"wait": 2, "sleep": 0.5},
    "cont_cartola":        {"wait": 2, "sleep": 0.8},
    "leer_fallecimiento":  {"wait": 2,  "sleep": 0.8},
    # --- CAMBIO R9 ---
    "expandir_caso":       {"wait": 2,  "sleep": 0.8}, # 3.0s -> 0.8s (eliminamos "ronquido")
    "prestaciones_tbody":  {"wait": 5,  "sleep": 0.8}, # 2.0s -> 5.0s (aumentamos "acecho")
    "ipd_tbody":           {"wait": 5,  "sleep": 0.8}, # 2.0s -> 5.0s (aumentamos "acecho")
    "oa_tbody":            {"wait": 5,  "sleep": 0.8}, # 2.0s -> 5.0s (aumentamos "acecho")
    # --- FIN CAMBIO R9 ---
    
    # --- Reintentos (Se mantienen sleeps largos) ---
    "reintento_1":         {"wait": 3,  "sleep": 2},
    "reintento_2":         {"wait": 3,  "sleep": 5.0},
    "reintento_3":         {"wait": 3,  "sleep": 10},
    "reintento_4":         {"wait": 3,  "sleep": 10},
    
    # --- Final ---
    "entre_pacientes":     {"wait": 0.3,  "sleep": 0.3}, 
}

# =============================== COLORES / CONSOLA (R8) ===================================
# (Sin cambios)
try:
    from colorama import Fore, Style, init as colorama_init
    colorama_init(autoreset=True)
    # --- CAMBIO R8: Usamos los colores LIGHT..._EX que son más brillantes/neón ---
    OK, WARN, INFO, ERR, DATA, CODE, RESET = (
        Fore.LIGHTGREEN_EX, Fore.LIGHTYELLOW_EX, Fore.LIGHTCYAN_EX,
        Fore.LIGHTRED_EX, Fore.LIGHTMAGENTA_EX, Fore.LIGHTBLUE_EX, Style.RESET_ALL
    )
    C_BARRA    = Fore.LIGHTYELLOW_EX # Dorado Neón
    C_INDICE   = Fore.LIGHTCYAN_EX   # Cyan Neón
    C_HORA     = Fore.LIGHTRED_EX    # Rojo Neón
    C_NOMBRE   = Fore.LIGHTGREEN_EX  # Verde Neón
    C_RUT      = Fore.LIGHTMAGENTA_EX# Fucsia Neón
    C_FECHA    = Fore.LIGHTYELLOW_EX # Amarillo Neón
    C_CYAN     = Fore.LIGHTCYAN_EX
    C_FUCSIA   = Fore.LIGHTMAGENTA_EX
    C_VERDE    = Fore.LIGHTGREEN_EX
    C_ROJO     = Fore.LIGHTRED_EX
    C_NARANJA  = Fore.LIGHTYELLOW_EX # Naranja Neón / Dorado
    # --- FIN CAMBIO R8 ---
    C_M1_LABEL = C_CYAN   # M1 usará Cyan Neón
    C_M2_LABEL = C_FUCSIA # M2 usará Fucsia Neón
    C_REVISADO_LBL = C_CYAN
    C_EXITO    = C_VERDE     # "Éxito" en Verde Neón
    C_SIN_CASO = C_NARANJA   # "Sin Caso" en Naranja Neón
    C_FALLO    = C_ROJO      # "Fallo" en Rojo Neón
    C_SI       = C_VERDE     # "Sí", "Con OA", etc. en Verde Neón
    C_NO       = C_ROJO      # "No", "Sin OA", "Sin IPD" en Rojo Neón
except Exception:
    class Dummy: RESET_ALL = ""
    Fore = Style = Dummy()
    OK = WARN = INFO = ERR = DATA = CODE = RESET = ""
    C_BARRA = C_INDICE = C_HORA = C_NOMBRE = C_RUT = C_FECHA = ""
    C_CYAN = C_FUCSIA = C_VERDE = C_ROJO = C_NARANJA = ""
    C_M1_LABEL = C_M2_LABEL = C_REVISADO_LBL = C_EXITO = C_SIN_CASO = C_FALLO = C_SI = C_NO = ""


# =============================== LOGGING Y ESPERAS =================================
# (Sin cambios)
def log_error(msg: str, tipo: str = "ERROR") -> None:
    """Imprime un mensaje de error o advertencia en la consola."""
    color = ERR if tipo == "ERROR" else WARN
    try:
        # --- CAMBIO R7 --- (Sutil) Añadimos un espacio para que no se pegue al emoji si lo tuviera
        print(color + f"[{tipo}] {datetime.now().strftime('%H:%M:%S')} :: {msg}" + RESET)
    except Exception:
        print(f"[{tipo}] {datetime.now().strftime('%H:%M:%S')} :: {msg}")

def espera(clave: str) -> None:
    """Realiza una Pausa Fija (time.sleep) 100% silenciosa."""
    cfg = ESPERAS.get(clave, {"sleep": 0.5})
    t = float(cfg.get("sleep", 0.5))
    if t > 0:
        time.sleep(t)

# =============================== EXCEPCIONES / HELPERS (R9) ===============================
# (Sin cambios)
class SpinnerStuck(Exception): pass
_ddmmyyyy = re.compile(r"\b(\d{2}/\d{2}/\d{4})\b", re.I)
def _norm(s: str) -> str:
    if not s: return ""
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c)).replace("\xa0"," ").lower().strip()
    return re.sub(r"[\s]+", " ", re.sub(r"[^a-z0-9\sáéíóúüñ]"," ", s))
def has_keyword(texto: str, kws: List[str]) -> bool:
    t = _norm(texto); return any(_norm(k) in t for k in kws)
def solo_fecha(x: Any) -> str:
    if isinstance(x, datetime): return x.strftime("%d/%m/%Y")
    s = str(x).split(" ")[0].replace("-", "/"); p = s.split("/")
    return f"{p[2]}/{p[1]}/{0}" if len(p[0])==4 else s
def dparse(s: str) -> Optional[datetime]:
    try: return datetime.strptime(s.split(" ")[0], "%d/%m/%Y")
    except: return None
def same_month(a: datetime, b: datetime) -> bool:
    return a.year==b.year and a.month==b.month
def en_vig(fecha_obj: Optional[datetime], dt: Optional[datetime]) -> bool:
    if not (fecha_obj and dt): return False
    return 0 <= (fecha_obj-dt).days <= VENTANA_VIGENCIA_DIAS
def join_tags(tags: List[str]) -> str:
    tags = [t.strip() for t in tags if t and str(t).strip()]
    return " + ".join(tags)
def pad_with_X(vals: List[str], n: int) -> List[str]:
    vals = list(vals)[:n]
    while len(vals) < n: vals.append("X")
    return vals
# def join_multi(vals: List[str], n: int) -> str: # <-- ELIMINADO R9 (Innecesario)
#     return " | ".join(pad_with_X(vals, n)) # <-- ELIMINADO R9 (Innecesario)
def normalizar_rut(rut: str) -> str:
    rut = rut.replace(".", "").replace("-", "").strip().upper()
    if not rut: return ""
    return rut[:-1] + "-" + rut[-1]
# def validar_rut_chileno(rut: str) -> bool: # <-- ELIMINADO R9 (Innecesario)
#     try:
#         r = rut.replace(".", "").replace("-", "").upper()
#         cuerpo, dv = r[:-1], r[-1]
#         suma, mult = 0, 2
#         for c in reversed(cuerpo):
#             suma += int(c)*mult
#             mult = 2 if mult==7 else mult+1
#         resto = suma % 11
#         dig = 11 - resto
#         dig = "0" if dig==11 else "K" if dig==10 else str(dig)
#         return dig == dv
#     except: return False

# =============================== XPATHs BASE / LOGIN (R9) =================================
# (Sin cambios)
XP_CONT_CARTOLA   = "/html/body/div/main/div[3]/div[2]/div[1]/div[5]/div[1]/div[2]"
XP_RUT_INPUT      = "/html/body/div/main/div[3]/div[3]/div/div/div[1]/div/div/div/input"
XP_BTN_BUSCAR     = "/html/body/div/main/div[3]/div[3]/div/div/div[1]/div/button"
XP_MENU_BUSCAR    = "/html/body/div/main/div[2]/nav/div[1]/div[2]/a[1]"
XP_MENU_CARTOLA   = "/html/body/div/main/div[2]/nav/div[1]/div[2]/a[2]"
XP_CHK_HITOS_GES  = "/html/body/div/main/div[3]/div[2]/div[1]/div[5]/div[1]/div/div/label/input"
LOGIN_BTN_INGRESAR = [
    "//button[@type='submit' and contains(@class,'botonBase')][.//p[normalize-space(.)='Ingresar']]",
    "/html/body/div/div/div[2]/div[1]/form/div[3]/button",
]
LOGIN_SEL_UNIDAD_HEADER = [
    "//div[contains(@class,'filtroSelect__header')][.//p[contains(translate(normalize-space(.),'ÁÉÍÓÚ','áéíóú'),'seleccione una unidad')]]",
    "/html/body/div/div/div/div/div[1]/div/div[2]/div[1]/div/div[2]/div[1]",
]
LOGIN_OP_HOSPITAL = [
    "//div[contains(@class,'filtroSelect__option')][.//p[contains(translate(normalize-space(.),'ÁÉÍÓÚ','áéíóú'),'gustavo fricke')]]",
    "/html/body/div/div/div/div/div[1]/div/div[2]/div[1]/div/div[2]/div[2]/div",
]
LOGIN_TILE_INGRESO_SIGGES = [
    "/html/body/div/div/div/div/div[1]/div/div[2]/div[2]/div[2]/div[3]",
    "//div[contains(@class,'boxItems__item')][.//p[normalize-space(.)='INGRESO SIGGES CONFIDENCIAL']]"
]
LOGIN_BTN_CONECTAR = [
    "//button[contains(@class,'botonBase')][.//p[contains(translate(normalize-space(.),'ÁÉÍÓÚ','áéíóú'),'conectese')]]",
    "/html/body/div/div/div/div/div[1]/div/div[2]/div[3]/button",
]
IPD_TBODY_CANDIDATE_ABS = [
    "/html/body/div/main/div[3]/div[2]/div[1]/div[5]/div[1]/div[2]/div[2]/div[4]/div[2]/div/table/tbody",
    "/html/body/div/main/div[3]/div[2]/div[1]/div[5]/div[1]/div[2]/div[7]/div[4]/div[2]/div/table/tbody",
]
# XPATHS_CHECKBOX_NO_GES = [ # <-- ELIMINADO R9 (Innecesario)
#     "/html/body/div/main/div[3]/div[2]/div[1]/div[5]/div[2]/div[1]/div/label/input",
#     "//div[label[contains(normalize-space(.),'Hitos No GES')]]//input[@type='checkbox']",
# ]

# =============================== WRAPPER SIGGES (R8) ======================================
# (Sin cambios)
class SiggesDriver:
    def __init__(self, driver: webdriver.Remote) -> None:
        self.driver = driver

    # spinner
    def hay_spinner(self) -> bool:
        try: return bool(self.driver.find_elements(By.CSS_SELECTOR, SPINNER_CSS))
        except Exception: return False

    def esperar_spinner(self, appear_timeout=SPINNER_APPEAR_TIMEOUT,
                        clave_espera="spinner",
                        raise_on_timeout=False):
        
        cfg = ESPERAS.get(clave_espera, {"wait": 25.0, "sleep": 0.5})
        disappear_timeout = cfg["wait"]
        
        seen=False
        try:
            end = time.time() + appear_timeout 
            while time.time() < end:
                if self.hay_spinner(): 
                    seen=True; 
                    break
                time.sleep(0.51) 
        except Exception:
            seen=False
            
        if not seen:
            return
        
        try:
            WebDriverWait(self.driver, disappear_timeout).until(
                EC.invisibility_of_element_located((By.CSS_SELECTOR, SPINNER_CSS)))
            
            # === INSERCIÓN DE BLINDAJE (R8) ===
            # Regla "Trabajo Seguro y Perfecto": Pausa extra de 0.5s post-spinner
            # (Solo si el spinner fue visto y ya desapareció)
            time.sleep(0.5) 
            # === FIN INSERCIÓN ===

        except TimeoutException:
            if self.hay_spinner() and raise_on_timeout:
                raise SpinnerStuck(f"Spinner no desapareció en {disappear_timeout}s")
        
    # --- INICIO REFACTOR R7 (Volvemos a la lógica R5) ---
    def _find(self, locators: List[str], mode="clickable", clave_espera="default") -> Optional[Any]:
        """
        Lógica R5: Itera por los locators. Si uno falla, prueba el siguiente.
        Sin re-intentos dobles. Simple y efectivo.
        """
        cond = {"presence":EC.presence_of_element_located,
                "visible":EC.visibility_of_element_located}.get(mode, EC.element_to_be_clickable)
        
        # Usar un fallback de 3s si la clave no está
        cfg = ESPERAS.get(clave_espera, {"wait": 2})
        # Usar el timeout de la clave (ej: 'ipd_tbody' usará 5.0s)
        timeout = float(cfg.get("wait", 2)) 
        
        for xp in locators:
            try:
                # Intento único por locator, con el timeout completo
                return WebDriverWait(self.driver, timeout).until(cond((By.XPATH,xp)))
            except Exception: 
                # Si falla (por Timeout u otro), simplemente continúa con el siguiente locator
                continue
        
        # Si todos los locators fallan
        return None
    # --- FIN REFACTOR R7 ---

    def _click(self, locators: List[str], scroll=True, wait_spinner=True,
               clave_espera="default", 
               spinner_clave_espera="spinner", 
               raise_on_spinner_timeout=False) -> bool:
        
        # _find ahora usa la lógica R5 (simple) con el timeout de la clave
        el = self._find(locators, "clickable", clave_espera)
        if not el: return False
        
        try:
            if scroll:
                try: self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
                except Exception: pass
            
            el.click()
        except (ElementClickInterceptedException, ElementNotInteractableException, StaleElementReferenceException, WebDriverException):
            try: self.driver.execute_script("arguments[0].click();", el)
            except Exception: return False
        
        if wait_spinner:
            self.esperar_spinner(clave_espera=spinner_clave_espera, raise_on_timeout=raise_on_spinner_timeout)
        
        return True

    # sesión
    def sesion_cerrada(self) -> bool:
        try:
            url = self.driver.current_url or ""
            if "/#/login" in url: return True
            if self._find(LOGIN_BTN_INGRESAR, "visible", "login_check"): return True
        except Exception: pass
        return False

    def ir(self, url: str) -> None:
        try:
            self.driver.get(url)
        except TimeoutException:
            log_error(f"Timeout de carga de página en {url}. Deteniendo carga.", tipo="WARN")
            try: self.driver.execute_script("window.stop();")
            except Exception: pass
        except Exception as e:
            log_error(f"Error general en driver.get({url}): {e}", tipo="WARN")
            pass 

    def asegurar_en_busqueda(self) -> None:
        try:
            if self.sesion_cerrada():
                log_error("Sesión cerrada. Intentando login...", tipo="WARN")
                if not intentar_login(self, reintentar_una_vez=True):
                    raise Exception("Fallo de login irrecuperable.")
                self.ir(BUSQUEDA_URL)

            current_url = self.driver.current_url or ""
            
            if BUSQUEDA_URL in current_url:
                try:
                    # Usamos _find (R7)
                    if self._find([XP_RUT_INPUT], "presence", "nav_a_busqueda_fast"):
                        return
                except Exception:
                    pass 
                
                log_error("En URL de búsqueda, pero no se encontró input RUT. Forzando recarga.", tipo="WARN")
                self.ir(BUSQUEDA_URL)
            
            else:
                self.ir(BUSQUEDA_URL) 

            self.esperar_spinner(raise_on_timeout=False) 
            
            if not self._find([XP_RUT_INPUT], "presence", "nav_a_busqueda"):
                raise TimeoutException("No se pudo cargar la página de Búsqueda de Paciente (Timeout).")
            
        except TimeoutException:
            log_error("No se pudo cargar la página de Búsqueda de Paciente (Timeout).", tipo="ERROR")
            raise 
        except Exception as e:
            log_error(f"Error general en asegurar_en_busqueda: {e}", tipo="ERROR")
            raise
            
    def ir_a_cartola(self, raise_on_spinner_timeout=True) -> bool:
        return self._click(
            [XP_MENU_CARTOLA], 
            scroll=False, 
            wait_spinner=True,
            clave_espera="click_ir_a_cartola",
            spinner_clave_espera="spinner_largo", 
            raise_on_spinner_timeout=raise_on_spinner_timeout
        )

    def activar_hitos_ges(self) -> None:
        chk = self._find([XP_CHK_HITOS_GES], "presence", "click_activar_hitos")
        if not chk: return
        try:
            if not chk.is_selected():
                self._click([XP_CHK_HITOS_GES], True, True, clave_espera="click_activar_hitos")
        except Exception: pass

    # cartola
    def lista_de_casos_cartola(self) -> List[str]:
        root = self._find([XP_CONT_CARTOLA], "presence", "cont_cartola")
        if not root: return []
        
        out=[]
        try:
            for c in root.find_elements(By.CSS_SELECTOR,"div.contRow"):
                t = (c.text or "").strip()
                if t: out.append(t)
        except Exception: pass
        return out

    def _case_root(self, i:int) -> Optional[Any]:
        try: return self._find([f"({XP_CONT_CARTOLA})/div[{i+1}]"], "presence", "expandir_caso")
        except Exception: return None

    def expandir_caso(self, i:int) -> Optional[Any]:
        root = self._case_root(i)
        if not root: return None
        try:
            if not root.find_elements(By.XPATH, ".//table"):
                self._click(
                    [f"({XP_CONT_CARTOLA})/div[{i+1}]/div/label/input", f"({XP_CONT_CARTOLA})/div[{i+1}]"], 
                    wait_spinner=False, 
                    clave_espera="expandir_caso"
                )
        except Exception: pass
        
        espera("expandir_caso") # <-- PAUSA R9 (0.8s)
        return root

    def cerrar_caso_por_indice(self, i:int)->None:
        try:
            root = self._case_root(i)
            if root and root.find_elements(By.XPATH, ".//table"):
                self._click(
                    [f"({XP_CONT_CARTOLA})/div[{i+1}]/div/label/input", f"({XP_CONT_CARTOLA})/div[{i+1}]"], 
                    wait_spinner=False, 
                    clave_espera="expandir_caso"
                )
        except Exception: pass
        espera("expandir_caso")

    def _prestaciones_tbody(self, i:int) -> Optional[Any]:
        base = XP_CONT_CARTOLA
        return self._find(
            [f"({base})/div[{i+1}]/div[6]/div[2]/div/table/tbody", f"({base})/div[{i+1}]//table/tbody"],
            "presence",
            "prestaciones_tbody" # <-- USA "ACECHO" R9 (5s)
        )

    def leer_prestaciones_desde_tbody(self, tbody: Any) -> List[Dict[str,str]]:
        out=[]
        try:
            for tr in reversed(tbody.find_elements(By.TAG_NAME,"tr")):
                tds = tr.find_elements(By.TAG_NAME,"td")
                if len(tds)>8:
                    f = (tds[2].text.strip().split(" ")[0] if len(tds)>2 else "")
                    if not dparse(f):
                        for c in tds:
                            m=_ddmmyyyy.search(c.text)
                            if m: f = m.group(1); break
                    out.append({"fecha":f,"codigo":tds[7].text.strip(),"glosa":tds[8].text.strip()})
        except Exception: pass
        return out

    def leer_fallecimiento(self) -> Optional[datetime]:
        try:
            el_fallecido = self._find(
                ["//div[span[normalize-space(.)='Fecha de fallecimiento'] and p]"],
                "presence",
                "leer_fallecimiento"
            )
            if el_fallecido:
                f = dparse((el_fallecido.find_element(By.TAG_NAME,"p").text or "").strip())
                if f: return f
        except Exception: pass
        return None

    def leer_edad(self) -> Optional[int]:
        try:
            locators = [
                XP_EDAD_PACIENTE, # (Tu XPath absoluto)
                "//div[p[contains(translate(normalize-space(.),'AÑOS','años'),'años')]]/p",
                "//p[contains(normalize-space(.),'Años') or contains(normalize-space(.),'años')]",
            ]
            el = self._find(locators, "presence", "leer_edad")
            # --- FIN CAMBIO ---

            if not el: return None
            
            txt = (el.text or "").strip()
            # Busca el primer grupo de dígitos (\d+) en el texto que se parezca a '89 Años, ...'
            m = re.search(r"(\d+)\s*Añ", txt, re.I) # Busca el número seguido de " Añ" para asegurar que es la edad
            if not m:
                # Fallback: Busca el primer número si el patrón 'Años' no se encuentra cerca
                m = re.search(r"(\d+)", txt)
                
            if m:
                return int(m.group(1))
        except Exception:
            return None
        return None

    # IPD helpers (Sin cambios)
    def _find_section_label_p(self, root, needle: str):
        nd = _norm(needle)
        try:
            for el in root.find_elements(By.XPATH, ".//div/label/p"):
                if _norm(el.text) and nd.split('(')[0].strip() in _norm(el.text):
                    return el
        except Exception: pass
        return None
    def _tbody_from_label_p(self, p_el):
        for xp in ["../../../following-sibling::div[1]//table/tbody",
                   "../../following-sibling::div[1]//table/tbody",
                   "../following-sibling::div[1]//table/tbody",
                   "ancestor::div[1]/following-sibling::div[1]//table/tbody",
                   "ancestor::div[2]/following-sibling::div[1]//table/tbody",
                   "ancestor::div[3]/following-sibling::div[1]//table/tbody"]:
            try:
                tb = p_el.find_element(By.XPATH, xp)
                if tb: return tb
            except Exception: continue
        return None
    def _thead_from_tbody(self, tbody):
        try: return tbody.find_element(By.XPATH, "../thead")
        except Exception:
            try: return tbody.find_element(By.XPATH, "ancestor::table/thead")
            except Exception: return None
    def _map_indices_thead(self, thead_el) -> dict:
        mapping = {}
        try:
            ths = thead_el.find_elements(By.TAG_NAME, "th")
            for i, th in enumerate(ths):
                lab = _norm(th.text or "")
                if lab: mapping[lab] = i
        except Exception: pass
        return mapping
    def _thead_has_ipd(self, thead) -> bool:
        try:
            mp = self._map_indices_thead(thead)
            has_diag = any("diagnostico" in k for k in mp.keys())
            has_conf = any("confirma" in k or "descart" in k or "confirm" in k for k in mp.keys())
            return has_diag and has_conf
        except Exception: return False

    def leer_ipd_desde_caso(self, root: Any, n:int) -> Tuple[List[str], List[str], List[str]]:
        fechas=[]; estados=[]; diags=[]
        if not root:
            return pad_with_X([], n), pad_with_X([], n), pad_with_X([], n)
        try:
            p = self._find_section_label_p(root, "informes de proceso de diagnóstico")
            thead, tbody = None, None
            if p:
                try:
                    txt = (p.text or "").strip()
                    m = re.search(r"\((\d+)\)", txt)
                    if m and int(m.group(1)) == 0:
                        return (pad_with_X(["NO TIENE IPD"], n),
                                pad_with_X(["NO TIENE IPD"], n),
                                pad_with_X(["NO TIENE IPD"], n))
                except Exception: pass
                tbody = self._tbody_from_label_p(p); thead = self._thead_from_tbody(tbody) if tbody else None
            
            if not tbody:
                for xp in IPD_TBODY_CANDIDATE_ABS:
                    tb_rel = None
                    try: tb_rel = root.find_element(By.XPATH, "." + xp.split("/body")[1])
                    except Exception: pass
                    
                    if tb_rel: 
                        tb = tb_rel
                    else: 
                        # Usar _find (R7)
                        tb = self._find([xp], "presence", "ipd_tbody") # <-- USA "ACECHO" R9 (5s)
                    
                    if tb:
                        th = self._thead_from_tbody(tb)
                        if th and self._thead_has_ipd(th): 
                            tbody, thead = tb, th
                            break
            
            if not tbody: return pad_with_X([], n), pad_with_X([], n), pad_with_X([], n)

            idx_fecha, idx_estado, idx_diag = 2, 6, 7
            if thead:
                try:
                    mp = self._map_indices_thead(thead)
                    for k,v in mp.items():
                        if "fecha ipd" in k or ("fecha" in k and v != 0): idx_fecha = v
                    for k,v in mp.items():
                        if "confirma" in k or "descart" in k or "confirm" in k: idx_estado = v
                    for k,v in mp.items():
                        if "diagnostico" in k: idx_diag = v
                except Exception: pass
            
            rows = tbody.find_elements(By.TAG_NAME, "tr") or []
            parsed=[]
            for r in rows:
                try:
                    tds = r.find_elements(By.TAG_NAME, "td")
                    f_txt = (tds[idx_fecha].text or "").strip() if len(tds)>idx_fecha else ""
                    if not f_txt:
                        m = _ddmmyyyy.search(r.text or "")
                        if m: f_txt = m.group(1)
                    e_txt = (tds[idx_estado].text or "").strip() if len(tds)>idx_estado else ""
                    d_txt = (tds[idx_diag].text or "").replace("\n"," ").strip() if len(tds)>idx_diag else ""
                    f_dt = dparse(f_txt) or datetime.min
                    parsed.append((f_dt, f_txt, e_txt, d_txt))
                except Exception: continue
            
            parsed.sort(key=lambda x: x[0], reverse=True)
            parsed = parsed[:n]
            fechas  = [p[1] for p in parsed]
            estados = [p[2] for p in parsed]
            diags   = [p[3] for p in parsed]
        
        except Exception:
             return pad_with_X([], n), pad_with_X([], n), pad_with_X([], n)
        
        return pad_with_X(fechas, n), pad_with_X(estados, n), pad_with_X(diags, n)

    # ==============================================================================
    # --- LEER_OA_DESDE_CASO ---
    # (Sin cambios, ya estaba modificado para Código OA)
    # ==============================================================================
    def leer_oa_desde_caso(self, root: Any, n:int) -> Tuple[List[str], List[str], List[str], List[str]]: # <-- CAMBIO 1: Firma de tipo
        fechas=[]; derivados=[]; diags=[]; codigos=[] # <-- CAMBIO 2: Lista nueva
        if not root:
             # <-- CAMBIO 3: Devolver 4 listas
             return pad_with_X([], n), pad_with_X([], n), pad_with_X([], n), pad_with_X([], n)
        
        try:
            p = self._find_section_label_p(root, "ordenes de atencion")
            tbody = None
            if p:
                try:
                    txt = (p.text or "").strip()
                    m = re.search(r"\((\d+)\)", txt)
                    if m and int(m.group(1)) == 0:
                        # <-- CAMBIO 4: Devolver 4 listas (en caso 'NO TIENE OA')
                        return (pad_with_X(['NO TIENE OA'], n), 
                                pad_with_X(['NO TIENE OA'], n), 
                                pad_with_X(['NO TIENE OA'], n),
                                pad_with_X(['NO TIENE OA'], n))
                except Exception: pass
                tbody = self._tbody_from_label_p(p)
            
            if not tbody:
                xp_fallback = ".//div[div/label/p[contains(translate(normalize-space(.),'ÁÉÍÓÚ','áéíóú'),'ordenes de atencion')]]/following-sibling::div[1]//table/tbody"
                try: 
                    # Intentamos encontrar la tabla relativa al root
                    tbody = root.find_element(By.XPATH, xp_fallback)
                except Exception: 
                    # Si falla, usamos el "acecho" global (R9)
                    tbody = self._find([xp_fallback.replace(".//","//")], "presence", "oa_tbody")

            # <-- CAMBIO 5: Devolver 4 listas (si no se encuentra tbody)
            if not tbody: return pad_with_X([], n), pad_with_X([], n), pad_with_X([], n), pad_with_X([], n)

            rows = tbody.find_elements(By.TAG_NAME, "tr") or []
            parsed=[]
            for r in rows:
                try:
                    tds = r.find_elements(By.TAG_NAME, "td")
                    f_txt = (tds[2].text or "").split(" ")[0].strip() if len(tds)>2 else ""
                    # --- INICIO NUEVA LÍNEA ---
                    c_cod = (tds[9].text or "").strip() if len(tds)>9 else "" # <-- CAMBIO 6: Extraer el código (td 10)
                    # --- FIN NUEVA LÍNEA ---
                    d_para = (tds[8].text or "").strip() if len(tds)>8 else ""
                    d_diag = (tds[12].text or "").strip() if len(tds)>12 else ""
                    if not f_txt:
                        m = _ddmmyyyy.search(r.text or "")
                        if m: f_txt = m.group(1)
                    f_dt = dparse(f_txt) or datetime.min
                    
                    # <-- CAMBIO 7: Añadir el código al tuple
                    parsed.append((f_dt, f_txt, d_para, d_diag, c_cod)) 
                except Exception: continue
            
            parsed.sort(key=lambda x: x[0], reverse=True)
            parsed = parsed[:n]
            fechas   = [p[1] for p in parsed]
            derivados= [p[2] for p in parsed]
            diags    = [p[3] for p in parsed]
            codigos  = [p[4] for p in parsed] # <-- CAMBIO 8: Desempaquetar la nueva lista
        
        except Exception:
            # <-- CAMBIO 9: Devolver 4 listas (en caso de error)
            return pad_with_X([], n), pad_with_X([], n), pad_with_X([], n), pad_with_X([], n)
        
        # <-- CAMBIO 10: Devolver 4 listas (retorno final)
        return pad_with_X(fechas, n), pad_with_X(derivados, n), pad_with_X(diags, n), pad_with_X(codigos, n)

# =============================== MINI TABLA ==========================================
# (Sin cambios)
XPATH_MINI_TBODY_CANDIDATES = [
    "//div[@class='contBody maxW scroll']//div[contains(@class,'cardTable')]/table/tbody",
    "/html/body/div/main/div[3]/div[3]/div/div[2]/div/div/table/tbody",
]
def leer_mini_tabla_busqueda(driver: Any, stable_reads=MINI_TABLA_STABLE_READS) -> List[Dict[str,str]]:
    
    tbody=None
    for xp in XPATH_MINI_TBODY_CANDIDATES:
        try:
            # Usará el timeout de 3.0s de "mini_tabla_read" (wait: 3)
            tbody = WebDriverWait(driver, ESPERAS["mini_tabla_read"]["wait"]).until(
                EC.presence_of_element_located((By.XPATH,xp))
            )
            # Si lo encuentra, rompemos el bucle
            if tbody: break 
        except Exception: 
            # Si falla, probamos el siguiente XPath
            continue
        
    if not tbody: return []
    
    # --- Lógica de Lectura Estable (R5) ---
    last_snap_key = ""
    stable_count = 0
    final_snap = []
    
    reads_to_perform = max(stable_reads, 1)
    
    if reads_to_perform == 1:
        # Modo "nitro" (una sola lectura)
        try:
            final_snap = driver.execute_script("""
                const tb = arguments[0];
                if (!tb) return [];
                const rows = Array.from(tb.querySelectorAll('tr'));
                return rows.map(tr => {
                    const tds = tr.querySelectorAll('td');
                    if (tds.length < 4) return null;
                    const problema = tds[1].innerText.trim();
                    const decreto  = tds[2].innerText.trim();
                    const estado   = tds[3].innerText.trim();
                    return {problema, decreto, estado};
                }).filter(Boolean);
            """, tbody) or []
        except Exception: final_snap = []
    
    else:
        # Lógica de lecturas estables (para stable_reads >= 2)
        end_time = time.time() + 5.0 
        
        while time.time() < end_time:
            try:
                snap = driver.execute_script("""
                    const tb = arguments[0];
                    if (!tb) return [];
                    const rows = Array.from(tb.querySelectorAll('tr'));
                    return rows.map(tr => {
                        const tds = tr.querySelectorAll('td');
                        if (tds.length < 4) return null;
                        const problema = tds[1].innerText.trim();
                        const decreto  = tds[2].innerText.trim();
                        const estado   = tds[3].innerText.trim();
                        return {problema, decreto, estado};
                    }).filter(Boolean);
                """, tbody) or []
            except Exception: snap = []
            
            current_key = "|".join([r.get("problema","") for r in snap])
            final_snap = snap 
            
            if not current_key and not last_snap_key:
                stable_count = reads_to_perform 
            
            if current_key == last_snap_key:
                stable_count += 1
            else:
                stable_count = 0 
                
            last_snap_key = current_key
            
            if stable_count >= (reads_to_perform - 1): 
                break 
                
            time.sleep(0.55) # <--- Mini-pausa activa entre lecturas
    
    # --- FIN LÓGICA ROBUSTEZ (R5) ---

    out = [r for r in final_snap if "evento sin caso" not in _norm(r["problema"])]
    return out
# --- FIN REFACTOR R7 ---

# =============================== LÓGICA DE MISIONES ==================================
# (Sin cambios)
def elegir_caso_mas_reciente(casos: List[str], kws: List[str]) -> Tuple[Optional[int], Optional[str], Optional[datetime]]:
    best_i, best_t, best_f = None, None, None
    for i, t in enumerate(casos):
        if not has_keyword(t, kws): continue
        m=_ddmmyyyy.search(t); f = dparse(m.group(1)) if m else None
        if best_f is None or (f and f>best_f):
            best_i, best_t, best_f = i, (t.strip() if t else ""), (f or best_f)
    return best_i, best_t, best_f
def mensual_categoria(fechas: List[datetime], fobj: datetime) -> str:
    if not fechas: return "Sin Día"
    u=sorted(set(fechas)); md = any(f==fobj for f in u)
    pri=[f for f in u if f.day<=15]; seg=[f for f in u if f.day>=16]
    if md and len(u)==1: return "Mismo Día"
    if pri and not seg: return "Primera Quincena"
    if seg and not pri: return "Segunda Quincena"
    if pri and seg: return "Primera + Segunda"
    return "Sin Día"
def listar_habilitantes(prest: List[Dict[str,str]], cods: List[str], fobj: Optional[datetime]) -> List[Tuple[str, datetime]]:
    out=[]
    for p in prest:
        c = p.get("codigo","")
        if c in cods:
            f = dparse(p.get("fecha",""))
            if f and (not fobj or f <= fobj): out.append((c,f))
    return sorted(out, key=lambda x:x[1], reverse=True)

# ==============================================================================
# --- MODIFICACIÓN 1: COLS_MISION ---
# (Añadida "Fecha Excluyente")
# ==============================================================================
def cols_mision(m: Optional[Dict[str,Any]] = None) -> List[str]:
    # Esta lista define el ORDEN y NOMBRE FINAL de las columnas en Excel
    cols: List[str] = [
        "Fecha", "Rut", "Edad", 
        "Cod Obj", # <-- Nuevo Nombre
        "Familia", "Especialidad", "Caso Encontrado", 
        "Fecha Obj", # <-- Nuevo Nombre (era "Objetivo")
        "Mensual", 
        "Habilitante", # <-- Nuevo Nombre
        "Fecha Hab", # <-- Nuevo Nombre
        "Excluyente", # <-- Nueva Columna
        "Fecha Excluyente", # <-- NUEVA COLUMNA
        "Fecha IPD", "Estado IPD", 
        "Diag IPD", # <-- Nuevo Nombre
        "Fecha OA", 
        "Cod OA", # <-- Nuevo Nombre
        "Deri OA", # <-- Nuevo Nombre
        "Diag OA", # <-- Nuevo Nombre
        "Observación",
    ]
    
    # Esta parte se mantiene para los "Objetivo 2", "Objetivo 3", etc.
    if m is not None:
        objetivos_list: List[str] = []
        if m.get("objetivos"):
            objetivos_list = [str(c).strip() for c in m.get("objetivos", [])][:MAX_OBJETIVOS]
        else:
            obj = m.get("objetivo")
            objetivos_list = [str(obj).strip()] if obj else []
        
        # Añadimos los objetivos extra (si existen) después de "Fecha Obj"
        obj_idx = cols.index("Fecha Obj") + 1
        for idx in range(1, len(objetivos_list)):
            cols.insert(obj_idx, f"Objetivo {idx+1}")
            obj_idx += 1
            
    return cols

# ==============================================================================
# --- MODIFICACIÓN 2: VAC_ROW ---
# (Añadida "Fecha Excluyente")
# ==============================================================================
def vac_row(m: Dict[str,Any], fecha: str, rut: str, obs: str="") -> Dict[str,Any]:
    objetivos: List[str] = []
    if m.get("objetivos"):
        objetivos = [str(c).strip() for c in m.get("objetivos", [])][:MAX_OBJETIVOS]
    else:
        obj = m.get("objetivo")
        objetivos = [str(obj).strip()] if obj else []
    objetivos = objetivos[:MAX_OBJETIVOS]
    cod_objetivo = " | ".join([c for c in objetivos if c])
    
    # Los nombres de las claves aquí son INTERNOS. Se renombrarán antes de exportar.
    row: Dict[str,Any] = {
        "Fecha": fecha, "Rut": rut, "Edad": "", 
        "Cod Objetivo": cod_objetivo, # <-- Clave interna
        "Familia": m.get("familia", ""), "Especialidad": m.get("especialidad", ""),
        "Caso Encontrado": "Sin caso", 
        "Objetivo": "N/A", # <-- Clave interna
        "Mensual": "Sin Día",
        # "Seguimiento": "Sin Caso", # <-- ELIMINADO
        "Fecha IPD": "", "Estado IPD": "",
        "Diagnóstico IPD": "", # <-- Clave interna
        "Fecha OA": "", 
        "Derivado OA": "", # <-- Clave interna
        "Diagnóstico OA": "", # <-- Clave interna
        "Código OA": "", # <-- Clave interna
        "Habilitan": "", # <-- Clave interna
        "Fecha Háb": "", # <-- Clave interna
        # "Hab Vi": "", # <-- ELIMINADO
        "Excluyente": "", # <-- NUEVA CLAVE
        "Fecha Excluyente": "", # <-- NUEVA CLAVE
        "Observación": obs,
    }
    # Esta parte se mantiene
    for idx in range(1, len(objetivos)):
        col_name = f"Objetivo {idx+1}"
        row[col_name] = "N/A"
    return row

def add_vacios(filas_por_mision, MISSIONS, fecha, rut, obs):
    for i_m, m in enumerate(MISSIONS):
        filas_por_mision[i_m].append(vac_row(m, fecha, rut, obs))

# =============================== LOGIN / REINTENTOS ==================================
# (Sin cambios)
def safe_refresh(driver):
    try: driver.refresh()
    except Exception: pass
def intentar_login(sig: SiggesDriver, reintentar_una_vez=True) -> bool:
    try:
        sig._click(LOGIN_BTN_INGRESAR, scroll=False, wait_spinner=True, clave_espera="login_ingresar", spinner_clave_espera="spinner_largo")
        sig._click(LOGIN_SEL_UNIDAD_HEADER, scroll=True, wait_spinner=False, clave_espera="login_sel_unidad")
        sig._click(LOGIN_OP_HOSPITAL, scroll=True, wait_spinner=False, clave_espera="login_op_hospital")
        sig._click(LOGIN_TILE_INGRESO_SIGGES, scroll=True, wait_spinner=False, clave_espera="login_ingreso_sigges")
        sig._click(LOGIN_BTN_CONECTAR, scroll=True, wait_spinner=True, clave_espera="login_conectar", spinner_clave_espera="spinner_largo")
        if "actualizaciones" in sig.driver.current_url:
            return True
    except Exception as e:
        log_error(f"Fallo durante el login: {e}", tipo="WARN")
        pass 
    if reintentar_una_vez:
        log_error("Fallo de login, reintentando una vez tras refresh...", tipo="WARN")
        safe_refresh(sig.driver)
        espera("reintento_1") 
        return intentar_login(sig, reintentar_una_vez=False) 
    return False
def manejar_reintentos_inteligentes(sig: SiggesDriver, intento: int) -> None:
    if intento == 1:
        espera("reintento_1")
        try:
            sig.ir(BUSQUEDA_URL) 
            sig.esperar_spinner(clave_espera="spinner", raise_on_timeout=False)
            if sig.sesion_cerrada():
                intentar_login(sig, reintentar_una_vez=True)
                sig.ir(BUSQUEDA_URL) 
                sig.esperar_spinner(raise_on_timeout=False)
        except Exception: pass
        return
    if intento == 2:
        espera("reintento_2")
        try:
            if sig.hay_spinner(): 
                safe_refresh(sig.driver) 
            url = (sig.driver.current_url or "")
            if BUSQUEDA_URL in url:
                safe_refresh(sig.driver)
                sig.esperar_spinner(raise_on_timeout=False)
            elif any(x in url for x in ["/#/93", "/#/actualizaciones"]):
                sig.ir(BUSQUEDA_URL)
                sig.esperar_spinner(raise_on_timeout=False)
            elif "/login" in url or sig.sesion_cerrada():
                if intentar_login(sig, reintentar_una_vez=True):
                    sig.ir(BUSQUEDA_URL)
                    sig.esperar_spinner(raise_on_timeout=False)
            else:
                safe_refresh(sig.driver)
        except Exception: pass
        return
    if intento == 3:
        espera("reintento_3")
        try:
            sig.esperar_spinner(clave_espera="spinner_stuck", raise_on_timeout=False)
            sig.ir(BUSQUEDA_URL) 
            sig.esperar_spinner(raise_on_timeout=False)
        except Exception: pass
        return
    if intento == 4:
        espera("reintento_4")
        try:
            safe_refresh(sig.driver)
            if sig.sesion_cerrada(): intentar_login(sig, reintentar_una_vez=True)
            sig.ir(BUSQUEDA_URL) 
            sig.esperar_spinner(raise_on_timeout=False)
        except Exception: pass
        return

# =============================== ANALIZAR / EXPORT ===================================
# ==============================================================================
# --- MODIFICACIÓN 3: ANALIZAR_MISION ---
# (Añadida lógica para "Fecha Excluyente")
# ==============================================================================
def analizar_mision(sigges: SiggesDriver, m: Dict[str,Any], casos: List[str], fobj: datetime,
                    fecha: str, fall_dt: Optional[datetime], edad_paciente: Optional[int]) -> Dict[str,Any]:
    
    # --- 1. SETUP ---
    if m.get("objetivos"):
        objetivos = [str(c).strip() for c in m.get("objetivos", [])][:MAX_OBJETIVOS]
    else:
        obj = m.get("objetivo")
        objetivos = [str(obj).strip()] if obj else []
    objetivos = objetivos[:MAX_OBJETIVOS]
    frecuencia = str(m.get("frecuencia", "anual")).lower().strip() if m.get("frecuencia") else "anual"
    if frecuencia not in ("mensual", "anual", "vida"):
        frecuencia = "anual"
    edad_min = m.get("edad_min", None)
    edad_max = m.get("edad_max", None)
    excluyentes = m.get("excluyentes", []) or []
    usar_hab = m.get("usar_habilitantes", False)
    cods_hab = m.get("habilitantes", []) or []
    min_habs = m.get("min_habilitantes", 0)

    res = vac_row(m, fecha, "", "")
    res["Edad"] = str(edad_paciente) if edad_paciente is not None else ""

    idx, texto, _ = elegir_caso_mas_reciente(casos, m.get("keywords", []))
    if idx is None:
        res["Observación"] = "Sin caso relevante en mini-tabla"
        return res
    res["Caso Encontrado"] = texto or "Caso relevante"
    
    root = sigges.expandir_caso(idx)
    if not root:
        res["Observación"] = "Error al expandir caso"
        return res

    # --- 2. LECTURA DE DATOS (IPD, OA, PRESTACIONES) ---
    prestaciones = []
    try:
        if REVISAR_IPD:
            f_list, e_list, d_list = sigges.leer_ipd_desde_caso(root, IPD_FILAS_REVISION)
            res["Fecha IPD"]       = " | ".join(f_list)
            res["Estado IPD"]      = " | ".join(e_list)
            res["Diagnóstico IPD"] = " | ".join(d_list) # <-- Clave interna
        
        tb = sigges._prestaciones_tbody(idx)
        # IMPORTANTE: "prestaciones" contiene TODAS las prestaciones (no solo las filtradas)
        prestaciones = sigges.leer_prestaciones_desde_tbody(tb) if tb is not None else []
        
        if REVISAR_OA:
            f_list_oa, p_list_oa, d_list_oa, c_list_oa = sigges.leer_oa_desde_caso(root, OA_FILAS_REVISION) 
            res["Fecha OA"]        = " | ".join(f_list_oa)
            res["Derivado OA"]     = " | ".join(p_list_oa) # <-- Clave interna
            res["Diagnóstico OA"]  = " | ".join(d_list_oa) # <-- Clave interna
            res["Código OA"]       = " | ".join(c_list_oa) # <-- Clave interna
            
    except Exception as e:
        log_error(f"Error leyendo datos (IPD/OA/PREST) del caso {idx}: {e}", tipo="ERROR")
        res["Observación"] = f"Error interno leyendo caso: {e}"
    finally:
        sigges.cerrar_caso_por_indice(idx)
    
    # --- 3. PROCESO DE DATOS ---
    def filtrar_prestaciones_periodo(prest: List[Dict[str,str]]) -> List[Dict[str,Any]]:
        out = []
        for p in prest:
            cod = p.get("codigo", "")
            dt = dparse(p.get("fecha", ""))
            if not dt: continue
            if fobj and dt > fobj: continue
            if not REVISAR_HISTORIA_COMPLETA:
                if frecuencia == "vida":
                    if dt.year != fobj.year: continue
                elif frecuencia == "anual":
                    if dt.year != fobj.year: continue
                elif frecuencia == "mensual":
                    if dt.year != fobj.year or dt.month != fobj.month: continue
            else: pass
            out.append({**p, "dt": dt})
        return out

    prest_filtradas = filtrar_prestaciones_periodo(prestaciones)
    last_obj_dates: List[Optional[datetime]] = []
    count_obj: List[int] = []
    for i, cod in enumerate(objetivos):
        dts = [p["dt"] for p in prest_filtradas if p.get("codigo") == cod]
        dts = [d for d in dts if d and (not fobj or d <= fobj)]
        last_dt = max(dts) if dts else None
        last_obj_dates.append(last_dt)
        col_name = "Objetivo" if i == 0 else f"Objetivo {i+1}" # <-- Clave interna
        if last_dt:
            res[col_name] = last_dt.strftime("%d/%m/%Y")
        cnt = len(dts)
        count_obj.append(cnt)
        
    if objetivos:
        dt_list0 = [p["dt"] for p in prest_filtradas if p.get("codigo") == objetivos[0]]
        res["Mensual"] = mensual_categoria([d for d in dt_list0 if d and same_month(d, fobj)], fobj) if dt_list0 else "Sin Día"
    
    if usar_hab and cods_hab:
        habs = listar_habilitantes(prestaciones, cods_hab, fobj)
        if habs:
            cods = [c for c,_ in habs[:TOP_K_HABILITANTES]]
            fchs = [d.strftime("%d/%m/%Y") if d else "" for _,d in habs[:TOP_K_HABILITANTES]]
            res["Habilitan"] = " | ".join(cods) # <-- Clave interna
            res["Fecha Háb"] = " | ".join(fchs) # <-- Clave interna
            res["Hab Vi"]    = "Sí" if any(en_vig(fobj, d) for _,d in habs) else "No" # <-- Clave interna (se queda por si acaso)
        else:
            res["Hab Vi"]    = "No"

    # --- 4. LÓGICA DE OBSERVACIONES (MEJORADA) Y EXCLUYENTE ---
    obs_tags: List[str] = []

    # 4.1. Críticos (Fallecido, Edad)
    if fall_dt:
        obs_tags.append(f"Fallecido {fall_dt.strftime('%d/%m/%Y')}")

    if edad_paciente is not None:
        if edad_min is not None and edad_paciente < edad_min:
            obs_tags.append(f"EDAD: {edad_paciente} (Req: {edad_min}+)")
        if edad_max is not None and edad_paciente > edad_max:
            obs_tags.append(f"EDAD: {edad_paciente} (Req: {edad_max}-)")
    
    # --- 4.2. LÓGICA EXCLUYENTE (CON FECHA) ---
    if excluyentes:
        excl_found_list = [] # Lista para (datetime, code, date_str)
        # Usamos TODAS las prestaciones (sin filtrar por fecha) para buscar excluyentes
        for p in prestaciones: 
            cod = p.get("codigo")
            if cod in excluyentes:
                dt = dparse(p.get("fecha", ""))
                if dt:
                    # Guardamos la fecha-objeto, el código, y el string de fecha
                    excl_found_list.append((dt, cod, p.get("fecha", "")))
        
        if excl_found_list:
            # Ordenamos por fecha (más nueva primero)
            excl_found_list.sort(key=lambda x: x[0], reverse=True) 
            # Guardamos los códigos y fechas en sus columnas
            res["Excluyente"] = " | ".join([item[1] for item in excl_found_list])
            res["Fecha Excluyente"] = " | ".join([item[2] for item in excl_found_list])
            
    # 4.3. IPD
    if REVISAR_IPD:
        ipd_f = res.get("Fecha IPD", "")
        if "no tiene" in ipd_f.lower():
            obs_tags.append("IPD: No Tiene")
        elif not ipd_f:
            obs_tags.append("IPD: Sin Info") # No se encontró la tabla

    # 4.4. OA
    if REVISAR_OA:
        oa_f = res.get("Fecha OA", "")
        if "no tiene" in oa_f.lower():
            obs_tags.append("OA: No Tiene")
        elif not oa_f:
            obs_tags.append("OA: Sin Info") # No se encontró la tabla

    # 4.5. Habilitantes
    if usar_hab and cods_hab:
        if not res.get("Habilitan"):
            obs_tags.append("Habilitante: No Encontrado")
        elif res.get("Hab Vi") == "No":
            obs_tags.append("Habilitante: Expirado")
        
        # Revisión de Mínimos
        if min_habs and min_habs > 0:
            habs_all = listar_habilitantes(prestaciones, cods_hab, fobj)
            num_habs_encontrados = len(habs_all)
            if num_habs_encontrados < min_habs:
                obs_tags.append(f"Habilitante: Faltan (Req: {min_habs})")

    # 4.6. Objetivos
    if objetivos:
        main_obj_date = last_obj_dates[0] if last_obj_dates else None
        if not main_obj_date:
            obs_tags.append("Objetivo: No Encontrado")
        else:
            cnt = count_obj[0]
            freq_name = frecuencia.capitalize()
            if cnt > 0 and frecuencia != "vida":
                 obs_tags.append(f"Objetivo: {cnt} en {freq_name}")
    
    # 4.7. Objetivos secundarios
    for i, cnt in enumerate(count_obj[1:], 1): # Empezar desde el segundo objetivo
        if cnt > 0:
            freq_name = frecuencia.capitalize()
            obj_label = f"Objetivo {i+1}"
            if frecuencia != "vida":
                obs_tags.append(f"{obj_label}: {cnt} en {freq_name}")
    
    # --- 5. ENSAMBLE FINAL DE OBSERVACIÓN ---
    if obs_tags:
        res["Observación"] = join_tags(obs_tags)
    else:
        # Si no hay tags, es un caso "Limpio"
        if res["Observación"] != "Sin caso relevante en mini-tabla":
             res["Observación"] = "Revisado OK" # Tag positivo
            
    # --- LÓGICA DE "SEGUIMIENTO" (seg_tags) COMPLETAMENTE ELIMINADA ---
    
    return res

def short_err_name(e: Exception) -> str:
    # (Sin cambios)
    mapping = [(TimeoutException, "Fallo de Carga"),
               (ElementClickInterceptedException, "Fallo de Clickeo"),
               (ElementNotInteractableException, "Fallo de Clickeo"),
               (StaleElementReferenceException, "DOM Inestable"),
               (WebDriverException, "WebDriver"),
               (NoSuchWindowException, "Ventana Cerrada"),
               (SpinnerStuck, "Spinner Pegado")]
    for typ, name in mapping:
        if isinstance(e, typ): return name
    return "Error General"

def resumen_paciente(i: int, total: int, nombre: str, rut: str, fecha: str, 
                     flags: Dict[str,bool], resultados: List[Dict[str, Any]]) -> None:
    # (Sin cambios, esta función no usaba "Seguimiento")
    
    now = datetime.now().strftime("%H:%M") 
    b = C_BARRA + "|" + RESET 

    m1 = resultados[0] if len(resultados) > 0 else {}
    m2 = resultados[1] if len(resultados) > 1 else {}
    
    paciente_ok = flags.get("ok", False)
    paciente_saltado = flags.get("saltado", False)

    # --- M1 (Artrosis) ---
    m1_mini_val = "Sí" if (m1.get("Caso Encontrado") or "Sin caso") != "Sin caso" else "No"
    m1_mini_col = C_SI if m1_mini_val == "Sí" else C_NO
    
    m1_hab_val_raw = m1.get("Habilitan", "")
    m1_hab_val_display = m1_hab_val_raw.split('|')[0].strip() or "No" 
    m1_hab_col  = C_SI if m1_hab_val_display != "No" else C_NO
    
    m1_ipd_val  = (m1.get("Estado IPD", "").split("|")[0].strip() or "Sin IPD")
    if "no tiene" in m1_ipd_val.lower(): m1_ipd_val = "Sin IPD"
    m1_ipd_col  = C_SI if m1_ipd_val != "Sin IPD" else C_NO
    
    m1_oa_val   = "Con OA" if (m1.get("Fecha OA") and "no tiene" not in m1.get("Fecha OA", "").lower()) else "Sin OA"
    m1_oa_col   = C_SI if m1_oa_val == "Con OA" else C_NO
    
    # --- M2 (Endoprótesis) ---
    m2_mini_val = "Sí" if (m2.get("Caso Encontrado") or "Sin caso") != "Sin caso" else "No"
    m2_mini_col = C_SI if m2_mini_val == "Sí" else C_NO

    m2_hab_val_raw = m2.get("Habilitan", "")
    m2_hab_val_display = m2_hab_val_raw.split('|')[0].strip() or "No"
    m2_hab_col  = C_SI if m2_hab_val_display != "No" else C_NO
    
    m2_ipd_val  = (m2.get("Estado IPD", "").split("|")[0].strip() or "Sin IPD")
    if "no tiene" in m2_ipd_val.lower(): m2_ipd_val = "Sin IPD"
    m2_ipd_col  = C_SI if m2_ipd_val != "Sin IPD" else C_NO

    m2_oa_val   = "Con OA" if (m2.get("Fecha OA") and "no tiene" not in m2.get("Fecha OA", "").lower()) else "Sin OA"
    m2_oa_col   = C_SI if m2_oa_val == "Con OA" else C_NO

    # --- Partes de Impresión (Formato Cascada R6) ---
    
    part1 = (
        f"🔥 {C_INDICE}[{i}/{total}]{RESET} 🔥 {b} "
        f"⏳ {C_HORA}{now}{RESET} ⏳ {b} "
        f"🤹🏻  {C_NOMBRE}{nombre.upper()}{RESET} 🤹🏻 {b} "
        f"🪪  {C_RUT}{rut}{RESET} 🪪  {b}"
        f"🗓️  {C_FECHA}{fecha}{RESET} 🗓️" 
    )
    
    part2 = (
        f"📋 {C_M1_LABEL}M1-Mini: {m1_mini_col}{m1_mini_val}{RESET} 📋 {b} "
        f"📋 {C_M2_LABEL}M2-Mini: {m2_mini_col}{m2_mini_val}{RESET} 📋 {b} "
        f"♦️ {C_M1_LABEL} M1-Hab: {m1_hab_col}{m1_hab_val_display}{RESET} ♦️  {b}  "
        f"♦️  {C_M2_LABEL} M2-Hab: {m2_hab_col}{m2_hab_val_display}{RESET} ♦️"
    )
    
    part3 = (
        f"🔶 {C_M1_LABEL}M1-IPD: {m1_ipd_col}{m1_ipd_val}{RESET} 🔶 {b} "
        f"🔶 {C_M2_LABEL}M2-IPD: {m2_ipd_col}{m2_ipd_val}{RESET} 🔶 {b} "
        f"🔷 {C_M1_LABEL}M1-OA: {m1_oa_col}{m1_oa_val}{RESET} 🔷 {b} "
        f"🔷 {C_M2_LABEL}M2-OA: {m2_oa_col}{m2_oa_val}{RESET} 🔷"
    )
    
    # --- Línea de Estado Final (Lógica R9) ---
    def _get_status_for_mission(res: Dict[str, Any]) -> Tuple[str, str]:
        if not res: # Si la misión no tiene datos (ej. script interrumpido)
            return ("⚠️  Sin Datos ⚠️", C_NARANJA)
        
        mini_found = (res.get("Caso Encontrado") or "Sin caso") != "Sin caso"
        obs_txt = res.get("Observación", "")
        
        # --- Lógica de Observaciones MEJORADA (R10) ---
        # Excluyentes, Edad, o Fallecido son OBSERVACIONES CRÍTICAS
        # --- MODIFICACIÓN: AHORA BUSCA "Excluyente" EN SU PROPIA CELDA ---
        excl_txt = res.get("Excluyente", "") 
        obs_critica = (excl_txt or "EDAD:" in obs_txt or "Fallecido" in obs_txt)

        if not mini_found:
            # Usamos C_NARANJA para "Sin Caso" (coincide con ⚠️)
            return ("⚠️ Sin Caso ⚠️", C_NARANJA) 
        if obs_critica:
            return ("⚠️ Observaciones ⚠️", C_NARANJA)
        
        # Todo lo demás (incluyendo "Sin IPD", "Sin Habilitante", "Revisado OK") es Éxito.
        return ("✅  Éxito ✅", C_EXITO)

    status_msg = ""
    status_col = "" # Para el fallback
    part4 = ""      # La línea de estado final

    if paciente_saltado:
        status_msg = f"♻️ Paciente Saltado Automáticamente ({MAX_REINTENTOS_GRAVES_POR_PACIENTE} Reintentos) ♻️"
        status_col = C_ROJO 
        part4 = f"{status_col}{status_msg}{RESET}"
    elif not paciente_ok:
        status_msg = "❌ Paciente No Revisado (Error Inesperado) ❌"
        status_col = C_FALLO 
        part4 = f"{status_col}{status_msg}{RESET}"
    else:
        # Obtenemos estados individuales
        m1_status, m1_col = _get_status_for_mission(m1)
        m2_status, m2_col = _get_status_for_mission(m2)
        
        # Construimos el string con tu formato exacto
        # M1 (Fucsia), M2 (Cyan)
        part4 = (
            f"{C_ROJO}Paciente:{RESET} "
            f"{C_FUCSIA}Mision 1{RESET} - {m1_col}{m1_status}{RESET} {b} "
            f"{C_CYAN}Mision 2{RESET} - {m2_col}{m2_status}{RESET}"
        )
        
        # Determinar el estado de fallback (para el print sin colores)
        if m1_col == C_EXITO and m2_col == C_EXITO:
            status_col = C_EXITO
        else: # (Observaciones o Sin Caso)
            status_col = C_NARANJA

    try:
        print(f"{part1}") 
        print() 
        print(f"{part2}")
        print() 
        print(f"{part3}")
        print() 
        print(part4) # <-- part4 ahora es la línea de estado R9
        print(f"{C_BARRA}{'-' * 80}{RESET}")
    except Exception:
        # Lógica de fallback actualizada para R9
        fallback_status = "ERROR"
        if paciente_saltado: 
            fallback_status = "SALTADO"
        elif not paciente_ok: 
            fallback_status = "ERROR"
        else:
            if status_col == C_EXITO:
                fallback_status = "EXITO"
            elif status_col == C_NARANJA:
                fallback_status = "OBS/SIN_CASO"
            
        print(f"[{i}/{total}] {now} | {nombre} | {rut} | F:{fecha} | ESTADO: {fallback_status}")

# ==============================================================================
# --- FUNCIÓN NUEVA: APLICAR ESTILOS DE EXCEL ---
# (Añadida "Fecha Excluyente" al mapa de colores, Fuentes verificadas como "bold=True")
# ==============================================================================
def aplicar_estilos_excel(writer, sheet_name: str, new_headers: List[str]):
    """
    Aplica los colores y estilos solicitados a los encabezados de la hoja de Excel.
    """
    try:
        # Definición de Estilos (Colores ARGB)
        # Fuentes (TODAS EN NEGRIYTA)
        FONT_BLANCO_B = Font(color="FFFFFF", bold=True)
        FONT_NEGRO_B = Font(color="000000", bold=True)
        
        # Rellenos (Fondo)
        FILL_ROJO = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid") # IPD
        FILL_AZUL = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid") # OA
        FILL_AMARILLO = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # Habilitantes
        FILL_VERDE = PatternFill(start_color="008000", end_color="008000", fill_type="solid") # Paciente (Verde oscuro)
        FILL_NARANJA = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid") # Observaciones, etc.
        FILL_MORADO = PatternFill(start_color="800080", end_color="800080", fill_type="solid") # Caso, etc.
        
        # Mapeo de estilos por grupo
        estilos = {
            "ipd": {"fill": FILL_ROJO, "font": FONT_BLANCO_B},
            "oa": {"fill": FILL_AZUL, "font": FONT_BLANCO_B},
            "hab": {"fill": FILL_AMARILLO, "font": FONT_NEGRO_B},
            "pac": {"fill": FILL_VERDE, "font": FONT_BLANCO_B},
            "obs": {"fill": FILL_NARANJA, "font": FONT_NEGRO_B},
            "caso": {"fill": FILL_MORADO, "font": FONT_BLANCO_B},
        }
        
        # Mapeo de CADA encabezado a su grupo de estilo
        # (Usamos los nombres FINALES de 'cols_mision')
        mapa_colores = {
            # Paciente (Verde)
            "Fecha": estilos["pac"],
            "Rut": estilos["pac"],
            "Edad": estilos["pac"],
            
            # Caso (Morado)
            "Caso Encontrado": estilos["caso"],
            "Familia": estilos["caso"],
            "Especialidad": estilos["caso"],
            
            # Observaciones/Objetivos (Naranja)
            "Cod Obj": estilos["obs"],
            "Fecha Obj": estilos["obs"],
            "Mensual": estilos["obs"],
            "Excluyente": estilos["obs"],
            "Fecha Excluyente": estilos["obs"], # <-- NUEVA COLUMNA AÑADIDA
            "Observación": estilos["obs"],
            
            # Habilitantes (Amarillo)
            "Habilitante": estilos["hab"],
            "Fecha Hab": estilos["hab"],
            
            # IPD (Rojo)
            "Fecha IPD": estilos["ipd"],
            "Estado IPD": estilos["ipd"],
            "Diag IPD": estilos["ipd"],
            
            # OA (Azul)
            "Fecha OA": estilos["oa"],
            "Cod OA": estilos["oa"],
            "Deri OA": estilos["oa"],
            "Diag OA": estilos["oa"],
        }
        
        # Mapeo para objetivos extra (ej. "Objetivo 2")
        for h in new_headers:
            if "Objetivo" in h and h != "Fecha Obj": # "Objetivo 2", "Objetivo 3"...
                mapa_colores[h] = estilos["obs"]

        # Obtener la hoja de trabajo (worksheet) de openpyxl
        ws = writer.sheets[sheet_name]
        
        # Centrar todo
        alignment_center = Alignment(horizontal='center', vertical='center')

        # Recorrer la PRIMERA FILA (los encabezados)
        for cell in ws[1]:
            if cell.value in mapa_colores:
                estilo = mapa_colores[cell.value]
                cell.fill = estilo["fill"]
                cell.font = estilo["font"]
                cell.alignment = alignment_center

        # Ajustar anchos de columna (un pequeño bonus)
        for i, col_name in enumerate(new_headers, 1):
            letter = chr(64 + i) # A, B, C...
            width = len(col_name) + 4 # Ancho basado en el encabezado
            if col_name in ["Caso Encontrado", "Observación", "Diag IPD", "Diag OA"]:
                width = 30 # Columnas anchas
            elif col_name in ["Familia", "Especialidad", "Habilitante", "Excluyente", "Deri OA", "Fecha Excluyente"]:
                width = 20
            elif col_name in ["Fecha", "Rut", "Edad", "Cod Obj", "Fecha Obj", "Cod OA"]:
                width = 15
            ws.column_dimensions[letter].width = width

    except Exception as e:
        # Si falla el coloreo, no es crítico. Solo informamos.
        log_error(f"No se pudieron aplicar los estilos de color al Excel: {e}", tipo="WARN")

# ==============================================================================
# --- MODIFICACIÓN 4: EJECUTAR_REVISION ---
# (Lógica de renombrado y reordenado verificada como correcta)
# ==============================================================================
def ejecutar_revision()->None:
    # Driver
    opts = webdriver.EdgeOptions(); opts.debugger_address = DIRECCION_DEBUG_EDGE
    try: 
        driver = webdriver.Edge(service=Service(EDGE_DRIVER_PATH), options=opts)
        driver.set_page_load_timeout(ESPERAS["page_load"]["wait"])
    except Exception as e: 
        print(f"{ERR}Error conectando a Edge: {e}{RESET}")
        return
        
    sigges = SiggesDriver(driver)

    # Datos
    try: df = pd.read_excel(RUTA_ARCHIVO_ENTRADA)
    except Exception as e: 
        print(f"{ERR}Error leyendo Excel: {e}{RESET}")
        if CERRAR_NAVEGADOR_AL_FINAL: driver.quit()
        return

    filas_por_mision = defaultdict(list)
    total=len(df)
    log_eventos: List[Dict[str,str]] = []
    n_ok = n_skip = n_warn = 0

    for idx,row in df.iterrows():
        if (idx>0) and (idx % 60 == 0):
            try: gc.collect()
            except Exception: pass

        nombre = str(row.iloc[INDICE_COLUMNA_NOMBRE]).strip()
        rut    = normalizar_rut(str(row.iloc[INDICE_COLUMNA_RUT]).strip())
        fecha  = solo_fecha(row.iloc[INDICE_COLUMNA_FECHA]); fobj = dparse(fecha)

        intento_actual = 0
        paciente_resuelto = False
        paciente_saltado = False 
        hallo_mini = hallo_cartola = False
        resultados_misiones_paciente: List[Dict[str, Any]] = []

        while intento_actual < MAX_REINTENTOS_GRAVES_POR_PACIENTE and not paciente_resuelto:
            try:
                intento_actual += 1
                
                sigges.asegurar_en_busqueda()

                el = sigges._find([XP_RUT_INPUT], "presence", "find_rut")
                if not el: raise TimeoutException("Campo RUT no encontrado")
                
                try: el.clear()
                except StaleElementReferenceException:
                    el = sigges._find([XP_RUT_INPUT], "presence", "find_rut"); el.clear()
                
                espera("clear_rut") 
                el.send_keys(rut)
                espera("send_rut")  

                ok = sigges._click([XP_BTN_BUSCAR], scroll=False, wait_spinner=True,
                                   clave_espera="click_buscar", 
                                   spinner_clave_espera="spinner", 
                                   raise_on_spinner_timeout=True)
                if not ok: raise TimeoutException("Botón Buscar no clickeable")

                mini = leer_mini_tabla_busqueda(sigges.driver, stable_reads=MINI_TABLA_STABLE_READS)
                
                hallo_mini = any(has_keyword(r.get("problema",""), m.get("keywords", [])) for r in (mini or []) for m in MISSIONS)
                
                if not hallo_mini:
                    add_vacios(filas_por_mision, MISSIONS, fecha, rut, "Sin caso relevante en mini-tabla")
                    for m in MISSIONS:
                        resultados_misiones_paciente.append(vac_row(m, fecha, rut, "Sin caso relevante en mini-tabla"))
                    n_ok += 1
                    paciente_resuelto = True
                    break 

                try: 
                    edad_paciente = sigges.leer_edad()
                except Exception: 
                    edad_paciente = None

                if not sigges.ir_a_cartola(raise_on_spinner_timeout=True):
                    raise TimeoutException("No se pudo abrir Cartola")
                
                sigges.activar_hitos_ges()
                hallo_cartola = True 

                fall_dt = sigges.leer_fallecimiento()
                casos = sigges.lista_de_casos_cartola()

                for i_m, m in enumerate(MISSIONS):
                    # analizar_mision devuelve el dict con las claves INTERNAS
                    r = analizar_mision(sigges, m, casos, fobj, fecha, fall_dt, edad_paciente)
                    r.update({"Fecha": fecha, "Rut": rut})
                    if edad_paciente is not None:
                        r["Edad"] = str(edad_paciente)
                    filas_por_mision[i_m].append(r)
                    resultados_misiones_paciente.append(r) 

                paciente_resuelto = True; n_ok += 1

            except (TimeoutException, StaleElementReferenceException, ElementClickInterceptedException, ElementNotInteractableException, SpinnerStuck, WebDriverException) as e:
                error_txt = short_err_name(e)
                log_error(f"Error {error_txt} con {rut}. Reintentando ({intento_actual}/{MAX_REINTENTOS_GRAVES_POR_PACIENTE})...", tipo="WARN")
                if intento_actual >= 2:
                    log_eventos.append({"hora": datetime.now().strftime("%H:%M:%S"), "rut":rut, "nombre":nombre,
                                        "paso":"General","intento":str(intento_actual),"estado":error_txt,
                                        "mision":"*","obs":"error recuperable (>=2 intentos)"})
                if intento_actual >= MAX_REINTENTOS_GRAVES_POR_PACIENTE: break
                manejar_reintentos_inteligentes(sigges, intento_actual)

            except KeyboardInterrupt:
                log_error("Interrumpido manualmente (Ctrl+C).", tipo="ERROR")
                if CERRAR_NAVEGADOR_AL_FINAL: driver.quit()
                return

            except Exception as e:
                error_nombre = short_err_name(e)
                log_error(f"Error Inesperado ({error_nombre}) con {rut}. Reintentando ({intento_actual}/{MAX_REINTENTOS_GRAVES_POR_PACIENTE})...", tipo="WARN")
                if intento_actual >= 2:
                    log_eventos.append({"hora": datetime.now().strftime("%H:%M:%S"), "rut":rut, "nombre":nombre,
                                        "paso":"General","intento":str(intento_actual),"estado":error_nombre,
                                        "mision":"*","obs":f"error inesperado ({e})"})
                if intento_actual >= MAX_REINTENTOS_GRAVES_POR_PACIENTE: break
                manejar_reintentos_inteligentes(sigges, intento_actual)

        if not paciente_resuelto:
            paciente_saltado = True 
            add_vacios(filas_por_mision, MISSIONS, fecha, rut, "Paciente Saltado tras reintentos")
            for m in MISSIONS:
                resultados_misiones_paciente.append(vac_row(m, fecha, rut, "Paciente Saltado"))
            log_eventos.append({"hora": datetime.now().strftime("%H:%M:%S"), "rut":rut, "nombre":nombre,
                                "paso":"Final","intento":str(intento_actual),"estado":"Saltado",
                                "mision":"*","obs":"agotó reintentos"})
            n_skip += 1

        resumen_paciente(idx+1, total, nombre, rut, fecha,
                         {"mini":hallo_mini, "cartola":hallo_cartola, "ok":paciente_resuelto, "saltado": paciente_saltado},
                         resultados_misiones_paciente)
        
        espera("entre_pacientes") 

    # Exportar Excel y Log
    os.makedirs(RUTA_CARPETA_SALIDA, exist_ok=True)
    stamp = datetime.now().strftime("%d-%m-%Y_%H-%M")
    out_xlsx = os.path.join(RUTA_CARPETA_SALIDA, f"Rev_{NOMBRE_DE_LA_MISION}_{stamp}.xlsx")
    try:
        with pd.ExcelWriter(out_xlsx, engine="openpyxl") as w:
            for i_m, m in enumerate(MISSIONS):
                
                # 1. Obtiene la lista de columnas (nuevo orden, nuevos nombres)
                col_list = cols_mision(m)
                
                # 2. Crea el DF con los datos (claves internas)
                df_out = pd.DataFrame(filas_por_mision[i_m])
                
                # 3. Mapa de renombre: "Clave Interna" -> "Nombre Final en Excel"
                rename_map = {
                    "Cod Objetivo": "Cod Obj",
                    "Objetivo": "Fecha Obj",
                    "Habilitan": "Habilitante",
                    "Fecha Háb": "Fecha Hab",
                    "Diagnóstico IPD": "Diag IPD",
                    "Código OA": "Cod OA",
                    "Derivado OA": "Deri OA",
                    "Diagnóstico OA": "Diag OA"
                }
                df_out.rename(columns=rename_map, inplace=True)
                
                # 4. Re-indexa el DF para ORDENAR y FILTRAR según col_list
                # (Esto elimina Hab Vi y Seguimiento, y ordena todo)
                df_out = df_out.reindex(columns=col_list)
                
                # 5. Escribe en la hoja
                sheet_name = f"Mision {i_m+1}"
                df_out.to_excel(w, index=False, sheet_name=sheet_name)
                
                # 6. Llama a la nueva función de estilos
                aplicar_estilos_excel(w, sheet_name, col_list)

            # Carga Masiva (sin cambios)
            pd.DataFrame([], columns=["FECHA","RUT","DV","PRESTACIONES","TIPO","PS-FAM","ESPECIALIDAD"]).to_excel(w, index=False, sheet_name="Carga Masiva")
            
        print(f"{OK}Archivo generado (con colores!): {out_xlsx}{RESET}")
    except Exception as e:
        print(f"{ERR}Error generando archivo de salida: {e}{RESET}")

    # Log (sin cambios)
    log_path = os.path.join(RUTA_CARPETA_SALIDA, f"Log_{NOMBRE_DE_LA_MISION}_{stamp}.txt")
    try:
        with open(log_path, "w", encoding="utf-8") as f:
            f.write(f"LOG SIGGES — {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write("="*80 + "\n")
            f.write(f"Misión: {NOMBRE_DE_LA_MISION}\n")
            f.write(f"Procesados OK: {n_ok} | Saltados: {n_skip} | Con advertencias (>=2 intentos): {len(log_eventos)}\n\n")
            for it in log_eventos:
                linea = (f"[{it.get('hora','')}] Rut:{it.get('rut','')} Nombre:{it.get('nombre','')} "
                         f"Paso:{it.get('paso','')} Intento:{it.get('intento','')} "
                         f"Estado:{it.get('estado','')} Misión:{it.get('mision','')} "
                         f"Obs:{it.get('obs','')}\n")
                f.write(linea)
        print(f"{OK}Log guardado en: {log_path}{RESET}")
    except Exception as e:
        print(f"{ERR}No se pudo escribir log: {e}{RESET}")

    
    if CERRAR_NAVEGADOR_AL_FINAL:
        driver.quit()
    else:
        print(f"{INFO}Ejecución finalizada. El navegador (modo debug) sigue abierto.{RESET}")

if __name__=="__main__":
    ejecutar_revision()