from __future__ import annotations 
import os, re, time, unicodedata, gc
from datetime import datetime, timedelta
from typing import List, Optional, Dict, Tuple, Any
from collections import defaultdict

import pandas as pd
import warnings
warnings.filterwarnings("ignore", category=UserWarning)

# Para pintar el Excel al final
try:
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    # Definición dummy para evitar errores si openpyxl no está disponible
    def aplicar_estilos_excel(ws): pass
    pass 

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException, ElementClickInterceptedException, ElementNotInteractableException,
    StaleElementReferenceException, WebDriverException, NoSuchWindowException
)

# ==========================================================================================
#                                  PANEL DE CONTROL GENERAL
# ==========================================================================================

NOMBRE_DE_LA_MISION = "AIJ"
RUTA_ARCHIVO_ENTRADA = r"C:\Users\usuariohgf\OneDrive\Documentos\Trabajo Oficina\Trabajo\AIJ Nov 1.xlsx"
RUTA_CARPETA_SALIDA  = r"C:\Users\usuariohgf\OneDrive\Documentos\Trabajo Oficina\Trabajo\Revisiones, Falta Enviar, CM\Revisiones\Revisiones Nómincas"
DIRECCION_DEBUG_EDGE = "localhost:9222"
EDGE_DRIVER_PATH     = r"C:\\Windows\\System32\\msedgedriver.exe"

# --- Configuración Excel (Columnas Entrada 0-based) ---
INDICE_COLUMNA_FECHA  = 5
INDICE_COLUMNA_RUT    = 1
INDICE_COLUMNA_NOMBRE = 3

# --- Reglas de Negocio ---
VENTANA_VIGENCIA_DIAS = 90
MAX_REINTENTOS_GRAVES_POR_PACIENTE = 4

# --- Toggles de Columnas y Lógica ---
REVISAR_HISTORIA_COMPLETA: bool = False
MAX_OBJETIVOS: int = 5
MINI_TABLA_STABLE_READS = 2 
TOP_K_HABILITANTES = 3

REVISAR_IPD = False      # Si es False, no aparecen columnas IPD
REVISAR_OA  = True       # Si es False, no aparecen columnas OA
IPD_FILAS_REVISION = 1
OA_FILAS_REVISION  = 1

# Misiones
MISSIONS: List[Dict[str, Any]] = [
    {
        "nombre": "Artritis Idiopática Juvenil",
        "keywords": ["artritis idiopatica", "artritis idiopática", "idiopatica", "idiopática", "juvenil"],
        "objetivo": "2505290",
        "usar_habilitantes": False,
        "habilitantes": ["9001043", "0305091"],
        "familia": "63",
        "especialidad": "13-13-01",
        "frecuencia": "Mensual",
        "edad_min": None,
        "edad_max": None,
        "excluyentes": [], 
        "min_habilitantes": 1
    },
]

# ==========================================================================================
#                             PANEL DE CONTROL DE DIRECCIONES (XPATHs)
# ==========================================================================================

XPATHS = {
    # --- Navegación General ---
    "BASE_URL": "https://www.sigges.cl",
    "BUSQUEDA_URL": "https://www.sigges.cl/#/busqueda-de-paciente",
    "CARTOLA_URL": "https://www.sigges.cl/#/cartola-unificada-de-paciente",

    # --- Login ---
    "LOGIN_BTN_INGRESAR": [
        "//button[@type='submit' and contains(@class,'botonBase')][.//p[normalize-space(.)='Ingresar']]",
        "/html/body/div/div/div[2]/div[1]/form/div[3]/button",
    ],
    "LOGIN_SEL_UNIDAD_HEADER": [
        "//div[contains(@class,'filtroSelect__header')][.//p[contains(translate(normalize-space(.),'ÁÉÍÓÚ','áéíóú'),'seleccione una unidad')]]",
        "/html/body/div/div/div/div/div[1]/div/div[2]/div[1]/div/div[2]/div[1]",
    ],
    "LOGIN_OP_HOSPITAL": [
        "//div[contains(@class,'filtroSelect__option')][.//p[contains(translate(normalize-space(.),'ÁÉÍÓÚ','áéíóú'),'gustavo fricke')]]",
        "/html/body/div/div/div/div/div[1]/div/div[2]/div[1]/div/div[2]/div[2]/div",
    ],
    "LOGIN_TILE_INGRESO_SIGGES": [
        "/html/body/div/div/div/div/div[1]/div/div[2]/div[2]/div[2]/div[3]",
        "//div[contains(@class,'boxItems__item')][.//p[normalize-space(.)='INGRESO SIGGES CONFIDENCIAL']]"
    ],
    "LOGIN_BTN_CONECTAR": [
        "//button[contains(@class,'botonBase')][.//p[contains(translate(normalize-space(.),'ÁÉÍÓÚ','áéíóú'),'conectese')]]",
        "/html/body/div/div/div/div/div[1]/div/div[2]/div[3]/button",
    ],

    # --- Menú Lateral (Detector Inteligente) ---
    "MENU_CONTENEDOR": "/html/body/div/main/div[2]/nav/div[1]", 
    "MENU_ICONO_APERTURA": "/html/body/div/main/div[2]/nav/div[1]/div[1]", 
    "MENU_CLASS_ABIERTO": "cardOpen", 
    
    # --- Búsqueda Paciente ---
    "INPUT_RUT": "/html/body/div/main/div[3]/div[3]/div/div/div[1]/div/div/div/input",
    "BTN_BUSCAR": "/html/body/div/main/div[3]/div[3]/div/div/div[1]/div/button",
    "EDAD_PACIENTE": "/html/body/div/main/div[3]/div[3]/div/div[1]/div[7]/div[2]/div/p",
    
    # --- Cartola / Mini Tabla ---
    "CONT_CARTOLA": "/html/body/div/main/div[3]/div[2]/div[1]/div[5]/div[1]/div[2]",
    "CHK_HITOS_GES": "/html/body/div/main/div[3]/div[2]/div[1]/div[5]/div[1]/div/div/label/input",
    "MINI_TABLA_TBODY": [
        "//div[@class='contBody maxW scroll']//div[contains(@class,'cardTable')]/table/tbody",
        "/html/body/div/main/div[3]/div[3]/div/div[2]/div/div/table/tbody",
    ],
    
    # --- Navegación Directa ---
    "BTN_MENU_CARTOLA": "/html/body/div/main/div[2]/nav/div[1]/div[2]/a[2]",
    
    # --- Tablas Internas (Respaldo Absoluto) ---
    "OA_TBODY_FALLBACK": [
        "/html/body/div/main/div[3]/div[2]/div[1]/div[5]/div[1]/div[2]/div/div[5]/div[2]/div/table/tbody", # Genérico
        "/html/body/div/main/div[3]/div[2]/div[1]/div[5]/div[1]/div[2]/div[2]/div[5]/div[2]/div/table/tbody"  # Variante común
    ],
    "IPD_TBODY_FALLBACK": [
        "/html/body/div/main/div[3]/div[2]/div[1]/div[5]/div[1]/div[2]/div[2]/div[4]/div[2]/div/table/tbody",
        "/html/body/div/main/div[3]/div[2]/div[1]/div[5]/div[1]/div[2]/div[7]/div[4]/div[2]/div/table/tbody",
    ],
    "SPINNER_CSS": "dialog.loading",
}

# ==========================================================================================
#                                  PANEL DE ESPERAS
# ==========================================================================================

ESPERAS: Dict[str, Dict[str, float]] = {
    "page_load":           {"wait": 1, "sleep": 0.5},
    "default":             {"wait": 1, "sleep": 1}, 
    "spinner":             {"wait": 5.0, "sleep": 1}, 
    "spinner_largo":       {"wait": 120,"sleep": 3}, 
    "spinner_stuck":       {"wait": 300,"sleep": 3},
    "inicio_driver":       {"wait": 1,  "sleep": 0.5},
    "nav_a_busqueda":      {"wait": 1,  "sleep": 0.5}, 
    "nav_a_busqueda_fast": {"wait": 1,  "sleep": 0.5}, 
    "menu_lateral":        {"wait": 3,  "sleep": 1.0}, 
    "login_check":         {"wait": 2,  "sleep": 0.5},
    "login_ingresar":      {"wait": 2,  "sleep": 0.5},
    "login_sel_unidad":    {"wait": 2,  "sleep": 0.5},
    "login_op_hospital":   {"wait": 2,  "sleep": 0.5},
    "login_ingreso_sigges":{"wait": 2,  "sleep": 0.5},
    "login_conectar":      {"wait": 2,  "sleep": 1.0},
    "find_rut":            {"wait": 1, "sleep": 0.3},
    "clear_rut":           {"wait": 1,  "sleep": 0.3},
    "send_rut":            {"wait": 1,  "sleep": 0.3},
    "click_buscar":        {"wait": 2, "sleep": 0.5},
    "mini_tabla_read":     {"wait": 3,  "sleep": 3}, 
    "leer_edad":           {"wait": 2,  "sleep": 0.5},
    "click_ir_a_cartola":  {"wait": 2, "sleep": 0.5},
    "click_activar_hitos": {"wait": 2, "sleep": 0.3},
    "cont_cartola":        {"wait": 2, "sleep": 0.8},
    "leer_fallecimiento":  {"wait": 2,  "sleep": 0.3},
    "expandir_caso":       {"wait": 2,  "sleep": 0.5}, 
    "prestaciones_tbody":  {"wait": 5,  "sleep": 0.5}, 
    "ipd_tbody":           {"wait": 5,  "sleep": 0.5}, 
    "oa_tbody":            {"wait": 5,  "sleep": 0.5},
    "entre_pacientes":     {"wait": 0.3,  "sleep": 0.3}, 
}

# =============================== COLORES / CONSOLA ===================================
try:
    from colorama import Fore, Style, init as colorama_init
    colorama_init(autoreset=True)
    OK, WARN, INFO, ERR, DATA, CODE, RESET = (
        Fore.LIGHTGREEN_EX, Fore.LIGHTYELLOW_EX, Fore.LIGHTCYAN_EX,
        Fore.LIGHTRED_EX, Fore.LIGHTMAGENTA_EX, Fore.LIGHTBLUE_EX, Style.RESET_ALL
    )
    C_BARRA, C_INDICE, C_HORA, C_NOMBRE, C_RUT, C_FECHA = Fore.LIGHTYELLOW_EX, Fore.LIGHTCYAN_EX, Fore.LIGHTRED_EX, Fore.LIGHTGREEN_EX, Fore.LIGHTMAGENTA_EX, Fore.LIGHTYELLOW_EX
    C_CYAN, C_FUCSIA, C_VERDE, C_ROJO, C_NARANJA = Fore.LIGHTCYAN_EX, Fore.LIGHTMAGENTA_EX, Fore.LIGHTGREEN_EX, Fore.LIGHTRED_EX, Fore.LIGHTYELLOW_EX
    C_M1_LABEL, C_M2_LABEL = C_CYAN, C_FUCSIA
    C_EXITO, C_SIN_CASO, C_FALLO, C_SI, C_NO = C_VERDE, C_NARANJA, C_ROJO, C_VERDE, C_ROJO
except Exception:
    class Dummy: RESET_ALL = ""
    Fore = Style = Dummy()
    OK = WARN = INFO = ERR = DATA = CODE = RESET = C_BARRA = C_INDICE = C_HORA = C_NOMBRE = C_RUT = C_FECHA = ""
    C_CYAN = C_FUCSIA = C_VERDE = C_ROJO = C_NARANJA = C_M1_LABEL = C_M2_LABEL = C_EXITO = C_SIN_CASO = C_FALLO = C_SI = C_NO = ""

# =============================== HELPERS / UTILS =================================
def log_error(msg: str, tipo: str = "ERROR") -> None:
    color = ERR if tipo == "ERROR" else WARN
    print(f"{color}[{tipo}] {datetime.now().strftime('%H:%M:%S')} :: {msg}{RESET}")

def espera(clave: str) -> None:
    t = float(ESPERAS.get(clave, {"sleep": 0.5}).get("sleep", 0.5))
    if t > 0: time.sleep(t)

class SpinnerStuck(Exception): pass
_ddmmyyyy = re.compile(r"\b(\d{2}/\d{2}/\d{4})\b", re.I)

def _norm(s: str) -> str:
    if not s: return ""
    s = unicodedata.normalize("NFKD", s)
    return re.sub(r"[\s]+", " ", re.sub(r"[^a-z0-9\sáéíóúüñ]"," ", "".join(c for c in s if not unicodedata.combining(c)).lower().strip()))

def has_keyword(texto: str, kws: List[str]) -> bool:
    t = _norm(texto); return any(_norm(k) in t for k in kws)

def solo_fecha(x: Any) -> str:
    if isinstance(x, datetime): return x.strftime("%d/%m/%Y")
    s = str(x).split(" ")[0].replace("-", "/"); p = s.split("/")
    return f"{p[2]}/{p[1]}/{0}" if len(p[0])==4 else s

def dparse(s: str) -> Optional[datetime]:
    try: return datetime.strptime(s.split(" ")[0], "%d/%m/%Y")
    except: return None

def same_month(a: datetime, b: datetime) -> bool:
    return a.year==b.year and a.month==b.month

def en_vig(fecha_obj: Optional[datetime], dt: Optional[datetime]) -> bool:
    if not (fecha_obj and dt): return False
    return 0 <= (fecha_obj-dt).days <= VENTANA_VIGENCIA_DIAS

def join_tags(tags: List[str]) -> str:
    return " + ".join([t.strip() for t in tags if t and str(t).strip()])

def join_clean(vals: List[str], sep: str = " | ") -> str:
    clean = [v for v in vals if v and v.strip() and v != "X" and "NO TIENE" not in v]
    if not clean:
        if any("NO TIENE" in (v or "") for v in vals):
            if "OA" in vals[0]: return "Sin OA"
            if "IPD" in vals[0]: return "Sin IPD"
            return "" 
        return ""
    return sep.join(clean)

def normalizar_rut(rut: str) -> str:
    rut = rut.replace(".", "").replace("-", "").strip().upper()
    if not rut: return ""
    return rut[:-1] + "-" + rut[-1]

# =============================== WRAPPER SIGGES ======================================
class SiggesDriver:
    def __init__(self, driver: webdriver.Remote) -> None:
        self.driver = driver

    def hay_spinner(self) -> bool:
        try: return bool(self.driver.find_elements(By.CSS_SELECTOR, XPATHS["SPINNER_CSS"]))
        except: return False

    def esperar_spinner(self, appear_timeout=2, clave_espera="spinner", raise_on_timeout=False):
        cfg = ESPERAS.get(clave_espera, {"wait": 25.0})
        disappear_timeout = cfg["wait"]
        seen=False
        try:
            end = time.time() + appear_timeout 
            while time.time() < end:
                if self.hay_spinner(): seen=True; break
                time.sleep(0.2)
        except: seen=False
        if not seen: return
        try:
            WebDriverWait(self.driver, disappear_timeout).until(
                EC.invisibility_of_element_located((By.CSS_SELECTOR, XPATHS["SPINNER_CSS"])))
            time.sleep(0.5) 
        except TimeoutException:
            if self.hay_spinner() and raise_on_timeout: raise SpinnerStuck(f"Spinner pegado")

    def _find(self, locators: List[str], mode="clickable", clave_espera="default") -> Optional[Any]:
        cond = {"presence":EC.presence_of_element_located, "visible":EC.visibility_of_element_located}.get(mode, EC.element_to_be_clickable)
        timeout = float(ESPERAS.get(clave_espera, {"wait": 2})["wait"])
        for xp in locators:
            try: return WebDriverWait(self.driver, timeout).until(cond((By.XPATH,xp)))
            except: continue
        return None

    def _click(self, locators: List[str], scroll=True, wait_spinner=True, clave_espera="default", spinner_clave_espera="spinner", raise_on_spinner_timeout=False) -> bool:
        el = self._find(locators, "clickable", clave_espera)
        if not el: return False
        try:
            if scroll: self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
            el.click()
        except:
            try: self.driver.execute_script("arguments[0].click();", el)
            except: return False
        if wait_spinner: self.esperar_spinner(clave_espera=spinner_clave_espera, raise_on_timeout=raise_on_spinner_timeout)
        return True

    def sesion_cerrada(self) -> bool:
        try:
            if "/#/login" in (self.driver.current_url or ""): return True
            if self._find(XPATHS["LOGIN_BTN_INGRESAR"], "visible", "login_check"): return True
        except: pass
        return False

    def ir(self, url: str) -> None:
        try: self.driver.get(url)
        except: 
            try: self.driver.execute_script("window.stop();")
            except: pass

    def asegurar_menu_desplegado(self) -> None:
        try:
            nav_div = self._find([XPATHS["MENU_CONTENEDOR"]], "presence", "menu_lateral")
            if not nav_div: return
            clases = nav_div.get_attribute("class") or ""
            if XPATHS["MENU_CLASS_ABIERTO"] in clases: return
            self._click([XPATHS["MENU_ICONO_APERTURA"]], scroll=False, wait_spinner=False, clave_espera="menu_lateral")
            time.sleep(0.5) 
        except Exception: pass

    # --- FUNCIÓN DE NAVEGACIÓN ---
    def asegurar_en_busqueda(self) -> None:
        try:
            if self.sesion_cerrada():
                log_error("Sesión cerrada. Relogueando...", "WARN")
                if not intentar_login(self, reintentar_una_vez=True): raise Exception("Fallo Login")
                self.ir(XPATHS["BUSQUEDA_URL"])
            self.asegurar_menu_desplegado()
            curr = self.driver.current_url or ""
            if "busqueda-de-paciente" in curr:
                 if self._find([XPATHS["INPUT_RUT"]], "presence", "nav_a_busqueda_fast"): return
            self.esperar_spinner(raise_on_timeout=False)
            self.ir(XPATHS["BUSQUEDA_URL"])
            self.esperar_spinner(raise_on_timeout=False)
            if not self._find([XPATHS["INPUT_RUT"]], "presence", "nav_a_busqueda"):
                raise TimeoutException("No cargó búsqueda")
        except TimeoutException: raise
        except Exception as e: 
            log_error(f"Error nav busqueda: {e}"); raise

    def ir_a_cartola(self, raise_on_spinner_timeout=True) -> bool:
        self.asegurar_menu_desplegado()
        return self._click([XPATHS["BTN_MENU_CARTOLA"]], False, True, "click_ir_a_cartola", "spinner_largo", raise_on_spinner_timeout)

    def activar_hitos_ges(self) -> None:
        chk = self._find([XPATHS["CHK_HITOS_GES"]], "presence", "click_activar_hitos")
        try: 
            if chk and not chk.is_selected(): self._click([XPATHS["CHK_HITOS_GES"]], True, True, "click_activar_hitos")
        except: pass

    def lista_de_casos_cartola(self) -> List[str]:
        root = self._find([XPATHS["CONT_CARTOLA"]], "presence", "cont_cartola")
        if not root: return []
        try: return [(c.text or "").strip() for c in root.find_elements(By.CSS_SELECTOR,"div.contRow") if (c.text or "").strip()]
        except: return []

    def _case_root(self, i:int) -> Optional[Any]:
        try: return self._find([f"({XPATHS['CONT_CARTOLA']})/div[{i+1}]"], "presence", "expandir_caso")
        except: return None

    def expandir_caso(self, i:int) -> Optional[Any]:
        root = self._case_root(i)
        if not root: return None
        try:
            if not root.find_elements(By.XPATH, ".//table"):
                self._click([f"({XPATHS['CONT_CARTOLA']})/div[{i+1}]/div/label/input", f"({XPATHS['CONT_CARTOLA']})/div[{i+1}]"], False, "expandir_caso")
        except: pass
        espera("expandir_caso")
        return root

    def cerrar_caso_por_indice(self, i:int)->None:
        try:
            root = self._case_root(i)
            if root and root.find_elements(By.XPATH, ".//table"):
                self._click([f"({XPATHS['CONT_CARTOLA']})/div[{i+1}]/div/label/input"], False, "expandir_caso")
        except: pass
        espera("expandir_caso")

    def _prestaciones_tbody(self, i:int) -> Optional[Any]:
        return self._find([f"({XPATHS['CONT_CARTOLA']})/div[{i+1}]/div[6]/div[2]/div/table/tbody", f"({XPATHS['CONT_CARTOLA']})/div[{i+1}]//table/tbody"], "presence", "prestaciones_tbody")

    def leer_prestaciones_desde_tbody(self, tbody: Any) -> List[Dict[str,str]]:
        out=[]
        try:
            for tr in reversed(tbody.find_elements(By.TAG_NAME,"tr")):
                tds = tr.find_elements(By.TAG_NAME,"td")
                if len(tds)>8:
                    f = (tds[2].text.strip().split(" ")[0] if len(tds)>2 else "")
                    if not dparse(f):
                        for c in tds:
                            m=_ddmmyyyy.search(c.text)
                            if m: f = m.group(1); break
                    out.append({
                        "fecha":f, 
                        "codigo":tds[7].text.strip(), 
                        "glosa":tds[8].text.strip(),
                        "ref": tds[0].text.strip()
                    })
        except: pass
        return out

    def leer_fallecimiento(self) -> Optional[datetime]:
        try:
            el = self._find(["//div[span[normalize-space(.)='Fecha de fallecimiento'] and p]"], "presence", "leer_fallecimiento")
            if el: return dparse((el.find_element(By.TAG_NAME,"p").text or "").strip())
        except: pass
        return None

    def leer_edad(self) -> Optional[int]:
        try:
            el = self._find([XPATHS["EDAD_PACIENTE"]], "presence", "leer_edad")
            if el: 
                m = re.search(r"(\d+)", (el.text or "").strip())
                return int(m.group(1)) if m else None
        except: return None

    # --- HELPERS DE BÚSQUEDA DE SECCIONES (Robustez Old School) ---
    def _find_section_label_p(self, root, needle: str):
        nd = _norm(needle)
        try:
            for el in root.find_elements(By.XPATH, ".//div/label/p"):
                if _norm(el.text) and nd.split('(')[0].strip() in _norm(el.text):
                    return el
        except: pass
        return None
        
    def _tbody_from_label_p(self, p_el):
        # Busca la tabla hermana o sobrina
        for xp in ["../../../following-sibling::div[1]//table/tbody",
                   "../../following-sibling::div[1]//table/tbody",
                   "../following-sibling::div[1]//table/tbody",
                   "ancestor::div[1]/following-sibling::div[1]//table/tbody"]:
            try:
                tb = p_el.find_element(By.XPATH, xp)
                if tb: return tb
            except: continue
        return None

    # --- LECTOR DE IPD (REPARADO) ---
    def leer_ipd_desde_caso(self, root: Any, n:int) -> Tuple[List[str], List[str], List[str]]:
        if not root: return [], [], []
        try:
            tbody = None
            # Intento 1: Etiqueta inteligente
            p = self._find_section_label_p(root, "informes de proceso de diagnóstico")
            if p: tbody = self._tbody_from_label_p(p)
            
            # Intento 2: Fallbacks absolutos
            if not tbody:
                 for xp in XPATHS["IPD_TBODY_FALLBACK"]:
                     try: 
                        # Busqueda relativa al root para no agarrar otros casos
                         tbody = root.find_element(By.XPATH, "." + xp.split("/main")[1]) 
                         if tbody: break
                     except: continue

            if not tbody: return [], [], []

            rows = tbody.find_elements(By.TAG_NAME, "tr") or []
            parsed=[]
            for r in rows:
                try:
                    tds = r.find_elements(By.TAG_NAME, "td")
                    f_txt = (tds[2].text or "").strip()
                    e_txt = (tds[6].text or "").strip()
                    d_txt = (tds[7].text or "").strip()
                    f_dt = dparse(f_txt) or datetime.min
                    parsed.append((f_dt, f_txt, e_txt, d_txt))
                except: continue
            parsed.sort(key=lambda x: x[0], reverse=True)
            return [p[1] for p in parsed[:n]], [p[2] for p in parsed[:n]], [p[3] for p in parsed[:n]]
        except: return [], [], []

    # --- LECTOR DE OA (REPARADO) ---
    def leer_oa_desde_caso(self, root: Any, n:int) -> Tuple[List[str], List[str], List[str], List[str], List[str]]:
        if not root: return [], [], [], [], []
        try:
            tbody = None
            # Intento 1: Etiqueta inteligente
            p = self._find_section_label_p(root, "ordenes de atencion")
            if p: tbody = self._tbody_from_label_p(p)
            
            # Intento 2: Fallback absoluto
            if not tbody:
                for xp in XPATHS["OA_TBODY_FALLBACK"]:
                    try:
                        tbody = root.find_element(By.XPATH, "." + xp.split("/main")[1])
                        if tbody: break
                    except: continue

            if not tbody: return [], [], [], [], []

            rows = tbody.find_elements(By.TAG_NAME, "tr") or []
            parsed=[]
            for r in rows:
                try:
                    tds = r.find_elements(By.TAG_NAME, "td")
                    # Index 0 para Folio es td:nth-child(1)
                    folio = (tds[0].text or "").strip()
                    f_txt = (tds[2].text or "").split(" ")[0].strip()
                    deriv = (tds[8].text or "").strip()
                    cod   = (tds[9].text or "").strip()
                    diag  = (tds[12].text or "").strip()
                    
                    f_dt = dparse(f_txt) or datetime.min
                    parsed.append((f_dt, f_txt, deriv, diag, cod, folio))
                except: continue
            
            parsed.sort(key=lambda x: x[0], reverse=True)
            parsed = parsed[:n]
            
            return (
                [p[1] for p in parsed], # Fecha
                [p[2] for p in parsed], # Deriv
                [p[3] for p in parsed], # Diag
                [p[4] for p in parsed], # Codigo
                [p[5] for p in parsed]  # Folio
            )
        except: return [], [], [], [], []

# =============================== MINI TABLA ==========================================
def leer_mini_tabla_busqueda(driver: Any, stable_reads=2) -> List[Dict[str,str]]:
    tbody=None
    for xp in XPATHS["MINI_TABLA_TBODY"]:
        try: 
            tbody = WebDriverWait(driver, ESPERAS["mini_tabla_read"]["wait"]).until(EC.presence_of_element_located((By.XPATH,xp)))
            if tbody: break
        except: continue
    if not tbody: return []
    try:
        rows = tbody.find_elements(By.TAG_NAME, "tr")
        out = []
        for tr in rows:
            tds = tr.find_elements(By.TAG_NAME, "td")
            if len(tds)<4: continue
            out.append({"problema":tds[1].text, "decreto":tds[2].text, "estado":tds[3].text})
        return [r for r in out if "evento sin caso" not in _norm(r["problema"])]
    except: return []

# =============================== LÓGICA DE MISIONES ==================================

def elegir_caso_mas_reciente(casos: List[str], kws: List[str]) -> Tuple[Optional[int], Optional[str], Optional[datetime]]:
    best_i, best_t, best_f = None, None, None
    for i, t in enumerate(casos):
        if not has_keyword(t, kws): continue
        m=_ddmmyyyy.search(t); f = dparse(m.group(1)) if m else None
        if best_f is None or (f and f>best_f): best_i, best_t, best_f = i, (t.strip() if t else ""), (f or best_f)
    return best_i, best_t, best_f

def listar_habilitantes(prest: List[Dict[str,str]], cods: List[str], fobj: Optional[datetime]) -> List[Tuple[str, datetime]]:
    out=[]
    for p in prest:
        c = p.get("codigo","")
        if c in cods:
            f = dparse(p.get("fecha",""))
            if f and (not fobj or f <= fobj): out.append((c,f))
    return sorted(out, key=lambda x:x[1], reverse=True)

def cols_mision(m: Dict[str,Any]) -> List[str]:
    cols = ["Fecha", "Rut", "Nombre", "Edad", "F Obj", "Familia", "Especialidad", "Caso Encontrado", "Mensual", "C Hab", "F Hab", "Hab Vi"]
    if m.get("excluyentes"): cols += ["C Excluyente", "F Excluyente"]
    if REVISAR_IPD: cols += ["Fecha IPD", "Estado IPD", "Diagnóstico IPD"]
    if REVISAR_OA: cols += ["Código OA", "Fecha OA", "Folio OA", "Derivado OA", "Diagnóstico OA"]
    cols += ["Observación", "Observación FOLIO"]
    return cols

def vac_row(m: Dict[str,Any], fecha: str, rut: str, nombre: str, obs: str="") -> Dict[str,Any]:
    r = {c: "" for c in cols_mision(m)}
    r["Fecha"] = fecha
    r["Rut"] = rut
    r["Nombre"] = nombre
    r["Observación"] = obs
    r["Caso Encontrado"] = "Sin caso"
    r["Mensual"] = "Sin Día"
    r["Familia"] = m.get("familia","")
    r["Especialidad"] = m.get("especialidad","")
    return r

def analizar_mision(sigges: SiggesDriver, m: Dict[str,Any], casos: List[str], fobj: datetime,
                    fecha: str, fall_dt: Optional[datetime], edad_paciente: Optional[int], nombre: str) -> Dict[str,Any]:
    res = vac_row(m, fecha, "", nombre, "")
    res["Edad"] = str(edad_paciente) if edad_paciente is not None else ""

    idx, texto, _ = elegir_caso_mas_reciente(casos, m.get("keywords", []))
    if idx is None:
        res["Observación"] = "Sin caso"
        return res
    res["Caso Encontrado"] = texto
    
    root = sigges.expandir_caso(idx)
    if not root:
        res["Observación"] = "Error expandir"
        return res

    prestaciones = []
    folios_oa_encontrados = [] # (folio, fecha_dt, codigo, derivado, fecha_str)
    try:
        if REVISAR_IPD:
            f_list, e_list, d_list = sigges.leer_ipd_desde_caso(root, IPD_FILAS_REVISION)
            res["Fecha IPD"] = join_clean(f_list)
            res["Estado IPD"] = join_clean(e_list)
            res["Diagnóstico IPD"] = join_clean(d_list)
        
        if REVISAR_OA:
            f_oa, p_oa, d_oa, c_oa, fol_oa = sigges.leer_oa_desde_caso(root, OA_FILAS_REVISION)
            res["Fecha OA"] = join_clean(f_oa)
            res["Derivado OA"] = join_clean(p_oa)
            res["Diagnóstico OA"] = join_clean(d_oa)
            res["Código OA"] = join_clean(c_oa)
            res["Folio OA"] = join_clean(fol_oa)
            
            # Recolectar datos completos para el formato detallado
            for i, fol in enumerate(fol_oa):
                if fol and f_oa[i]:
                    dt_oa = dparse(f_oa[i])
                    if dt_oa:
                        # Guardamos todo el contexto para formatear bonito después
                        folios_oa_encontrados.append((fol, dt_oa, c_oa[i], p_oa[i], f_oa[i]))

        tb = sigges._prestaciones_tbody(idx)
        prestaciones = sigges.leer_prestaciones_desde_tbody(tb) if tb else []
    except Exception as e:
        res["Observación"] = f"Error lectura: {e}"
    finally:
        sigges.cerrar_caso_por_indice(idx)

    # --- Lógica Datos ---
    objetivos = m.get("objetivos") or ([m.get("objetivo")] if m.get("objetivo") else [])
    objetivos = [str(c).strip() for c in objetivos][:MAX_OBJETIVOS]
    
    fechas_obj_strs = []
    fechas_obj_all_dts = []
    for cod in objetivos:
        dts = [p["dt"] for p in [{**p, "dt":dparse(p["fecha"])} for p in prestaciones] if p["dt"] and p.get("codigo")==cod and (not fobj or p["dt"]<=fobj)]
        last = max(dts) if dts else None
        if last: 
            fechas_obj_strs.append(last.strftime("%d/%m/%Y"))
            fechas_obj_all_dts.append(last)
        else:
            fechas_obj_strs.append("")
    res["F Obj"] = " | ".join(fechas_obj_strs)

    if fechas_obj_all_dts:
        dts_mes = [d for d in fechas_obj_all_dts if same_month(d, fobj)]
        res["Mensual"] = "1 en Mes" if dts_mes else "Sin Día"
    else:
        res["Mensual"] = "Sin Día"

    habs_cfg = m.get("habilitantes", [])
    habs_found = []
    if m.get("usar_habilitantes") and habs_cfg:
        habs_found = listar_habilitantes(prestaciones, habs_cfg, fobj)
        if habs_found:
            res["C Hab"] = join_clean([h[0] for h in habs_found[:TOP_K_HABILITANTES]])
            res["F Hab"] = join_clean([h[1].strftime("%d/%m/%Y") for h in habs_found[:TOP_K_HABILITANTES]])
            res["Hab Vi"] = "Vigente" if any(en_vig(fobj, h[1]) for h in habs_found) else "No Vigente"
        else:
             res["Hab Vi"] = "Sin Habilitante"
    
    excl_list = m.get("excluyentes", [])
    if excl_list:
        excl_found = [(p["codigo"], p["fecha"]) for p in prestaciones if p.get("codigo") in excl_list]
        if excl_found:
            res["C Excluyente"] = join_clean([x[0] for x in excl_found])
            res["F Excluyente"] = join_clean([x[1] for x in excl_found])
    
    # --- FORMATEO DE OBSERVACIÓN DE FOLIOS (NUEVO FORMATO SOLICITADO) ---
    obs_folio_list = []
    if REVISAR_OA and folios_oa_encontrados:
        ahora = datetime.now()
        un_ano_atras = ahora - timedelta(days=365)
        refs_prestaciones = []
        
        # Cacheamos referencias de prestaciones recientes
        for p in prestaciones:
            p_dt = dparse(p.get("fecha",""))
            if p_dt and p_dt >= un_ano_atras: refs_prestaciones.append(_norm(p.get("ref","")))
            
        for folio, dt_oa, codigo, derivado, fecha_str in folios_oa_encontrados:
            if dt_oa >= un_ano_atras:
                folio_clean = _norm(folio).replace("oa","").strip()
                found = False
                for ref in refs_prestaciones:
                    if folio_clean in ref: found = True; break
                
                if found:
                    # Formato: Fol 82437861 / Cód 2504092 / Fec 28/07/2025
                    obs_folio_list.append(f"Fol {folio} / Cód {codigo} / Fec {fecha_str}")

    res["Observación FOLIO"] = " | ".join(obs_folio_list)

    obs_tags = []
    if fall_dt: obs_tags.append("Fallecido")
    emin, emax = m.get("edad_min"), m.get("edad_max")
    if edad_paciente is not None:
        if (emin and edad_paciente < emin) or (emax and edad_paciente > emax):
            obs_tags.append("No cumple con la Edad")
    if not fechas_obj_all_dts: obs_tags.append("Sin Objetivo")
    if REVISAR_OA and not res.get("Código OA"): obs_tags.append("Sin OA")
    if REVISAR_IPD and not res.get("Fecha IPD"): obs_tags.append("Sin IPD")

    res["Observación"] = join_tags(obs_tags)
    return res

# =============================== LOGIN / REINTENTOS ==================================
def safe_refresh(driver):
    try: driver.refresh()
    except: pass

def intentar_login(sig: SiggesDriver, reintentar_una_vez=True) -> bool:
    try:
        sig._click(XPATHS["LOGIN_BTN_INGRESAR"], False, True, "login_ingresar", "spinner_largo")
        sig._click(XPATHS["LOGIN_SEL_UNIDAD_HEADER"], True, False, "login_sel_unidad")
        sig._click(XPATHS["LOGIN_OP_HOSPITAL"], True, False, "login_op_hospital")
        sig._click(XPATHS["LOGIN_TILE_INGRESO_SIGGES"], True, False, "login_ingreso_sigges")
        sig._click(XPATHS["LOGIN_BTN_CONECTAR"], True, True, "login_conectar", "spinner_largo")
        sig.asegurar_menu_desplegado()
        if "actualizaciones" in sig.driver.current_url: return True
    except Exception as e:
        log_error(f"Fallo login: {e}", "WARN")
    if reintentar_una_vez:
        safe_refresh(sig.driver)
        # Nota: Usamos una clave genérica en lugar de las que no estaban definidas
        espera("default") 
        return intentar_login(sig, False)
    return False

# =============================== EXPORTAR CON ESTILOS (RESTAURADA Y CORREGIDA) =================================
def aplicar_estilos_excel(ws):
    styles = {
        "Rosa":    {"fill": "FFC0CB", "font_color": "000000", "bold": True},
        "Celeste": {"fill": "87CEEB", "font_color": "000000", "bold": True},
        "Naranja": {"fill": "FFA500", "font_color": "000000", "bold": True},
        "Verde":   {"fill": "90EE90", "font_color": "FFFFFF", "bold": True}, 
        "Amarillo":{"fill": "FFFF00", "font_color": "000000", "bold": True},
        "Salmon":  {"fill": "FA8072", "font_color": "000000", "bold": True},
        "Rojo":    {"fill": "FF0000", "font_color": "FFFFFF", "bold": True},
        "Azul":    {"fill": "0000FF", "font_color": "FFFFFF", "bold": True},
    }

    for cell in ws[1]:
        val = str(cell.value).strip()
        st_key = None
        
        # Jerarquía estricta
        if "OA" in val: st_key = "Azul"  # Manda sobre todo lo demás
        elif "IPD" in val: st_key = "Rojo"
        elif "Excluyente" in val: st_key = "Salmon"
        elif "Hab" in val: st_key = "Amarillo"
        elif "Obj" in val: st_key = "Celeste"
        elif "Caso" in val or "Mensual" in val: st_key = "Verde"
        elif "Familia" in val or "Especialidad" in val: st_key = "Naranja"
        else: st_key = "Rosa" # Default para Fecha, Rut, Obs, Edad

        if st_key:
            s = styles[st_key]
            cell.fill = PatternFill(start_color=s["fill"], end_color=s["fill"], fill_type="solid")
            cell.font = Font(color=s["font_color"], bold=s["bold"])
            cell.alignment = Alignment(horizontal='center')

# --- TERMINAL INTACTA (RESTO DE RESUMEN PACIENTE ABAJO) ---
def resumen_paciente(i: int, total: int, nombre: str, rut: str, fecha: str, 
                     flags: Dict[str,bool], resultados: List[Dict[str, Any]]) -> None:
    
    now = datetime.now().strftime("%H:%M") 
    b = C_BARRA + "|" + RESET 

    m1 = resultados[0] if len(resultados) > 0 else {}
    m2 = resultados[1] if len(resultados) > 1 else {}
    
    paciente_ok = flags.get("ok", False)
    paciente_saltado = flags.get("saltado", False)

    m1_mini_val = "Sí" if (m1.get("Caso Encontrado") or "Sin caso") != "Sin caso" else "No"
    m1_mini_col = C_SI if m1_mini_val == "Sí" else C_NO
    
    m1_hab_val_display = (m1.get("Hab Vi") or "No").split('|')[0]
    m1_hab_col  = C_SI if "Vigente" in m1_hab_val_display else C_NO
    
    m1_ipd_val  = (m1.get("Estado IPD", "") or "Sin IPD")
    if "no tiene" in m1_ipd_val.lower(): m1_ipd_val = "Sin IPD"
    m1_ipd_col  = C_SI if m1_ipd_val != "Sin IPD" else C_NO
    
    m1_oa_val   = "Con OA" if m1.get("Código OA") else "Sin OA"
    m1_oa_col   = C_SI if m1_oa_val == "Con OA" else C_NO
    
    m2_mini_val = "Sí" if (m2.get("Caso Encontrado") or "Sin caso") != "Sin caso" else "No"
    m2_mini_col = C_SI if m2_mini_val == "Sí" else C_NO

    m2_hab_val_display = (m2.get("Hab Vi") or "No").split('|')[0]
    m2_hab_col  = C_SI if "Vigente" in m2_hab_val_display else C_NO
    
    m2_ipd_val  = (m2.get("Estado IPD", "") or "Sin IPD")
    if "no tiene" in m2_ipd_val.lower(): m2_ipd_val = "Sin IPD"
    m2_ipd_col  = C_SI if m2_ipd_val != "Sin IPD" else C_NO

    m2_oa_val   = "Con OA" if m2.get("Código OA") else "Sin OA"
    m2_oa_col   = C_SI if m2_oa_val == "Con OA" else C_NO

    part1 = (
        f"🔥 {C_INDICE}[{i}/{total}]{RESET} 🔥 {b} "
        f"⏳ {C_HORA}{now}{RESET} ⏳ {b} "
        f"🤹🏻  {C_NOMBRE}{nombre.upper()}{RESET} 🤹🏻 {b} "
        f"🪪  {C_RUT}{rut}{RESET} 🪪  {b}"
        f"🗓️  {C_FECHA}{fecha}{RESET} 🗓️" 
    )
    part2 = (
        f"📋 {C_M1_LABEL}M1-Mini: {m1_mini_col}{m1_mini_val}{RESET} 📋 {b} "
        f"📋 {C_M2_LABEL}M2-Mini: {m2_mini_col}{m2_mini_val}{RESET} 📋 {b} "
        f"♦️ {C_M1_LABEL} M1-Hab: {m1_hab_col}{m1_hab_val_display}{RESET} ♦️  {b}  "
        f"♦️  {C_M2_LABEL} M2-Hab: {m2_hab_col}{m2_hab_val_display}{RESET} ♦️"
    )
    part3 = (
        f"🔶 {C_M1_LABEL}M1-IPD: {m1_ipd_col}{m1_ipd_val}{RESET} 🔶 {b} "
        f"🔶 {C_M2_LABEL}M2-IPD: {m2_ipd_col}{m2_ipd_val}{RESET} 🔶 {b} "
        f"🔷 {C_M1_LABEL}M1-OA: {m1_oa_col}{m1_oa_val}{RESET} 🔷 {b} "
        f"🔷 {C_M2_LABEL}M2-OA: {m2_oa_col}{m2_oa_val}{RESET} 🔷"
    )
    def _get_status_for_mission(res: Dict[str, Any]) -> Tuple[str, str]:
        if not res: return ("⚠️  Sin Datos ⚠️", C_NARANJA)
        mini_found = (res.get("Caso Encontrado") or "Sin caso") != "Sin caso"
        obs_txt = res.get("Observación", "")
        obs_critica = "Excluyentes" in obs_txt or "Edad" in obs_txt or "Fallecido" in obs_txt
        if not mini_found: return ("⚠️ Sin Caso ⚠️", C_NARANJA) 
        if obs_critica: return ("⚠️  Observaciones ⚠️", C_NARANJA)
        return ("✅  Éxito ✅", C_EXITO)

    status_msg = ""; status_col = ""; part4 = ""      
    if paciente_saltado:
        status_msg = f"♻️ Paciente Saltado Automáticamente ({MAX_REINTENTOS_GRAVES_POR_PACIENTE} Reintentos) ♻️"
        status_col = C_ROJO 
        part4 = f"{status_col}{status_msg}{RESET}"
    elif not paciente_ok:
        status_msg = "❌ Paciente No Revisado (Error Inesperado) ❌"
        status_col = C_FALLO 
        part4 = f"{status_col}{status_msg}{RESET}"
    else:
        m1_status, m1_col = _get_status_for_mission(m1)
        m2_status, m2_col = _get_status_for_mission(m2)
        part4 = (
            f"{C_ROJO}Paciente:{RESET} "
            f"{C_FUCSIA}Mision 1{RESET} - {m1_col}{m1_status}{RESET} {b} "
            f"{C_CYAN}Mision 2{RESET} - {m2_col}{m2_status}{RESET}"
        )
        if m1_col == C_EXITO and m2_col == C_EXITO: status_col = C_EXITO
        else: status_col = C_NARANJA
    try:
        print(f"{part1}"); print(); print(f"{part2}"); print(); print(f"{part3}"); print(); print(part4); print(f"{C_BARRA}{'-' * 80}{RESET}")
    except: pass

# =============================== MAIN LOOP DE REVISIÓN =================================
def ejecutar_revision()->None:
    opts = webdriver.EdgeOptions(); opts.debugger_address = DIRECCION_DEBUG_EDGE
    try: 
        driver = webdriver.Edge(service=Service(EDGE_DRIVER_PATH), options=opts)
        driver.set_page_load_timeout(ESPERAS["page_load"]["wait"])
    except Exception as e: print(f"{ERR}Error Edge: {e}{RESET}"); return
    sigges = SiggesDriver(driver)
    try: df = pd.read_excel(RUTA_ARCHIVO_ENTRADA)
    except Exception as e: print(f"{ERR}Error Excel: {e}{RESET}"); return

    filas_por_mision = defaultdict(list)
    total = len(df)
    for idx,row in df.iterrows():
        if (idx>0 and idx%60==0): gc.collect()
        nombre = str(row.iloc[INDICE_COLUMNA_NOMBRE]).strip()
        rut    = normalizar_rut(str(row.iloc[INDICE_COLUMNA_RUT]).strip())
        fecha  = solo_fecha(row.iloc[INDICE_COLUMNA_FECHA]); fobj = dparse(fecha)
        
        intento = 0
        resuelto = False
        saltado = False
        res_paci = []

        # --- LOOP DE REINTENTOS INTELIGENTES ---
        while intento < MAX_REINTENTOS_GRAVES_POR_PACIENTE and not resuelto:
            intento += 1
            try:
                # --- LOGICA DE REINTENTOS (PRE-BUSQUEDA) ---
                if intento == 2:
                     # Reintento 1 (falló el 1): Esperar 3 seg -> ir busqueda
                     time.sleep(3)
                     sigges.ir(XPATHS["BUSQUEDA_URL"])
                elif intento == 3:
                     # Reintento 2 (falló el 2): Esperar 3s -> busqueda -> esperar 2s -> escribir
                     time.sleep(3)
                     sigges.ir(XPATHS["BUSQUEDA_URL"])
                     time.sleep(2)
                elif intento >= 4:
                     # Reintento 3 y 4: F5 -> esperar 2s -> busqueda -> esperar 2s -> escribir
                     safe_refresh(sigges.driver)
                     time.sleep(2)
                     sigges.ir(XPATHS["BUSQUEDA_URL"])
                     time.sleep(2)

                # --- PROCESO DE BÚSQUEDA ---
                sigges.asegurar_en_busqueda()
                
                el = sigges._find([XPATHS["INPUT_RUT"]], "presence", "find_rut")
                if not el: raise Exception("No se encontró input RUT")
                el.clear(); espera("clear_rut")
                el.send_keys(rut); espera("send_rut")
                sigges._click([XPATHS["BTN_BUSCAR"]], False, True, "click_buscar", "spinner")

                mini = leer_mini_tabla_busqueda(sigges.driver, MINI_TABLA_STABLE_READS)
                hallo_mini = any(has_keyword(r.get("problema",""), m.get("keywords", [])) for r in mini for m in MISSIONS)
                
                if not hallo_mini:
                    for m in MISSIONS: res_paci.append(vac_row(m, fecha, rut, nombre, "Sin caso en mini-tabla"))
                    resuelto = True
                else:
                    edad = sigges.leer_edad()
                    if not sigges.ir_a_cartola(): raise TimeoutException("Fallo ir a cartola")
                    sigges.activar_hitos_ges()
                    
                    fall_dt = sigges.leer_fallecimiento()
                    casos = sigges.lista_de_casos_cartola()
                    
                    res_paci = [] 
                    for m in MISSIONS:
                        r = analizar_mision(sigges, m, casos, fobj, fecha, fall_dt, edad, nombre)
                        r["Rut"] = rut; r["Fecha"] = fecha; r["Nombre"] = nombre
                        res_paci.append(r)

                    resuelto = True

            except (TimeoutException, WebDriverException, Exception) as e:
                log_error(f"Intento {intento} fallido para {rut}: {e}", "WARN")
                # El loop continuará e incrementará 'intento'
            except KeyboardInterrupt: return

        if not resuelto:
            saltado = True
            res_paci = []
            for m in MISSIONS: res_paci.append(vac_row(m, fecha, rut, nombre, "Paciente Saltado Automáticamente"))
        
        for i_m, r in enumerate(res_paci): filas_por_mision[i_m].append(r)
        resumen_paciente(idx+1, total, nombre, rut, fecha, {"ok":resuelto, "saltado":saltado}, res_paci)
        espera("entre_pacientes")

    stamp = datetime.now().strftime("%d-%m-%Y_%H-%M")
    out_xlsx = os.path.join(RUTA_CARPETA_SALIDA, f"Rev_{NOMBRE_DE_LA_MISION}_{stamp}.xlsx")
    try:
        with pd.ExcelWriter(out_xlsx, engine="openpyxl") as w:
            for i_m, m in enumerate(MISSIONS):
                cols = cols_mision(m)
                df_out = pd.DataFrame(filas_por_mision[i_m]).reindex(columns=cols)
                df_out.to_excel(w, index=False, sheet_name=f"Mision {i_m+1}")
                aplicar_estilos_excel(w.sheets[f"Mision {i_m+1}"])
        print(f"{OK}Guardado: {out_xlsx}{RESET}")
    except Exception as e: print(f"{ERR}Error guardado: {e}{RESET}")

if __name__=="__main__":
    ejecutar_revision()