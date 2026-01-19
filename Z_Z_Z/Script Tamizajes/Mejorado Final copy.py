"""
Script de automatización SIGGES - Tamizajes
Autor: Nozh (Refactorizado por Gemini)
Versión: 2.0 - Arquitectura mejorada con clase encapsulada
"""
from __future__ import annotations

import os
import re
import sys
import time
from typing import List, Dict, Any, Optional, Tuple, Callable
from datetime import datetime
from contextlib import contextmanager

# ───────── Selenium ─────────
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.edge.service import Service
from selenium.common.exceptions import (
    TimeoutException, WebDriverException,
    ElementNotInteractableException, StaleElementReferenceException, NoSuchElementException
)

# ───────── Excel / Consola ─────────
try:
    import openpyxl
except ImportError:
    print("❌ Falta 'openpyxl'. Instala con: pip install openpyxl")
    raise

try:
    from colorama import Fore, Style, init as colorama_init
    colorama_init(autoreset=True)
    C_OK, C_ERR, C_DIM = Fore.GREEN, Fore.RED, Style.DIM
    C_CNT, C_RUT, C_DATE, C_NAME = Fore.MAGENTA, Fore.CYAN, Fore.LIGHTRED_EX, Fore.YELLOW
except ImportError:
    class _Dummy:
        RESET_ALL = ""
    C_OK = C_ERR = C_DIM = C_CNT = C_RUT = C_DATE = C_NAME = _Dummy()
    Fore = Style = _Dummy()

# ╔═══════════════════════════════════════════════════════════════════╗
# ║ 🧭 PANEL DE CONTROL GENERAL                                       ║
# ╚═══════════════════════════════════════════════════════════════════╝
PANEL: Dict[str, Any] = {
    "identidad": {
        "autor": "Nozh (Modo Dios por Gemini)",
        "version": "2.0",
        "emojis": {"ok": "✅", "err": "❌", "warn": "⚠️", "retry": "🔁", "clock": "⏰"},
    },

    "rutas": {
        "url_login":  "https://www.sigges.cl/#/login",
        "url_home":   "https://www.sigges.cl/#/actualizaciones",
        "url_masivo": "https://www.sigges.cl/#/ingreso-de-prestaciones-otorgadas-integradas-masivas",
        "carpeta_logs": r"C:\Users\usuariohgf\OneDrive\Documentos\Trabajo Oficina\Revisiones, Falta Enviar, CM, Listos\Revisiones\Revision Tamizajes",
        "excel_entrada": r"C:\Users\usuariohgf\OneDrive\Documentos\Trabajo Oficina\Archivos\Tamizajes\Tamizajes.xlsx"
    },

    "selectores": {
        "navbar": "/html/body/div/main/div[2]/nav/div[2]",
        "codigo_input": "/html/body/div/main/div[3]/div[2]/div/div[4]/div[2]/div[1]/div[1]/div/div[1]/div/input",
        "buscar_codigo": "/html/body/div/main/div[3]/div[2]/div/div[4]/div[2]/div[1]/div[1]/div/div[2]/button",
        "codigo_primera_opcion": "/html/body/div/main/div[3]/div[2]/div/div[4]/div[2]/div[1]/div[2]/div/div/div[2]/div[1]",
        "fecha_input": "/html/body/div/main/div[3]/div[2]/div/div[4]/div[4]/div[1]/div/div[1]/div/input",
        "hora_input":  "/html/body/div/main/div[3]/div[2]/div/div[4]/div[4]/div[2]/div/div[2]/div/input",
        "unidad_select": "/html/body/div/main/div[3]/div[2]/div/div[4]/div[5]/div[1]/div/select",
        "unidad_option_2": "/html/body/div/main/div[3]/div[2]/div/div[4]/div[5]/div[1]/div/select/option[2]",
        "especialidad_select": "/html/body/div/main/div[3]/div[2]/div/div[4]/div[5]/div[2]/div/select",
        "especialidad_option_92": "/html/body/div/main/div[3]/div[2]/div/div[4]/div[5]/div[2]/div/select/option[92]",
        "rut_input": "/html/body/div/main/div[3]/div[2]/div/div[6]/div[1]/div[1]/div[2]/input",
        "buscar_run": "/html/body/div/main/div[3]/div[2]/div/div[6]/div[1]/div[1]/button[1]",
        "alerta_multi_1": "/html/body/div[1]/main/div[3]/div[2]/div/div[6]/div[1]/div/div/div/p",
        "alerta_multi_2": "//div[@data-color='rojo']/p[contains(.,'El RUN especificado tiene más de un paciente asociado')]",
        "caso_select_orig": "/html/body/div/main/div[3]/div[2]/div/div[6]/div[2]/div[2]/div/select",
        "orden_atencion_orig": "/html/body/div/main/div[3]/div[2]/div/div[6]/div[3]/div/div[3]/div/label/div",
        "folio_select_orig": "/html/body/div/main/div[3]/div[2]/div/div[6]/div[3]/div/div[5]/div/select",
        "agregar_paciente_orig": "/html/body/div/main/div[3]/div[2]/div/div[6]/div[5]/div[1]/button",
        "caso_select_patch": "/html/body/div[1]/main/div[3]/div[2]/div/div[6]/div[3]/div[2]/div/select",
        "orden_atencion_patch": "/html/body/div/main/div[3]/div[2]/div/div[6]/div[4]/div/div[3]/div/label/div",
        "folio_select_patch": "/html/body/div/main/div[3]/div[2]/div/div[6]/div[4]/div/div[5]/div/select",
        "agregar_paciente_patch": "/html/body/div/main/div[3]/div[2]/div/div[6]/div[6]/div[1]/button",
        "agregar_paciente_no_table": "/html/body/div/main/div[3]/div[2]/div/div[6]/div[4]/div[1]/button",
        "tbody_agregados": "/html/body/div/main/div[3]/div[2]/div/div[6]/div[4]/div/table/tbody",
        "btn_eliminar_paciente": [
            "/html/body/div/main/div[3]/div[2]/div/div[6]/div[5]/div[2]/button",
            "#root > main > div.content > div.contBody > div > div:nth-child(9) > div:nth-child(5) > div:nth-child(2) > button"
        ],
        "btn_grabar": "/html/body/div/main/div[3]/div[2]/div/div[7]/div/button[1]",
        "tbody_grabados": "/html/body/div/main/div[3]/div[2]/div/div[1]/div/table/tbody",
        "btn_volver": "/html/body/div/main/div[3]/div[2]/div/div[2]/button",
        "spinner": [
            (By.CSS_SELECTOR, "#root > dialog.loading > div"),
            (By.XPATH, "//*[@id='root']/dialog[2]/div"),
            (By.XPATH, "/html/body/div/dialog[2]/div"),
            (By.CSS_SELECTOR, "div.circulo"),
        ],
    },

    "tabla_masiva": {
        "td_checkbox": 1,
        "td_rut": 4,
        "td_fecha": 11,
        "td_nombre": 5,
    },

    "prestacion": {
        "codigo": "0801101",
        "hora_fija": "14:00",
        "caso_prioritario": "Cáncer Cervicouterino Segmento Proceso de Diagnóstico",
        "caso_sospecha": "Caso en Sospecha",
        "caso_cerrado": "Caso Cerrado",
    },

    "webdriver": {
        "driver_path": r"C:\Windows\System32\msedgedriver.exe",
        "debugger_address": "localhost:9222",
        "implicit_wait": 0,
        "page_load_timeout": 60,
        "script_timeout": 60,
    },

    "timeouts": {
        "clickable_default": 10,
        "presence_default": 10,
        "spinner_max": 180,
        "quick_verify_timeout": 5,
        "post_table_timeout": 15,
        "healthcheck_timeout": 5,
    },

    "esperas": {
        "sleep_min": 1,
        "post_click": 1,
        "post_escribir": 1,
        "post_clear": 1,
        "after_navegar_masivo": 1,
        "poll_spinner": 1,
        "espera_eliminar_dup": 3.0,
    },

    "reintentos": {
        "redirigir_masivo_en_cada_reintento": True,
        "delays_por_intento": [2, 3, 5, 8, 15, 20, 30, 45],
        "max_intentos": 12,
    },

    "verificacion": {
        "checkpoint_every": 50,
        "lote_tamano": 100,
        "leer_top_pre": 100,
    },
}


# ╔═══════════════════════════════════════════════════════════════════╗
# ║ Utilidades puras (sin dependencia de estado)                      ║
# ╚═══════════════════════════════════════════════════════════════════╝
def normalizar_rut(s: str) -> str:
    """Normaliza RUT: solo dígitos y K mayúscula."""
    return re.sub(r"[^0-9kK]", "", str(s)).upper()


def normalizar_fecha(valor: Any) -> str:
    """
    Normaliza cualquier valor de fecha a formato DD-MM-YYYY.
    Acepta: datetime, string con varios formatos, o None.
    """
    if valor is None:
        return "SIN_FECHA"
    
    # Si es datetime, formatear directamente
    if hasattr(valor, "strftime"):
        return valor.strftime("%d-%m-%Y")
    
    s = str(valor).strip().replace("/", "-").replace(".", "-")
    
    # Intentar parsear formatos conocidos
    formatos = ("%d-%m-%Y", "%Y-%m-%d", "%d-%m-%y", "%d-%m-%Y %H:%M:%S")
    for fmt in formatos:
        try:
            return datetime.strptime(s, fmt).strftime("%d-%m-%Y")
        except ValueError:
            continue
    
    # Fallback: regex para extraer componentes
    match = re.search(r"(\d{1,2})-(\d{1,2})-(\d{2,4})", s)
    if match:
        dd, mm, yy = match.groups()
        yy = f"20{yy}" if len(yy) == 2 else yy
        try:
            return datetime(int(yy), int(mm), int(dd)).strftime("%d-%m-%Y")
        except ValueError:
            pass
    
    return s if s else "SIN_FECHA"


def short_error(e: Exception) -> str:
    """Resume un error en texto corto legible."""
    mensajes = {
        TimeoutException: "Timeout esperando elemento",
        ElementNotInteractableException: "Elemento no interactuable",
        StaleElementReferenceException: "Elemento obsoleto ('stale')",
        NoSuchElementException: "Elemento no encontrado",
    }
    for exc_type, msg in mensajes.items():
        if isinstance(e, exc_type):
            return msg
    msg = re.sub(r"\s+", " ", str(e)).strip()
    return (msg[:120] + "…") if len(msg) > 120 else msg


def asegurar_carpeta(path: str) -> bool:
    """Crea carpeta si no existe. Retorna True si exitoso."""
    try:
        os.makedirs(path, exist_ok=True)
        return True
    except OSError as e:
        print(f"{C_ERR}No pude crear carpeta {path}: {e}{Style.RESET_ALL}")
        return False


# ╔═══════════════════════════════════════════════════════════════════╗
# ║ Clase Principal: SiggesAutomation                                 ║
# ╚═══════════════════════════════════════════════════════════════════╝
class SiggesAutomation:
    """Encapsula toda la lógica de automatización SIGGES."""
    
    def __init__(self, config: Dict[str, Any]):
        self.config = config
        self.driver: Optional[webdriver.Edge] = None
        self.parche_activo: bool = False
        self.results_problemas: List[Dict[str, str]] = []
        self.eventos_sesion: List[str] = []
        self.session_ts: str = datetime.now().strftime("%Y-%m-%d_%H.%M.%S")
        self._emo = config["identidad"]["emojis"]
    
    def __enter__(self) -> "SiggesAutomation":
        self._iniciar_driver()
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb) -> None:
        self._guardar_problemas_excel()
        # No cerramos el driver porque es sesión de debug compartida
    
    def _validar_config(self) -> None:
        """Valida que la configuración esencial exista."""
        excel = self.config["rutas"]["excel_entrada"]
        if not os.path.exists(excel):
            raise FileNotFoundError(f"Excel no encontrado: {excel}")
        
        driver_path = self.config["webdriver"]["driver_path"]
        if not os.path.exists(driver_path):
            raise FileNotFoundError(f"EdgeDriver no encontrado: {driver_path}")
    
    def _iniciar_driver(self) -> None:
        """Inicia el WebDriver de Edge en modo debug."""
        cfg = self.config["webdriver"]
        opts = webdriver.EdgeOptions()
        opts.debugger_address = cfg["debugger_address"]
        service = Service(cfg["driver_path"])
        
        try:
            self.driver = webdriver.Edge(service=service, options=opts)
            self.driver.implicitly_wait(cfg["implicit_wait"])
            self.driver.set_page_load_timeout(cfg["page_load_timeout"])
            self.driver.set_script_timeout(cfg["script_timeout"])
        except WebDriverException as e:
            raise RuntimeError(f"No pude iniciar WebDriver: {e}") from e
    
    # ─────────────────────────────────────────────────────────────────
    # Helpers de espera y clicks
    # ─────────────────────────────────────────────────────────────────
    def _pausa(self, nombre: str = "sleep_min") -> None:
        """Pausa configurable desde PANEL['esperas']."""
        sec = self.config.get("esperas", {}).get(nombre, 0)
        if sec > 0:
            time.sleep(float(sec))
    
    def _click(self, el) -> None:
        """Click robusto con scroll y fallback JS."""
        try:
            self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
            el.click()
        except (ElementNotInteractableException, WebDriverException):
            self.driver.execute_script("arguments[0].click();", el)
        self._pausa("post_click")
    
    def _escribir(self, el, texto: str, clear: bool = True) -> None:
        """Escribe texto en elemento con clear opcional."""
        if clear:
            try:
                el.clear()
            except WebDriverException:
                pass
            self._pausa("post_clear")
        el.send_keys(texto)
        self._pausa("post_escribir")
    
    def _spinner_activo(self) -> bool:
        """Verifica si algún spinner está visible."""
        for by, sel in self.config["selectores"]["spinner"]:
            try:
                el = self.driver.find_element(by, sel)
                if el and el.is_displayed():
                    return True
            except (NoSuchElementException, StaleElementReferenceException):
                continue
        return False
    
    def _esperar_spinner_desaparezca(self, max_seg: Optional[int] = None) -> bool:
        """Espera activa hasta que el spinner desaparezca."""
        if max_seg is None:
            max_seg = self.config["timeouts"]["spinner_max"]
        
        t0 = time.time()
        while time.time() - t0 < max_seg:
            try:
                if not self._spinner_activo():
                    return True
            except WebDriverException:
                return True
            self._pausa("poll_spinner")
        
        raise TimeoutException(f"Spinner sigue activo después de {max_seg}s")
    
    def _smart_wait(self, condition: Callable, timeout: int, name: str):
        """Espera spinner + condición específica."""
        self._esperar_spinner_desaparezca()
        try:
            return WebDriverWait(self.driver, timeout).until(condition)
        except TimeoutException:
            raise TimeoutException(f"Timeout esperando '{name}'")
    
    def _esperar_clickable_xp(self, xp: str, timeout: Optional[int] = None):
        """Espera elemento clickable por XPATH."""
        timeout = timeout or self.config["timeouts"]["clickable_default"]
        return self._smart_wait(
            EC.element_to_be_clickable((By.XPATH, xp)),
            timeout, f"clickable: {xp[:40]}..."
        )
    
    def _esperar_presence_xp(self, xp: str, timeout: Optional[int] = None):
        """Espera presencia de elemento por XPATH."""
        timeout = timeout or self.config["timeouts"]["presence_default"]
        return self._smart_wait(
            EC.presence_of_element_located((By.XPATH, xp)),
            timeout, f"presence: {xp[:40]}..."
        )
    
    # ─────────────────────────────────────────────────────────────────
    # Navegación y sesión
    # ─────────────────────────────────────────────────────────────────
    def _en_pagina_login(self) -> bool:
        """Detecta si estamos en login (sesión cerrada)."""
        try:
            if "#/login" in self.driver.current_url.lower():
                return True
        except WebDriverException:
            pass
        try:
            WebDriverWait(self.driver, 1).until(
                EC.presence_of_element_located((By.XPATH, self.config["selectores"]["navbar"]))
            )
            return False
        except TimeoutException:
            return True
    
    def _navegar_a_masivo(self) -> None:
        """Navega a la pantalla de ingreso masivo."""
        self.driver.get(self.config["rutas"]["url_masivo"])
        self._pausa("after_navegar_masivo")
        try:
            self._esperar_clickable_xp(self.config["selectores"]["codigo_input"], timeout=15)
        except TimeoutException as e:
            print(f"{C_ERR}Página 'Masivo' no cargó. {short_error(e)}{Style.RESET_ALL}")
            if self._en_pagina_login():
                self._reloguear_y_ir_a_masivo()
            else:
                raise
    
    def _reloguear_y_ir_a_masivo(self, indice: Optional[int] = None) -> None:
        """Intenta re-login automático cuando la sesión expira."""
        print(f"{self._emo['warn']} Sesión cerrada. Reingresando…")
        try:
            self.driver.get(self.config["rutas"]["url_login"])
            
            # Intentar click en botón Ingresar
            try:
                btn = WebDriverWait(self.driver, 10).until(
                    EC.element_to_be_clickable(
                        (By.XPATH, "//*[@class='botonBase' and @data-color='verde']/p[contains(.,'Ingresar')]/..")
                    )
                )
                self._click(btn)
            except TimeoutException:
                pass
            
            # Seleccionar establecimiento
            try:
                header = WebDriverWait(self.driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//div[@class='filtroSelect__header leftCenter']"))
                )
                self._click(header)
                opt = self.driver.find_element(
                    By.XPATH, "//div[@class='filtroSelect__options']/div[contains(@class,'filtroSelect__option')]"
                )
                self._click(opt)
            except (TimeoutException, NoSuchElementException):
                pass
            
            # Seleccionar perfil INGRESO SIGGES
            try:
                items = self.driver.find_elements(By.XPATH, "//div[contains(@class,'boxItems__items')]/div")
                elegido = next((el for el in items if "INGRESO SIGGES" in (el.text or "")), None)
                elegido = elegido or (items[0] if items else None)
                if elegido:
                    self._click(elegido)
            except (NoSuchElementException, IndexError):
                pass
            
            # Click botón azul confirmar
            try:
                btn = self.driver.find_element(By.XPATH, "//button[@class='botonBase' and @data-color='azulP']")
                self._click(btn)
            except NoSuchElementException:
                pass
            
            self._navegar_a_masivo()
            if not self._en_pagina_login():
                print(f"{self._emo['ok']} Sesión reanudada.")
                self.eventos_sesion.append(f"[{datetime.now():%H:%M:%S}] Relogueo OK (paciente #{indice or '-'}).")
        
        except WebDriverException as e:
            print(f"{C_ERR}Error en relogueo: {short_error(e)}{Style.RESET_ALL}")
            self._health_check_to_log("reloguear", extra=short_error(e))
    
    def _cerrar_popup_si_aplica(self, timeout: float = 2.0) -> None:
        """Cierra popups de confirmación si aparecen."""
        try:
            btn = WebDriverWait(self.driver, timeout).until(
                EC.element_to_be_clickable(
                    (By.XPATH, "//dialog//button[contains(.,'Aceptar') or contains(.,'OK') or contains(.,'Sí')]")
                )
            )
            self._click(btn)
        except TimeoutException:
            pass
    
    # ─────────────────────────────────────────────────────────────────
    # Pasos de formulario
    # ─────────────────────────────────────────────────────────────────
    def _seleccionar_codigo(self) -> None:
        """Selecciona código de prestación."""
        S = self.config["selectores"]
        inp = self._esperar_clickable_xp(S["codigo_input"])
        self._escribir(inp, self.config["prestacion"]["codigo"])
        try:
            btn = self._esperar_clickable_xp(S["buscar_codigo"])
            self._click(btn)
        except TimeoutException:
            pass
        first = self._esperar_clickable_xp(S["codigo_primera_opcion"])
        self._click(first)
    
    def _set_fecha_y_hora(self, fecha_str: str) -> None:
        """Establece fecha y hora en formulario."""
        S = self.config["selectores"]
        f = self._esperar_clickable_xp(S["fecha_input"])
        self._escribir(f, normalizar_fecha(fecha_str))
        h = self._esperar_clickable_xp(S["hora_input"])
        self._escribir(h, self.config["prestacion"]["hora_fija"])
    
    def _seleccionar_unidad_y_especialidad(self) -> None:
        """Selecciona unidad y especialidad."""
        S = self.config["selectores"]
        un = self._esperar_clickable_xp(S["unidad_select"])
        self._click(un)
        opu = self._esperar_clickable_xp(S["unidad_option_2"])
        self._click(opu)
        es = self._esperar_clickable_xp(S["especialidad_select"])
        self._click(es)
        ope = self._esperar_clickable_xp(S["especialidad_option_92"])
        self._click(ope)
    
    def _alerta_multi_paciente_presente(self, timeout: float = 1.0) -> bool:
        """Detecta alerta de RUN con múltiples pacientes."""
        S = self.config["selectores"]
        xpaths = [S["alerta_multi_1"], S["alerta_multi_2"]]
        
        for xp in xpaths:
            try:
                el = WebDriverWait(self.driver, timeout).until(
                    EC.presence_of_element_located((By.XPATH, xp))
                )
                if el and "más de un paciente" in (el.text or ""):
                    return True
            except TimeoutException:
                continue
        
        # Búsqueda genérica
        try:
            els = self.driver.find_elements(By.XPATH, "//*[contains(text(),'más de un paciente')]")
            return bool(els)
        except NoSuchElementException:
            return False
    
    def _get_xpaths_para(self, nombre: str) -> List[str]:
        """Obtiene XPATHs según parche activo."""
        S = self.config["selectores"]
        mapping = {
            "caso": ("caso_select_orig", "caso_select_patch"),
            "orden": ("orden_atencion_orig", "orden_atencion_patch"),
            "folio": ("folio_select_orig", "folio_select_patch"),
            "agregar": ("agregar_paciente_orig", "agregar_paciente_patch"),
        }
        if nombre not in mapping:
            raise ValueError(f"Nombre inválido: {nombre}")
        
        orig_key, patch_key = mapping[nombre]
        if self.parche_activo:
            return [S[patch_key], S[orig_key]]
        return [S[orig_key], S[patch_key]]
    
    def _ingresar_run_y_seleccionar_caso(self, rut_norm: str) -> None:
        """Ingresa RUT y selecciona caso GES."""
        S = self.config["selectores"]
        P = self.config["prestacion"]
        
        inp = self._esperar_clickable_xp(S["rut_input"])
        self._escribir(inp, rut_norm)
        btn = self._esperar_clickable_xp(S["buscar_run"])
        self._click(btn)
        self._pausa("sleep_min")
        
        self.parche_activo = self._alerta_multi_paciente_presente(timeout=1.0)
        if self.parche_activo:
            print("🟥 Aviso: RUN con múltiples pacientes. Activando PARCHE.")
        
        caso_xp = self._get_xpaths_para("caso")[0]
        caso = self._esperar_clickable_xp(caso_xp)
        
        # Buscar caso prioritario
        encontrado = False
        for opt in caso.find_elements(By.TAG_NAME, "option"):
            texto = opt.text or ""
            if P["caso_prioritario"] in texto and P["caso_sospecha"] in texto:
                self._click(opt)
                encontrado = True
                break
        
        # Fallback a caso cerrado
        if not encontrado:
            for opt in caso.find_elements(By.TAG_NAME, "option"):
                texto = opt.text or ""
                if P["caso_prioritario"] in texto and P["caso_cerrado"] in texto:
                    self._click(opt)
                    encontrado = True
                    break
        
        # Último fallback: cualquier opción válida
        if not encontrado:
            for opt in caso.find_elements(By.TAG_NAME, "option"):
                if not opt.get_attribute("disabled"):
                    val = (opt.get_attribute("value") or "").strip()
                    if val and val != "0":
                        self._click(opt)
                        break
    
    def _seleccionar_orden_y_folio(self) -> None:
        """Selecciona orden de atención y folio."""
        orden_xp = self._get_xpaths_para("orden")[0]
        orden = self._esperar_clickable_xp(orden_xp)
        self._click(orden)
        
        folio_xp = self._get_xpaths_para("folio")[0]
        folio = self._esperar_clickable_xp(folio_xp)
        self._click(folio)
        
        # Esperar que carguen opciones
        WebDriverWait(self.driver, 10).until(
            lambda d: len(d.find_element(By.XPATH, folio_xp).find_elements(By.TAG_NAME, "option")) >= 1
        )
        
        # Seleccionar primer folio válido
        for opt in folio.find_elements(By.TAG_NAME, "option"):
            if opt.get_attribute("disabled"):
                continue
            val = (opt.get_attribute("value") or "").strip()
            if val and val != "0":
                self._click(opt)
                return
        
        raise RuntimeError("No hay folios válidos disponibles")
    
    def _agregar_paciente(self) -> None:
        """Hace click en botón Agregar paciente."""
        S = self.config["selectores"]
        btn = None
        
        # Intentar encontrar tabla PRE y usar botón correspondiente
        try:
            self.driver.find_element(By.XPATH, S["tbody_agregados"])
            btn = self._esperar_clickable_xp(self._get_xpaths_para("agregar")[0])
        except NoSuchElementException:
            pass
        
        # Fallbacks
        if not btn:
            fallbacks = [
                S["agregar_paciente_no_table"],
                "//button[contains(.,'Agregar') or .//p[contains(.,'Agregar')]]"
            ]
            for xp in fallbacks:
                try:
                    btn = self._esperar_clickable_xp(xp)
                    break
                except TimeoutException:
                    continue
        
        if not btn:
            raise RuntimeError("No se encontró botón Agregar")
        
        self._click(btn)
        self.parche_activo = False
    
    # ─────────────────────────────────────────────────────────────────
    # Tabla PRE / POST
    # ─────────────────────────────────────────────────────────────────
    def _leer_filas_agregados(self, limit: Optional[int] = None) -> List[Dict[str, Any]]:
        """Lee filas de la tabla PRE-grabado."""
        S, T = self.config["selectores"], self.config["tabla_masiva"]
        filas: List[Dict[str, Any]] = []
        
        try:
            tbody = self.driver.find_element(By.XPATH, S["tbody_agregados"])
        except NoSuchElementException:
            return filas
        
        trs = tbody.find_elements(By.XPATH, "./tr")
        if limit:
            trs = trs[:limit]
        
        for tr in trs:
            rut = self._leer_td_text(tr, T["td_rut"])
            fecha = self._leer_td_text(tr, T["td_fecha"])
            nombre = self._leer_td_text(tr, T["td_nombre"])
            filas.append({
                "tr": tr,
                "rut": rut,
                "rut_norm": normalizar_rut(rut),
                "fecha": fecha,
                "fecha_norm": normalizar_fecha(fecha),
                "nombre": nombre
            })
        return filas
    
    @staticmethod
    def _leer_td_text(tr, td_idx: int) -> str:
        """Lee texto de una celda TD."""
        try:
            td = tr.find_element(By.XPATH, f"./td[{td_idx}]")
            return (td.text or "").strip()
        except NoSuchElementException:
            return ""
    
    def _quick_verify_after_add(self, rut_norm: str) -> bool:
        """Verificación rápida de que el paciente fue agregado."""
        timeout = self.config["timeouts"]["quick_verify_timeout"]
        S = self.config["selectores"]
        
        try:
            filas_antes = len(self.driver.find_elements(By.XPATH, f"{S['tbody_agregados']}/tr"))
        except NoSuchElementException:
            filas_antes = 0
        
        try:
            WebDriverWait(self.driver, timeout).until(
                lambda d: (
                    len(d.find_elements(By.XPATH, f"{S['tbody_agregados']}/tr")) > filas_antes and
                    any(r["rut_norm"] == rut_norm for r in self._leer_filas_agregados(limit=30))
                )
            )
            return True
        except TimeoutException:
            return any(r["rut_norm"] == rut_norm for r in self._leer_filas_agregados(limit=30))
    
    def _detectar_duplicados(self, rows: List[Dict]) -> Dict[Tuple[str, str], List[int]]:
        """Detecta duplicados por (RUT, fecha)."""
        m: Dict[Tuple[str, str], List[int]] = {}
        for i, r in enumerate(rows):
            key = (r["rut_norm"], r["fecha_norm"])
            m.setdefault(key, []).append(i)
        return {k: v for k, v in m.items() if len(v) > 1}
    
    def _eliminar_duplicados_uno_a_uno(self) -> List[Dict[str, str]]:
        """Elimina duplicados de la tabla PRE."""
        eliminados: List[Dict[str, str]] = []
        espera = self.config["esperas"]["espera_eliminar_dup"]
        T = self.config["tabla_masiva"]
        
        while True:
            rows = self._leer_filas_agregados()
            dups = self._detectar_duplicados(rows)
            if not dups:
                break
            
            hubo_eliminado = False
            for key, idxs in dups.items():
                for i in idxs[1:2]:  # Solo uno por iteración
                    fila = rows[i]
                    tr = fila["tr"]
                    
                    # Marcar checkbox
                    try:
                        cb = tr.find_element(By.XPATH, f"./td[{T['td_checkbox']}]/input[@type='checkbox']")
                    except NoSuchElementException:
                        try:
                            cb = tr.find_element(By.CSS_SELECTOR, "td:first-child input[type='checkbox']")
                        except NoSuchElementException:
                            continue
                    
                    self._click(cb)
                    
                    # Click botón eliminar
                    if self._click_boton_eliminar():
                        eliminados.append({
                            "Rut": fila["rut"],
                            "Nombre": fila["nombre"],
                            "Fecha": fila["fecha_norm"]
                        })
                        time.sleep(espera)
                        hubo_eliminado = True
                        break
            
            if not hubo_eliminado:
                break
        
        return eliminados
    
    def _click_boton_eliminar(self) -> bool:
        """Click en botón eliminar paciente."""
        for path in self.config["selectores"]["btn_eliminar_paciente"]:
            try:
                by = By.XPATH if path.startswith("/") else By.CSS_SELECTOR
                btn = WebDriverWait(self.driver, 5).until(
                    EC.element_to_be_clickable((by, path))
                )
                self._click(btn)
                return True
            except TimeoutException:
                continue
        return False
    
    def _obtener_ruts_post_grabar(self) -> List[str]:
        """Lee RUTs de tabla POST-grabado."""
        S = self.config["selectores"]
        ruts: List[str] = []
        
        try:
            tbody = WebDriverWait(self.driver, self.config["timeouts"]["post_table_timeout"]).until(
                EC.presence_of_element_located((By.XPATH, S["tbody_grabados"]))
            )
        except TimeoutException:
            print(f"{C_ERR}No apareció tabla 'Grabados'.{Style.RESET_ALL}")
            return ruts
        
        for tr in tbody.find_elements(By.XPATH, "./tr"):
            try:
                td1 = tr.find_element(By.XPATH, "./td[1]").text.strip()
                ruts.append(normalizar_rut(td1))
            except NoSuchElementException:
                continue
        
        return ruts
    
    # ─────────────────────────────────────────────────────────────────
    # Logging
    # ─────────────────────────────────────────────────────────────────
    def _log_problema(self, rut: str, nombre: str, fecha: str, obs: str, etapa: str) -> None:
        """Registra problema para reporte final."""
        self.results_problemas.append({
            "Rut": rut, "Nombre": nombre, "Fecha": fecha,
            "Observación": obs, "Etapa": etapa
        })
    
    def _health_check_to_log(self, contexto: str, extra: Optional[str] = None) -> None:
        """Escribe archivo de diagnóstico."""
        try:
            asegurar_carpeta(self.config["rutas"]["carpeta_logs"])
            fname = f"HEALTH_{datetime.now():%Y-%m-%d_%H.%M.%S}.txt"
            fpath = os.path.join(self.config["rutas"]["carpeta_logs"], fname)
            
            with open(fpath, "w", encoding="utf-8") as f:
                f.write(f"===== HEALTH CHECK =====\n")
                f.write(f"Fecha/Hora: {datetime.now():%d-%m-%Y %H:%M:%S}\n")
                f.write(f"Contexto: {contexto}\n")
                if extra:
                    f.write(f"Extra: {extra}\n")
                f.write(f"URL: {getattr(self.driver, 'current_url', '<desconocido>')}\n")
                f.write(f"Spinner: {self._spinner_activo()}\n")
        except OSError:
            pass
    
    def _guardar_problemas_excel(self, path_out: Optional[str] = None) -> None:
        """Guarda reporte de problemas en Excel."""
        if not self.results_problemas:
            return
        
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Problemas"
            
            # Headers
            headers = ["Rut", "Nombre", "Fecha", "Observación", "Etapa"]
            ws.append(headers)
            
            # Data
            for p in self.results_problemas:
                ws.append([p.get(h, "") for h in headers])
            
            # Guardar
            if not path_out:
                path_out = os.path.join(
                    self.config["rutas"]["carpeta_logs"],
                    f"Problemas_{self.session_ts}.xlsx"
                )
            wb.save(path_out)
            print(f"{C_OK}Problemas guardados en: {path_out}{Style.RESET_ALL}")
        except OSError as e:
            print(f"{C_ERR}No pude guardar problemas: {e}{Style.RESET_ALL}")
    
    def _escribir_log_lote(self, n_lote: int, pacientes: List, eliminados: List,
                           faltantes: List, extras: List, dups: Dict, ruts_post: List) -> None:
        """Escribe log detallado del lote."""
        try:
            fname = f"Lote_{n_lote}_{self.session_ts}.txt"
            fpath = os.path.join(self.config["rutas"]["carpeta_logs"], fname)
            
            with open(fpath, "w", encoding="utf-8") as f:
                f.write(f"===== LOTE #{n_lote} =====\n")
                f.write(f"Fecha: {datetime.now():%d-%m-%Y %H:%M:%S}\n")
                f.write(f"Pacientes Excel: {len(pacientes)}\n")
                f.write(f"Eliminados duplicados: {len(eliminados)}\n")
                f.write(f"Faltantes PRE: {len(faltantes)}\n")
                f.write(f"Extras PRE: {len(extras)}\n")
                f.write(f"Duplicados PRE: {len(dups)}\n")
                f.write(f"RUTs POST: {len(ruts_post)}\n")
        except OSError:
            pass
    
    # ─────────────────────────────────────────────────────────────────
    # Procesamiento
    # ─────────────────────────────────────────────────────────────────
    def _delay_por_intento(self, idx: int) -> int:
        """Calcula delay según número de intento."""
        arr = self.config["reintentos"]["delays_por_intento"]
        return arr[min(idx - 1, len(arr) - 1)]
    
    def _procesar_paciente(self, i: int, total: int, fila: Tuple) -> None:
        """Procesa un paciente individual con reintentos."""
        rut_raw = str(fila[0])
        rut_norm = normalizar_rut(rut_raw)
        nombre = fila[1] if fila[1] else "SIN_NOMBRE"
        fecha_val = normalizar_fecha(fila[2])
        
        print(f"{C_CNT}[{i}/{total}] {datetime.now():%H:%M:%S}{Style.RESET_ALL} | "
              f"{C_NAME}{nombre}{Style.RESET_ALL} | "
              f"{C_RUT}{rut_raw}{Style.RESET_ALL} | "
              f"{C_DATE}{fecha_val}{Style.RESET_ALL}")
        
        max_intentos = self.config["reintentos"]["max_intentos"]
        
        for intento in range(1, max_intentos + 1):
            try:
                if self.config["reintentos"]["redirigir_masivo_en_cada_reintento"]:
                    self._navegar_a_masivo()
                    if self._en_pagina_login():
                        self._reloguear_y_ir_a_masivo(indice=i)
                
                self._seleccionar_codigo()
                self._set_fecha_y_hora(fecha_val)
                self._seleccionar_unidad_y_especialidad()
                self._ingresar_run_y_seleccionar_caso(rut_norm)
                self._seleccionar_orden_y_folio()
                self._agregar_paciente()
                
                if not self._quick_verify_after_add(rut_norm):
                    self._log_problema(rut_raw, nombre, fecha_val,
                                       "QuickCheck: no encontrado en PRE", "quick")
                    print(f"{C_DIM}QuickCheck: no visto aún.{Style.RESET_ALL}")
                return
            
            except Exception as e:
                self._cerrar_popup_si_aplica(2.0)
                
                if self._en_pagina_login():
                    self._reloguear_y_ir_a_masivo(indice=i)
                
                if not isinstance(e, TimeoutException) and self._spinner_activo():
                    print(f"{self._emo['warn']} Esperando spinner pegado…")
                    try:
                        self._esperar_spinner_desaparezca()
                    except TimeoutException:
                        print(f"{C_ERR}Spinner se quedó pegado.{Style.RESET_ALL}")
                
                if intento < max_intentos:
                    d = self._delay_por_intento(intento)
                    print(f"{C_ERR}{short_error(e)}{Style.RESET_ALL} "
                          f"{self._emo['retry']} {intento}/{max_intentos} · {d}s")
                    self._health_check_to_log("procesar_paciente", extra=short_error(e))
                    time.sleep(d)
                else:
                    self._log_problema(rut_raw, nombre, fecha_val,
                                       f"Fallo tras {max_intentos} intentos: {short_error(e)}", "final")
                    print(f"{C_ERR}Fallo definitivo para este paciente.{Style.RESET_ALL}")
                    self._health_check_to_log("procesar_final", extra=short_error(e))
                    return
    
    def _checkpoint_50(self, i_procesados: int) -> None:
        """Checkpoint cada 50 pacientes."""
        top = self._leer_filas_agregados(limit=50)
        dups = self._detectar_duplicados(top)
        if dups:
            print(f"{self._emo['warn']} Checkpoint {i_procesados}: {len(dups)} claves duplicadas")
        else:
            print(f"🧹 Checkpoint {i_procesados}: Sin duplicados")
    
    def _procesar_lote(self, n_lote: int, pacientes: List[Tuple]) -> None:
        """Procesa un lote completo: elimina dups, graba, verifica."""
        # Normalizar datos del lote
        excel_norm = [(normalizar_rut(str(r[0])), r[1] or "SIN_NOMBRE", normalizar_fecha(r[2]))
                      for r in pacientes]
        
        eliminados = self._eliminar_duplicados_uno_a_uno()
        filas_pre = self._leer_filas_agregados(limit=self.config["verificacion"]["leer_top_pre"])
        
        # Comparar PRE con Excel
        set_excel = {(r[0], r[2]) for r in excel_norm}
        set_pre = {(f["rut_norm"], f["fecha_norm"]) for f in filas_pre}
        faltantes = list(set_excel - set_pre)
        extras = list(set_pre - set_excel)
        dups = self._detectar_duplicados(filas_pre)
        
        print(f"🔎 Lote #{n_lote} PRE — faltantes={len(faltantes)}, extras={len(extras)}, dups={len(dups)}")
        
        # Grabar
        btn_grabar = self._esperar_clickable_xp(self.config["selectores"]["btn_grabar"])
        self._click(btn_grabar)
        
        try:
            self._esperar_spinner_desaparezca()
        except TimeoutException as e:
            print(f"{Fore.YELLOW}Timeout spinner post-grabar: {e}{Style.RESET_ALL}")
        
        # Verificar POST
        ruts_post = self._obtener_ruts_post_grabar()
        set_post = set(ruts_post)
        set_excel_rut = {r[0] for r in excel_norm}
        
        print(f"📊 Lote #{n_lote} POST — registrados={len(set_post)} | "
              f"faltan={len(set_excel_rut - set_post)} | extras={len(set_post - set_excel_rut)}")
        
        # Volver a pantalla masivo
        btn_volver = self._esperar_clickable_xp(self.config["selectores"]["btn_volver"])
        self._click(btn_volver)
        self._esperar_clickable_xp(self.config["selectores"]["codigo_input"], timeout=15)
        
        # Log
        self._escribir_log_lote(n_lote, excel_norm, eliminados, faltantes, extras, dups, ruts_post)
    
    def run(self) -> None:
        """Ejecuta el proceso completo."""
        print(f"{self._emo['clock']} Iniciando v{self.config['identidad']['version']} "
              f"por {self.config['identidad']['autor']} {self._emo['clock']}")
        
        self._validar_config()
        asegurar_carpeta(self.config["rutas"]["carpeta_logs"])
        self._navegar_a_masivo()
        
        # Cargar Excel
        wb = openpyxl.load_workbook(self.config["rutas"]["excel_entrada"])
        ws = wb.active
        filas = list(ws.iter_rows(min_row=2, values_only=True))
        valid = [r for r in filas if r and r[0]]
        total = len(valid)
        
        if total == 0:
            print(f"{C_ERR}Excel sin pacientes.{Style.RESET_ALL}")
            return
        
        tamanio = self.config["verificacion"]["lote_tamano"]
        lote_actual = 1
        
        for i, fila in enumerate(valid, start=1):
            self._procesar_paciente(i, total, fila)
            
            if i % self.config["verificacion"]["checkpoint_every"] == 0:
                self._checkpoint_50(i)
            
            if (i % tamanio == 0) or (i == total):
                ini = i - ((i - 1) % tamanio) - 1
                pacientes_lote = valid[ini:i]
                print("— — — — — — — — — — — — — — — —")
                print(f"🏁 Lote #{lote_actual} ({len(pacientes_lote)} pacientes)…")
                try:
                    self._procesar_lote(lote_actual, pacientes_lote)
                except Exception as e:
                    print(f"{C_ERR}Error lote #{lote_actual}: {short_error(e)}{Style.RESET_ALL}")
                    self._health_check_to_log("procesar_lote", extra=short_error(e))
                    try:
                        self._navegar_a_masivo()
                    except WebDriverException:
                        print(f"{C_ERR}Fallo catastrófico.{Style.RESET_ALL}")
                        raise
                lote_actual += 1
        
        print(f"{C_OK}Proceso completado. Pacientes: {total}. Lotes: {lote_actual - 1}.{Style.RESET_ALL}")


# ╔═══════════════════════════════════════════════════════════════════╗
# ║ Entry Point                                                       ║
# ╚═══════════════════════════════════════════════════════════════════╝
def main():
    try:
        with SiggesAutomation(PANEL) as automation:
            automation.run()
    except FileNotFoundError as e:
        print(f"{C_ERR}Archivo no encontrado: {e}{Style.RESET_ALL}")
        sys.exit(1)
    except RuntimeError as e:
        print(f"{C_ERR}Error de ejecución: {e}{Style.RESET_ALL}")
        sys.exit(1)
    except KeyboardInterrupt:
        print(f"\n{C_DIM}Interrumpido por usuario.{Style.RESET_ALL}")
        sys.exit(0)


if __name__ == "__main__":
    main()
